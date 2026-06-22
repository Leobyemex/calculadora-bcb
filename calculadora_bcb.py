#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calculadora do Cidadão - Correção de Valores (v2.1)
====================================================
Reprodução não-oficial em Python/Tkinter da Calculadora do BCB,
estendida com modo Lote, Demonstrativo Previdenciário (FUNFIN),
exportação XLSX estilizada e geração de PDF.

- Dados em tempo real via API pública do BCB (SGS).
- Sem dependências externas obrigatórias (apenas stdlib).
- XLSX requer openpyxl (opcional). PDF requer reportlab (opcional).
"""

import csv
import json
import os
import re
import threading
import tkinter as tk
import urllib.error
import urllib.request
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal, getcontext, ROUND_HALF_UP
from tkinter import filedialog, messagebox, ttk

# precisão decimal generosa para os cálculos
getcontext().prec = 30

# openpyxl é opcional (apenas para XLSX)
XLSX_IMPORT_ERROR = None
try:
    import openpyxl
    from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side,
                                 NamedStyle)
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except Exception as _e:
    HAS_XLSX = False
    XLSX_IMPORT_ERROR = f"{type(_e).__name__}: {_e}"

# reportlab é opcional (apenas para PDF)
PDF_IMPORT_ERROR = None
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape, portrait
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak, Image)
    from reportlab.platypus.flowables import KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas as rl_canvas
    HAS_PDF = True
except Exception as _e:
    HAS_PDF = False
    PDF_IMPORT_ERROR = f"{type(_e).__name__}: {_e}"

# Log de diagnóstico ao lado do executável (ou script)
import sys as _sys
try:
    if getattr(_sys, "frozen", False):
        _log_dir = os.path.dirname(_sys.executable)
    else:
        _log_dir = os.path.dirname(os.path.abspath(__file__))
    _log_path = os.path.join(_log_dir, "calculadora_log.txt")
    with open(_log_path, "w", encoding="utf-8") as _lf:
        _lf.write(f"Python: {_sys.version}\n")
        _lf.write(f"Frozen (exe): {getattr(_sys, 'frozen', False)}\n")
        _lf.write(f"Executable: {_sys.executable}\n")
        _lf.write(f"HAS_XLSX: {HAS_XLSX}\n")
        if XLSX_IMPORT_ERROR:
            _lf.write(f"XLSX_IMPORT_ERROR: {XLSX_IMPORT_ERROR}\n")
        _lf.write(f"HAS_PDF: {HAS_PDF}\n")
        if PDF_IMPORT_ERROR:
            _lf.write(f"PDF_IMPORT_ERROR: {PDF_IMPORT_ERROR}\n")
except Exception:
    pass


# ===================== Configurações ===================== #

APP_TITLE = "Calculadora do Cidadão — Correção de Valores"
APP_VERSION  = "2.9.29"
GITHUB_REPO  = "Leobyemex/calculadora-bcb"

INDICES = {
    "IGP-M":  {"serie": 189,   "name": "IGP-M (FGV)",       "min": (1989, 6)},
    "IGP-DI": {"serie": 190,   "name": "IGP-DI (FGV)",      "min": (1944, 2)},
    "INPC":   {"serie": 188,   "name": "INPC (IBGE)",       "min": (1979, 4)},
    "IPCA":   {"serie": 433,   "name": "IPCA (IBGE)",       "min": (1980, 1)},
    "IPCA-E": {"serie": 10764, "name": "IPCA-E (IBGE)",     "min": (1992, 1)},
    "IPC-BR": {"serie": 191,   "name": "IPC-BRASIL (FGV)",  "min": (1990, 1)},
    "IPC-SP": {"serie": 193,   "name": "IPC-SP (FIPE)",     "min": (1942, 11)},
    "SELIC":  {"serie": 11,    "name": "Selic (BCB)",       "min": (1986, 6)},
}

# Paleta institucional BCB
COLOR_BCB_BLUE = "#003366"
COLOR_BCB_BLUE_DARK = "#001f3f"
COLOR_BCB_BLUE_LIGHT = "#0055a5"
COLOR_BG = "#e8eaed"
COLOR_PANEL = "#ffffff"
COLOR_HEADER_BG = "#f3f3f3"
COLOR_RESULT_BG = "#f4f8fc"
COLOR_RESULT_OK = "#006400"
COLOR_ERROR  = "#c0392b"
COLOR_VERDE  = "#1a7a1a"   # verde para valores que somam ao total
COLOR_TEXT = "#222222"
COLOR_SUBTLE = "#666666"
COLOR_TABLE_HEADER = "#003366"
COLOR_TABLE_ALT = "#f0f4f8"


# ===================== Auto-update helpers =========================== #

def _versao_maior(nova: str, atual: str) -> bool:
    """Retorna True se `nova` for estritamente maior que `atual`.
    Compara segmentos inteiros separados por ponto (ex: "2.9.20" > "2.9.19").
    """
    try:
        p_nova  = [int(x) for x in nova.strip().lstrip("v").split(".")]
        p_atual = [int(x) for x in atual.strip().lstrip("v").split(".")]
        # Padeia o mais curto com zeros
        n = max(len(p_nova), len(p_atual))
        p_nova  += [0] * (n - len(p_nova))
        p_atual += [0] * (n - len(p_atual))
        return p_nova > p_atual
    except Exception:
        return False


# ===================== Helpers de parsing/format ===================== #

def parse_month_year(s):
    """MM/AAAA -> (month, year) ou None"""
    if not s:
        return None
    m = re.match(r"^\s*(\d{1,2})/(\d{4})\s*$", s)
    if not m:
        return None
    mes, ano = int(m.group(1)), int(m.group(2))
    if not (1 <= mes <= 12) or not (1900 <= ano <= 2100):
        return None
    return mes, ano


def parse_date_br(s):
    """DD/MM/AAAA -> date ou None"""
    if not s:
        return None
    m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$", s)
    if not m:
        return None
    d, mes, ano = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(ano, mes, d)
    except ValueError:
        return None


def parse_valor_br(s):
    """1.234,56 -> Decimal('1234.56'); '' -> None; inválido -> ValueError"""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    clean = s.replace(" ", "").replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return Decimal(clean)
    except Exception as e:
        raise ValueError(f"Valor inválido: {s}") from e


def fmt_brl(v):
    if v is None:
        return ""
    s = f"{float(v):,.2f}"
    return s.replace(",", "_").replace(".", ",").replace("_", ".")


def fmt_fator(v):
    if v is None:
        return ""
    s = f"{float(v):,.7f}"
    return s.replace(",", "_").replace(".", ",").replace("_", ".")


def fmt_percent(v):
    if v is None:
        return ""
    s = f"{float(v):,.4f}"
    return s.replace(",", "_").replace(".", ",").replace("_", ".")


def fmt_competencia(mes, ano):
    """(3, 2024) -> 'Março/2024'"""
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    return f"{meses[mes-1]}/{ano}"


# ===================== API BCB (com cache) ===================== #

_api_cache = {}


def _fetch_serie_raw(serie, data_inicial, data_final, timeout=30):
    """Chama a API SGS do BCB. Retorna lista [{data, valor}]."""
    url = (
        f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{serie}/dados"
        f"?formato=json&dataInicial={data_inicial}&dataFinal={data_final}"
    )
    req = urllib.request.Request(
        url, headers={"Accept": "application/json",
                      "User-Agent": "CalculadoraCidadao/2.0"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read().decode("utf-8", errors="replace")
    data = json.loads(body)
    return data if isinstance(data, list) else []


def fetch_serie_em_lotes(serie, data_inicial_str, data_final_str, timeout=30,
                        progress_cb=None):
    """API limita a ~10 anos; pagina em lotes. data_*_str em DD/MM/AAAA."""
    cache_key = (serie, data_inicial_str, data_final_str)
    if cache_key in _api_cache:
        return _api_cache[cache_key]

    di = datetime.strptime(data_inicial_str, "%d/%m/%Y").date()
    df = datetime.strptime(data_final_str, "%d/%m/%Y").date()

    todos = []
    cur = di
    while cur <= df:
        try:
            end = date(cur.year + 9, cur.month, min(cur.day, 28))
        except ValueError:
            end = date(cur.year + 9, cur.month, 28)
        if end > df:
            end = df

        if progress_cb:
            progress_cb(f"Consultando BCB: {cur.strftime('%m/%Y')} → {end.strftime('%m/%Y')}")

        lote = _fetch_serie_raw(
            serie,
            cur.strftime("%d/%m/%Y"),
            end.strftime("%d/%m/%Y"),
            timeout=timeout,
        )
        if lote:
            todos.extend(lote)

        cur = end + timedelta(days=1)

    # dedup preservando ordem
    seen = set()
    out = []
    for d in todos:
        key = d.get("data")
        if key in seen:
            continue
        seen.add(key)
        out.append(d)

    _api_cache[cache_key] = out
    return out


def limpar_cache_api():
    _api_cache.clear()


# ===================== Cálculos básicos ===================== #

def calcular_indice(indice_key, ini, fim, valor, progress_cb=None):
    """
    ini, fim = (mes, ano)
    valor = Decimal ou None
    Retorna dict com resultado.
    """
    cfg = INDICES[indice_key]
    ini_mes, ini_ano = ini
    fim_mes, fim_ano = fim

    min_ano, min_mes = cfg["min"]
    if (ini_ano, ini_mes) < (min_ano, min_mes):
        raise ValueError(
            f"O índice {cfg['name']} está disponível a partir de "
            f"{min_mes:02d}/{min_ano}."
        )
    if (ini_ano, ini_mes) > (fim_ano, fim_mes):
        raise ValueError("Data inicial deve ser anterior ou igual à data final.")

    last_day = monthrange(fim_ano, fim_mes)[1]
    di_str = f"01/{ini_mes:02d}/{ini_ano}"
    df_str = f"{last_day:02d}/{fim_mes:02d}/{fim_ano}"

    dados_brutos = fetch_serie_em_lotes(cfg["serie"], di_str, df_str,
                                        progress_cb=progress_cb)
    if not dados_brutos:
        raise ValueError("A API do BCB não retornou dados para o período.")

    ini_num = ini_ano * 12 + ini_mes
    fim_num = fim_ano * 12 + fim_mes

    dados_filtrados = []
    for d in dados_brutos:
        try:
            _, mes_s, ano_s = d["data"].split("/")
            num = int(ano_s) * 12 + int(mes_s)
            if ini_num <= num <= fim_num:
                dados_filtrados.append(d)
        except (KeyError, ValueError):
            continue

    if not dados_filtrados:
        raise ValueError("Não há dados publicados para o intervalo informado.")

    fator = Decimal("1")
    for d in dados_filtrados:
        v = Decimal(str(d["valor"]))
        fator *= (Decimal("1") + v / Decimal("100"))

    variacao = (fator - Decimal("1")) * Decimal("100")
    valor_corrigido = (valor * fator) if valor is not None else None

    return {
        "indice_key": indice_key,
        "indice_nome": cfg["name"],
        "periodo": f"{ini_mes:02d}/{ini_ano} a {fim_mes:02d}/{fim_ano}",
        "meses": len(dados_filtrados),
        "fator": fator,
        "variacao": variacao,
        "valor_informado": valor,
        "valor_corrigido": valor_corrigido,
    }


def calcular_selic(ini, fim, valor, progress_cb=None):
    """ini, fim = date. valor = Decimal ou None."""
    if ini > fim:
        raise ValueError("Data inicial deve ser anterior ou igual à data final.")
    if ini < date(1986, 6, 4):
        raise ValueError("Para a Selic, informe períodos a partir de 04/06/1986.")

    di_str = ini.strftime("%d/%m/%Y")
    df_str = fim.strftime("%d/%m/%Y")

    dados = fetch_serie_em_lotes(11, di_str, df_str, progress_cb=progress_cb)
    if not dados:
        raise ValueError("A API do BCB não retornou dados de Selic.")

    # IMPORTANTE: a série SGS=11 do BCB já retorna a Selic em % AO DIA
    # (taxa diária pronta), NÃO em % ao ano. Por isso o fator é apenas
    # (1 + selic_dia/100), sem elevar a 1/252.
    # O fator acumulado é arredondado a 8 casas decimais a cada passo,
    # seguindo a metodologia da Calculadora do Cidadão (BCB).
    fator = Decimal("1")
    quant_8 = Decimal("0.00000001")
    dias_processados = 0
    for d in dados:
        try:
            val_str = str(d["valor"]).strip().replace(",", ".")
            selic_dia = Decimal(val_str)
            fator_dia = Decimal("1") + selic_dia / Decimal("100")
            fator = (fator * fator_dia).quantize(quant_8,
                                                rounding=ROUND_HALF_UP)
            dias_processados += 1
        except (KeyError, ValueError, ArithmeticError):
            continue

    variacao = (fator - Decimal("1")) * Decimal("100")
    valor_corrigido = (valor * fator) if valor is not None else None

    return {
        "indice_key": "SELIC",
        "indice_nome": "Selic (BCB)",
        "periodo": f"{di_str} a {df_str}",
        "meses": dias_processados,  # dias úteis efetivamente processados
        "fator": fator,
        "variacao": variacao,
        "valor_informado": valor,
        "valor_corrigido": valor_corrigido,
    }


def calcular_unificado(indice_key, ini_mes, ini_ano, fim_mes, fim_ano,
                      valor, progress_cb=None):
    """
    Calcula tanto índice quanto Selic usando MM/AAAA como entrada.
    Para Selic, usa o primeiro dia do mês inicial e último dia do mês final.
    """
    if indice_key == "SELIC":
        ini = date(ini_ano, ini_mes, 1)
        last = monthrange(fim_ano, fim_mes)[1]
        fim = date(fim_ano, fim_mes, last)
        return calcular_selic(ini, fim, valor, progress_cb=progress_cb)
    else:
        return calcular_indice(indice_key, (ini_mes, ini_ano),
                              (fim_mes, fim_ano), valor, progress_cb=progress_cb)


# ===================== Cálculo do Demonstrativo Previdenciário ===================== #

def _produtorio_indice_meses(dados, mes_ini, ano_ini, mes_fim, ano_fim):
    """Calcula produtório do índice entre dois meses (inclusive)."""
    ini_num = ano_ini * 12 + mes_ini
    fim_num = ano_fim * 12 + mes_fim
    if ini_num > fim_num:
        return Decimal("1"), 0

    fator = Decimal("1")
    meses_contados = 0
    for d in dados:
        try:
            _, mes_s, ano_s = d["data"].split("/")
            num = int(ano_s) * 12 + int(mes_s)
            if ini_num <= num <= fim_num:
                v = Decimal(str(d["valor"]))
                fator *= (Decimal("1") + v / Decimal("100"))
                meses_contados += 1
        except (KeyError, ValueError):
            continue
    return fator, meses_contados


def _produtorio_selic_dias(dados, data_ini, data_fim):
    """Calcula produtório da Selic diária entre datas (inclusive).
    Retorna (fator, dias_processados).

    IMPORTANTE: a série SGS=11 do BCB já retorna a Selic em % AO DIA
    (taxa diária pronta), NÃO em % ao ano. Então o fator de cada dia é
    simplesmente (1 + selic_dia/100), sem elevar a 1/252.

    O fator acumulado é arredondado a 8 casas decimais a cada passo,
    seguindo a metodologia da Calculadora do Cidadão (BCB)."""
    if data_ini > data_fim:
        return Decimal("1"), 0

    fator = Decimal("1")
    quant_8 = Decimal("0.00000001")  # 8 casas decimais
    dias = 0
    for d in dados:
        try:
            day_s, mes_s, ano_s = d["data"].split("/")
            dt = date(int(ano_s), int(mes_s), int(day_s))
            if data_ini <= dt <= data_fim:
                val_str = str(d["valor"]).strip().replace(",", ".")
                selic_dia = Decimal(val_str)
                fator_dia = Decimal("1") + selic_dia / Decimal("100")
                fator = (fator * fator_dia).quantize(quant_8,
                                                    rounding=ROUND_HALF_UP)
                dias += 1
        except (KeyError, ValueError, ArithmeticError):
            continue
    return fator, dias



def _aliquota_para_data(periodos, data):
    """Retorna (aliq_seg, aliq_pat) para a data, baseado na lista de períodos.
    Retorna (None, None) se não encontrar período compatível."""
    for p in sorted(periodos, key=lambda x: x["data_ini"]):
        fim = p.get("data_fim")
        if data >= p["data_ini"] and (fim is None or data <= fim):
            return p["aliq_seg"], p["aliq_pat"]
    return None, None


def _aliquota_proporcional_mes(periodos, mes, ano):
    """Retorna (aliq_seg, aliq_pat) para o mês/ano da competência.
    Quando um período termina ou começa no meio do mês (ex: ago/2005, mar/2019),
    calcula a alíquota proporcional pelos dias de cada vigência no mês."""
    dias_mes = monthrange(ano, mes)[1]
    d_ini_mes = date(ano, mes, 1)
    d_fim_mes = date(ano, mes, dias_mes)

    periodos_sorted = sorted(periodos, key=lambda x: x["data_ini"])

    total_seg_dias = Decimal("0")
    total_pat_dias = Decimal("0")
    dias_cobertos = 0

    for p in periodos_sorted:
        p_ini = p["data_ini"]
        p_fim = p.get("data_fim")

        inicio = max(p_ini, d_ini_mes)
        fim    = min(p_fim, d_fim_mes) if p_fim else d_fim_mes

        if inicio > d_fim_mes or fim < d_ini_mes:
            continue

        dias = (fim - inicio).days + 1
        total_seg_dias += p["aliq_seg"] * Decimal(dias)
        total_pat_dias += p["aliq_pat"] * Decimal(dias)
        dias_cobertos  += dias

    if dias_cobertos == 0:
        return None, None

    return (total_seg_dias / Decimal(dias_cobertos),
            total_pat_dias / Decimal(dias_cobertos))


def calcular_demonstrativo(config, competencias, progress_cb=None):
    """
    config:
      - indice_key: 'IPCA' ou 'SELIC' (ou outro índice)
      - data_atualizacao: date
      - aliquota_seg: Decimal (ex: 0.14)
      - aliquota_pat: Decimal (ex: 0.28)
      - dia_vencimento: int (ex: 5)
      - juros_pagamento_pct: Decimal (ex: 0.01) [só usado se Selic]
      - multa_limite_pct: Decimal (ex: 0.20) [só usado se Selic]
      - juros_mora_pct: Decimal (ex: 0.01) [só usado para IPCA - 1% ao mês]
      - moeda: str (padrão 'R$')

    competencias: list of dicts:
      - mes, ano
      - base_calculo: Decimal
      - descricao (opcional)
      - vencimento_custom (opcional): date - se ausente, usa dia X do mês seguinte

    Retorna:
      - linhas: list of dicts com tudo calculado
      - totais: dict com totais
    """
    indice_key = config["indice_key"]
    data_atual = config["data_atualizacao"]
    aliq_seg = Decimal(str(config.get("aliquota_seg", "0.14")))
    aliq_pat = Decimal(str(config.get("aliquota_pat", "0.28")))
    periodos_aliq = config.get("periodos_aliquota") or []
    dia_venc = int(config.get("dia_vencimento", 5))
    # juros mensais (1% padrão) - editável; mesma taxa para IPCA (juros de mora ao mês)
    # e Selic (1% no mês do pagamento). Retrocompat: aceita juros_mora_pct ou juros_pagamento_pct.
    juros_mes = config.get("juros_mensais_pct")
    if juros_mes is None:
        juros_mes = config.get("juros_mora_pct",
                              config.get("juros_pagamento_pct", "0.01"))
    juros_mes = Decimal(str(juros_mes))
    # Multa Selic - 0,33% por dia de atraso, limitada por multa_limite_pct
    multa_dia = Decimal(str(config.get("multa_diaria_pct", "0.0033")))
    multa_limite = Decimal(str(config.get("multa_limite_pct", "0.20")))

    # 1) Determinar o range completo de meses/datas necessários
    vencimentos = []
    for c in competencias:
        mes, ano = c["mes"], c["ano"]
        if c.get("vencimento_custom"):
            v = c["vencimento_custom"]
        else:
            # dia X do mês seguinte ao da competência
            if mes == 12:
                mes_venc, ano_venc = 1, ano + 1
            else:
                mes_venc, ano_venc = mes + 1, ano
            ult_dia = monthrange(ano_venc, mes_venc)[1]
            v = date(ano_venc, mes_venc, min(dia_venc, ult_dia))
        vencimentos.append(v)

    venc_min = min(vencimentos)
    venc_max = max(vencimentos)

    # 2) Buscar TODA a série uma única vez
    if indice_key == "SELIC":
        # Selic é diária - precisa do primeiro vencimento até a data de atualização
        di_str = venc_min.strftime("%d/%m/%Y")
        df_str = data_atual.strftime("%d/%m/%Y")
        if progress_cb:
            progress_cb(f"Buscando série Selic ({di_str} → {df_str})...")
        dados = fetch_serie_em_lotes(11, di_str, df_str, progress_cb=progress_cb)
        # Diagnóstico em arquivo de log
        try:
            import sys as _sys
            _ld = (os.path.dirname(_sys.executable) if getattr(_sys, "frozen", False)
                  else os.path.dirname(os.path.abspath(__file__)))
            with open(os.path.join(_ld, "calculo_log.txt"), "a", encoding="utf-8") as _lf:
                _lf.write(f"\n=== {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
                _lf.write(f"Selic chamada: {di_str} a {df_str}\n")
                _lf.write(f"Dias retornados pela API: {len(dados) if dados else 0}\n")
                if dados:
                    _lf.write(f"  Primeiro: {dados[0].get('data')} -> {dados[0].get('valor')}\n")
                    _lf.write(f"  Último:   {dados[-1].get('data')} -> {dados[-1].get('valor')}\n")
        except Exception:
            pass
        if not dados:
            raise ValueError("A API do BCB não retornou dados de Selic.")
        if progress_cb:
            progress_cb(f"Selic recebida: {len(dados)} dias úteis. Calculando...")
    else:
        cfg = INDICES[indice_key]
        # Índice mensal - precisa do mês do vencimento mais antigo até o mês de atualização
        ini_mes_busca, ini_ano_busca = venc_min.month, venc_min.year
        fim_mes_busca, fim_ano_busca = data_atual.month, data_atual.year
        last_day = monthrange(fim_ano_busca, fim_mes_busca)[1]
        di_str = f"01/{ini_mes_busca:02d}/{ini_ano_busca}"
        df_str = f"{last_day:02d}/{fim_mes_busca:02d}/{fim_ano_busca}"
        if progress_cb:
            progress_cb(f"Buscando série {cfg['name']} ({di_str} → {df_str})...")
        dados = fetch_serie_em_lotes(cfg["serie"], di_str, df_str,
                                    progress_cb=progress_cb)
        if not dados:
            raise ValueError(f"A API do BCB não retornou dados de {cfg['name']}.")

    # 3) Para cada competência, calcular tudo
    linhas = []
    for i, c in enumerate(competencias):
        mes, ano = c["mes"], c["ano"]
        base = Decimal(str(c["base_calculo"]))
        venc = vencimentos[i]
        desc = c.get("descricao") or fmt_competencia(mes, ano)

        # Valor devido (Segurado e Patronal)
        # Se houver períodos de alíquota, busca a alíquota correta para a data do vencimento
        if periodos_aliq:
            _as, _ap = _aliquota_proporcional_mes(periodos_aliq, mes, ano)
            aliq_seg_c = _as if _as is not None else aliq_seg
            aliq_pat_c = _ap if _ap is not None else aliq_pat
        else:
            aliq_seg_c = aliq_seg
            aliq_pat_c = aliq_pat
        valor_seg = (base * aliq_seg_c).quantize(Decimal("0.01"))
        valor_pat = (base * aliq_pat_c).quantize(Decimal("0.01"))

        # Fator de correção entre vencimento e atualização
        if indice_key == "SELIC":
            fator, _qtd = _produtorio_selic_dias(dados, venc, data_atual)
            dias_atraso = max(0, (data_atual - venc).days)
            meses_atraso = Decimal(dias_atraso) / Decimal("30")
        else:
            # Da competência (do mês do vencimento) até o mês da atualização
            fator, _qtd = _produtorio_indice_meses(
                dados,
                venc.month, venc.year,
                data_atual.month, data_atual.year,
            )
            dias_atraso = max(0, (data_atual - venc).days)
            # Demonstrativo Previdenciário (via IPCA): a contagem de meses
            # para os juros de mora segue a base 30/360:
            #   meses = (Y2−Y1)*12 + (M2−M1) + (D2 − D1)/30
            # com dias limitados a 30 (final do mês).
            if data_atual >= venc:
                d1 = min(venc.day, 30)
                d2 = min(data_atual.day, 30)
                meses_atraso = (
                    Decimal(data_atual.year - venc.year) * Decimal("12")
                    + Decimal(data_atual.month - venc.month)
                    + Decimal(d2 - d1) / Decimal("30")
                )
                if meses_atraso < 0:
                    meses_atraso = Decimal("0")
            else:
                meses_atraso = Decimal("0")

        # Valor atualizado
        atual_seg = (valor_seg * fator).quantize(Decimal("0.0001"))
        atual_pat = (valor_pat * fator).quantize(Decimal("0.0001"))

        # Juros / multa
        if indice_key == "SELIC":
            # Juros 1% (configurável) sobre o valor devido, no mês do pagamento
            juros_seg = (valor_seg * juros_mes).quantize(Decimal("0.0001"))
            juros_pat = (valor_pat * juros_mes).quantize(Decimal("0.0001"))
            # Multa: 0,33% por dia de atraso (configurável), com teto
            taxa_multa = multa_dia * Decimal(dias_atraso)
            if taxa_multa > multa_limite:
                taxa_multa = multa_limite
            multa_seg = (valor_seg * taxa_multa).quantize(Decimal("0.0001"))
            multa_pat = (valor_pat * taxa_multa).quantize(Decimal("0.0001"))
            total_seg = (atual_seg + juros_seg + multa_seg).quantize(Decimal("0.01"))
            total_pat = (atual_pat + juros_pat + multa_pat).quantize(Decimal("0.01"))
        else:
            # Juros mensais (configurável, default 1%) × meses de atraso, sobre o valor atualizado
            juros_seg = (atual_seg * juros_mes * meses_atraso).quantize(Decimal("0.0001"))
            juros_pat = (atual_pat * juros_mes * meses_atraso).quantize(Decimal("0.0001"))
            multa_seg = Decimal("0")
            multa_pat = Decimal("0")
            total_seg = (atual_seg + juros_seg).quantize(Decimal("0.01"))
            total_pat = (atual_pat + juros_pat).quantize(Decimal("0.01"))

        situacao = "EM ATRASO" if venc < data_atual else "A REPASSAR"

        linhas.append({
            "descricao": desc,
            "competencia": fmt_competencia(mes, ano),
            "base": base,
            "vencimento": venc,
            "data_atualizacao": data_atual,
            "fator": fator,
            "meses_atraso": meses_atraso,
            # Segurado
            "valor_devido_seg": valor_seg,
            "valor_atual_seg": atual_seg,
            "juros_seg": juros_seg,
            "multa_seg": multa_seg,
            "total_seg": total_seg,
            # Patronal
            "valor_devido_pat": valor_pat,
            "valor_atual_pat": atual_pat,
            "juros_pat": juros_pat,
            "multa_pat": multa_pat,
            "total_pat": total_pat,
            # Geral
            "total_geral": (total_seg + total_pat).quantize(Decimal("0.01")),
            "situacao": situacao,
        })

    # 4) Totais
    totais = {
        "qtd_competencias": len(linhas),
        "base_total": sum((l["base"] for l in linhas), Decimal("0")),
        "valor_devido_seg": sum((l["valor_devido_seg"] for l in linhas), Decimal("0")),
        "valor_atual_seg": sum((l["valor_atual_seg"] for l in linhas), Decimal("0")),
        "juros_seg": sum((l["juros_seg"] for l in linhas), Decimal("0")),
        "multa_seg": sum((l["multa_seg"] for l in linhas), Decimal("0")),
        "total_seg": sum((l["total_seg"] for l in linhas), Decimal("0")),
        "valor_devido_pat": sum((l["valor_devido_pat"] for l in linhas), Decimal("0")),
        "valor_atual_pat": sum((l["valor_atual_pat"] for l in linhas), Decimal("0")),
        "juros_pat": sum((l["juros_pat"] for l in linhas), Decimal("0")),
        "multa_pat": sum((l["multa_pat"] for l in linhas), Decimal("0")),
        "total_pat": sum((l["total_pat"] for l in linhas), Decimal("0")),
        "total_geral": sum((l["total_geral"] for l in linhas), Decimal("0")),
    }

    # 5) Honorários (sobre o total geral atualizado)
    aplicar_honor = bool(config.get("aplicar_honorarios", False))
    honor_pct = Decimal(str(config.get("honorarios_pct", "0")))
    if aplicar_honor and honor_pct > 0:
        honorarios = (totais["total_geral"] * honor_pct).quantize(Decimal("0.01"))
    else:
        honorarios = Decimal("0.00")
    totais["aplicar_honorarios"] = aplicar_honor
    totais["honorarios_pct"] = honor_pct
    totais["honorarios"] = honorarios
    totais["total_com_honorarios"] = (totais["total_geral"] + honorarios).quantize(Decimal("0.01"))

    return {
        "config": config,
        "linhas": linhas,
        "totais": totais,
        "indice_nome": (INDICES[indice_key]["name"] if indice_key in INDICES
                       else indice_key),
    }


# ===================== Cobrança Amigável ===================== #
#
# Atualização de um débito ÚNICO (valor cheio do termo) por IPCA, com:
# - Multa 10% (configurável: sobre valor atualizado ou original)
# - Juros 1% a.m. a partir da data de notificação (configurável)
# - Honorários 10% (sobre o total atualizado + multa + juros)
# - Abatimento de parcelas pagas (com ou sem correção)
# - Abatimento de 13º como crédito (corrigido por IPC-FIPE)

def calcular_cobranca_amigavel(config, progress_cb=None):
    """
    config:
      - data_origem: date (data do débito original)
      - valor_origem: Decimal
      - data_atualizacao: date
      - indice_key: 'IPCA' (padrão) ou outro
      - aplicar_multa: bool
      - multa_pct: Decimal (padrão 0.10)
      - multa_sobre: 'atualizado' ou 'original'
      - aplicar_juros: bool
      - juros_mensais_pct: Decimal (padrão 0.01)
      - data_notificacao: date | None (origem dos juros)
      - juros_desde: 'notificacao' ou 'origem'
      - aplicar_honorarios: bool
      - honorarios_pct: Decimal (padrão 0.10)
      - parcelas_pagas: list of {data: date, valor: Decimal} | []
      - parcelas_corrigir: bool (True = trazer cada parcela por IPCA da data
                                  de pagamento → data_atualizacao; False = soma simples)
      - credito_13: dict | None: {data_fato_gerador: date, valor: Decimal} (IPC-FIPE)

    Retorna dict completo do cálculo.
    """
    data_origem = config["data_origem"]
    valor_origem = Decimal(str(config["valor_origem"]))
    data_atual = config["data_atualizacao"]
    indice_key = config.get("indice_key", "IPCA")

    # --- Modo "várias competências" (opcional) ---------------------------
    # Se vier uma lista de competências (cada uma com data + valor), o débito
    # passa a ser a soma delas; cada competência é corrigida pelo IPCA da sua
    # própria data até a data de atualização. Caso contrário, mantém o modo
    # de débito único de sempre (data_origem + valor_origem).
    competencias_in = config.get("competencias") or []
    usar_competencias = len(competencias_in) > 0
    if usar_competencias:
        # data_origem efetiva = a competência mais antiga (para buscar a série
        # e validar abatimentos); valor_origem = soma nominal das competências.
        datas_comp = [c["data"] for c in competencias_in]
        data_origem = min(datas_comp)
        valor_origem = sum(
            (Decimal(str(c["valor"])) for c in competencias_in), Decimal("0")
        ).quantize(Decimal("0.01"))

    if data_atual < data_origem:
        raise ValueError("Data de Atualização deve ser igual ou posterior à Data Origem.")

    # 1) Fator de correção principal (IPCA) data_origem → data_atualizacao
    cfg_idx = INDICES[indice_key]
    last_day = monthrange(data_atual.year, data_atual.month)[1]
    di_str = f"01/{data_origem.month:02d}/{data_origem.year}"
    df_str = f"{last_day:02d}/{data_atual.month:02d}/{data_atual.year}"
    if progress_cb:
        progress_cb(f"Buscando {cfg_idx['name']}...")
    dados_idx = fetch_serie_em_lotes(cfg_idx["serie"], di_str, df_str,
                                    progress_cb=progress_cb)
    if not dados_idx:
        raise ValueError(f"A API do BCB não retornou dados de {cfg_idx['name']}.")

    detalhe_competencias = []
    if usar_competencias:
        # Corrige cada competência a partir do mês do VENCIMENTO (= mês seguinte
        # ao da competência), alinhando com a lógica do Demonstrativo Previdenciário.
        # Ex.: competência jul/23 → vencimento ago/23; IPCA começa em ago/23.
        valor_atualizado = Decimal("0.00")
        for c in competencias_in:
            c_data = c["data"]
            c_valor = Decimal(str(c["valor"]))
            if c_data > data_atual:
                raise ValueError(
                    f"A competência {c_data.strftime('%m/%Y')} é posterior à "
                    f"Data de Atualização. Verifique as datas.")
            # Mês do vencimento = mês seguinte ao da competência
            if c_data.month == 12:
                venc_month, venc_year = 1, c_data.year + 1
            else:
                venc_month, venc_year = c_data.month + 1, c_data.year
            f_c, _ = _produtorio_indice_meses(
                dados_idx,
                venc_month, venc_year,
                data_atual.month, data_atual.year,
            )
            c_corrigido = (c_valor * f_c).quantize(Decimal("0.01"))
            detalhe_competencias.append({
                "data": c_data,
                "venc_month": venc_month,
                "venc_year": venc_year,
                "descricao": c.get("descricao", ""),
                "valor": c_valor,
                "fator": f_c,
                "valor_corrigido": c_corrigido,
            })
            valor_atualizado += c_corrigido
        valor_atualizado = valor_atualizado.quantize(Decimal("0.01"))
        # Fator "médio" só para exibição (valor atualizado / valor nominal)
        if valor_origem > 0:
            fator = (valor_atualizado / valor_origem)
        else:
            fator = Decimal("1")
        qtd_meses = len(competencias_in)
    else:
        fator, qtd_meses = _produtorio_indice_meses(
            dados_idx,
            data_origem.month, data_origem.year,
            data_atual.month, data_atual.year,
        )
        valor_atualizado = (valor_origem * fator).quantize(Decimal("0.01"))

    # 2) Multa
    aplicar_multa = bool(config.get("aplicar_multa", False))
    multa_pct = Decimal(str(config.get("multa_pct", "0.10")))
    multa_sobre = config.get("multa_sobre", "atualizado")
    if aplicar_multa:
        base_multa = valor_atualizado if multa_sobre == "atualizado" else valor_origem
        multa = (base_multa * multa_pct).quantize(Decimal("0.01"))
    else:
        multa = Decimal("0.00")
        base_multa = Decimal("0.00")

    # 3) Juros
    aplicar_juros = bool(config.get("aplicar_juros", False))
    juros_mes = Decimal(str(config.get("juros_mensais_pct", "0.01")))
    juros_desde = config.get("juros_desde", "notificacao")
    data_notif = config.get("data_notificacao")
    if aplicar_juros:
        if juros_desde == "notificacao":
            if not data_notif:
                raise ValueError(
                    "Para aplicar juros a partir da notificação, "
                    "informe a Data de Notificação.")
            data_inicio_juros = data_notif
        else:
            data_inicio_juros = data_origem

        if usar_competencias:
            # Modo "várias competências": cada competência tem seu próprio
            # cálculo de juros em MESES INTEIROS.
            #
            # Quando juros_desde == "notificacao":
            #   meses = (y_atual - y_notif)*12 + (m_atual - m_notif)
            #   se day_atual < day_notif → meses -= 1  (BCB não conta o mês
            #   incompleto quando o dia de atualização ainda não atingiu o dia
            #   da notificação no mês corrente).
            #
            # Quando juros_desde == "origem":
            #   meses = (y_atual - y_comp)*12 + (m_atual - m_comp)
            #   (usa o mês da competência — sem correção por dia)
            juros = Decimal("0.00")
            total_meses = Decimal("0")
            qtd = Decimal(len(detalhe_competencias))
            for det in detalhe_competencias:
                # Base: SEMPRE o mês da competência (igual ao BCB).
                # A data de notificação não substitui a base — serve apenas
                # para a correção pelo dia quando o dia de atualização ainda
                # não atingiu o dia da notificação no mês corrente.
                ref_month = det["data"].month
                ref_year  = det["data"].year
                meses_c   = (data_atual.year - ref_year) * 12 + \
                            (data_atual.month - ref_month)
                # Correção de dia: aplica SOMENTE no modo "notificacao".
                # Ex.: notif 14/07, update 08/06 → day 8 < 14 → -1 mês.
                if (juros_desde == "notificacao" and data_notif
                        and data_atual.day < data_notif.day):
                    meses_c -= 1
                if meses_c < 0:
                    meses_c = 0
                meses_c_dec = Decimal(meses_c)
                juros_c = (det["valor_corrigido"] * juros_mes *
                          meses_c_dec).quantize(Decimal("0.01"))
                det["meses_juros"] = meses_c_dec
                det["juros"] = juros_c
                juros += juros_c
                total_meses += meses_c_dec
            meses_atraso_juros = (
                (total_meses / qtd) if qtd > 0 else Decimal("0")
            )
        else:
            # Modo "débito único" (comportamento original, com dias/30)
            if data_atual < data_inicio_juros:
                meses_atraso_juros = Decimal("0")
            else:
                dias_atraso = (data_atual - data_inicio_juros).days
                meses_atraso_juros = Decimal(dias_atraso) / Decimal("30")
            juros = (valor_atualizado * juros_mes * meses_atraso_juros).quantize(Decimal("0.01"))
    else:
        juros = Decimal("0.00")
        meses_atraso_juros = Decimal("0")
        data_inicio_juros = None

    # 4) Subtotal antes de honorários e abatimentos
    subtotal = (valor_atualizado + multa + juros).quantize(Decimal("0.01"))

    # 5) Honorários
    aplicar_honor = bool(config.get("aplicar_honorarios", False))
    honor_pct = Decimal(str(config.get("honorarios_pct", "0.10")))
    if aplicar_honor:
        honorarios = (subtotal * honor_pct).quantize(Decimal("0.01"))
    else:
        honorarios = Decimal("0.00")

    # 6) Parcelas pagas (abatimento)
    parcelas_pagas = config.get("parcelas_pagas") or []
    parcelas_corrigir = bool(config.get("parcelas_corrigir", True))
    detalhe_parcelas = []
    total_parcelas_corrigido = Decimal("0.00")
    for p in parcelas_pagas:
        p_data = p["data"]
        p_valor = Decimal(str(p["valor"]))
        # Validações: parcela não pode ser depois da atualização nem antes da origem
        if p_data > data_atual:
            raise ValueError(
                f"Parcela paga em {p_data.strftime('%d/%m/%Y')} é "
                f"posterior à Data de Atualização "
                f"({data_atual.strftime('%d/%m/%Y')}). Verifique a data.")
        if p_data < data_origem:
            raise ValueError(
                f"Parcela paga em {p_data.strftime('%d/%m/%Y')} é "
                f"anterior à Data de Origem do débito "
                f"({data_origem.strftime('%d/%m/%Y')}). Verifique a data.")
        if parcelas_corrigir:
            # Corrige cada parcela: do mês de pagamento até o mês da atualização
            f_p, _ = _produtorio_indice_meses(
                dados_idx,
                p_data.month, p_data.year,
                data_atual.month, data_atual.year,
            )
            p_corrigido = (p_valor * f_p).quantize(Decimal("0.01"))
        else:
            f_p = Decimal("1")
            p_corrigido = p_valor
        detalhe_parcelas.append({
            "data": p_data,
            "valor": p_valor,
            "fator": f_p,
            "valor_corrigido": p_corrigido,
        })
        total_parcelas_corrigido += p_corrigido
    total_parcelas_corrigido = total_parcelas_corrigido.quantize(Decimal("0.01"))

    # 7) 13º Salário (IPC-FIPE) — pode somar ou subtrair
    credito_13 = config.get("credito_13")
    credito_13_corrigido = Decimal("0.00")
    credito_13_fator = Decimal("1")
    credito_13_op = "subtrair"
    if credito_13 and credito_13.get("valor"):
        c13_data = credito_13["data_fato_gerador"]
        c13_valor = Decimal(str(credito_13["valor"]))
        credito_13_op = credito_13.get("operacao", "subtrair")
        # Buscar série IPC-FIPE separadamente
        cfg_ipc = INDICES["IPC-SP"]
        last_d = monthrange(data_atual.year, data_atual.month)[1]
        di_ipc = f"01/{c13_data.month:02d}/{c13_data.year}"
        df_ipc = f"{last_d:02d}/{data_atual.month:02d}/{data_atual.year}"
        if progress_cb:
            progress_cb("Buscando IPC-FIPE para o 13º...")
        dados_ipc = fetch_serie_em_lotes(cfg_ipc["serie"], di_ipc, df_ipc,
                                        progress_cb=progress_cb)
        if not dados_ipc:
            raise ValueError("A API do BCB não retornou dados de IPC-FIPE para o 13º.")
        credito_13_fator, _ = _produtorio_indice_meses(
            dados_ipc,
            c13_data.month, c13_data.year,
            data_atual.month, data_atual.year,
        )
        credito_13_corrigido = (c13_valor * credito_13_fator).quantize(Decimal("0.01"))

    # 8) Total final
    sinal_c13 = Decimal("1") if credito_13_op == "somar" else Decimal("-1")
    total = (subtotal + honorarios
            - total_parcelas_corrigido
            + sinal_c13 * credito_13_corrigido).quantize(Decimal("0.01"))

    return {
        "config": config,
        "indice_nome": cfg_idx["name"],
        "data_origem": data_origem,
        "valor_origem": valor_origem,
        "data_atualizacao": data_atual,
        "fator": fator,
        "meses_corrigidos": qtd_meses,
        "valor_atualizado": valor_atualizado,
        "usar_competencias": usar_competencias,
        "competencias": detalhe_competencias,
        "aplicar_multa": aplicar_multa,
        "multa": multa,
        "multa_pct": multa_pct,
        "multa_sobre": multa_sobre,
        "aplicar_juros": aplicar_juros,
        "juros": juros,
        "juros_mensais_pct": juros_mes,
        "juros_desde": juros_desde,
        "data_inicio_juros": data_inicio_juros,
        "meses_atraso_juros": meses_atraso_juros,
        "subtotal": subtotal,
        "aplicar_honorarios": aplicar_honor,
        "honorarios": honorarios,
        "honorarios_pct": honor_pct,
        "parcelas_pagas": detalhe_parcelas,
        "parcelas_corrigir": parcelas_corrigir,
        "total_parcelas_corrigido": total_parcelas_corrigido,
        "credito_13": (None if not credito_13 or not credito_13.get("valor")
                      else {
                          "data": credito_13["data_fato_gerador"],
                          "valor": Decimal(str(credito_13["valor"])),
                          "fator": credito_13_fator,
                          "valor_corrigido": credito_13_corrigido,
                          "operacao": credito_13_op,
                      }),
        "total": total,
    }


# ===================== Atraso de Parcela ===================== #
#
# Para UMA única parcela atrasada, cláusula contratual (Lei 13.275/2002):
# - Correção monetária IPCA do vencimento até o pagamento
# - Multa 10% sobre o valor atualizado
# - Juros 1% a.m. sobre o valor atualizado × meses de atraso

def calcular_atraso_parcela(config, progress_cb=None):
    """
    config:
      - data_vencimento: date
      - data_pagamento: date  (efetivo ou atualização)
      - valor_parcela: Decimal
      - multa_pct: Decimal (padrão 0.10)
      - juros_mensais_pct: Decimal (padrão 0.01)
      - indice_key: 'IPCA' (padrão)
    """
    data_venc = config["data_vencimento"]
    data_pag = config["data_pagamento"]
    valor = Decimal(str(config["valor_parcela"]))
    multa_pct = Decimal(str(config.get("multa_pct", "0.10")))
    juros_mes = Decimal(str(config.get("juros_mensais_pct", "0.01")))
    indice_key = config.get("indice_key", "IPCA")

    if data_pag < data_venc:
        raise ValueError("Data de Pagamento deve ser igual ou posterior ao Vencimento.")

    # Se foi pago no mesmo dia do vencimento, é "pago no prazo":
    # sem multa, sem juros, sem correção monetária (fator = 1).
    dias_atraso = (data_pag - data_venc).days
    if dias_atraso == 0:
        cfg_idx = INDICES[indice_key]
        return {
            "config": config,
            "indice_nome": cfg_idx["name"],
            "data_vencimento": data_venc,
            "data_pagamento": data_pag,
            "valor_parcela": valor,
            "fator": Decimal("1"),
            "meses_corrigidos": 0,
            "meses_atraso": Decimal("0"),
            "dias_atraso": 0,
            "valor_atualizado": valor,
            "multa_pct": multa_pct,
            "multa": Decimal("0.00"),
            "juros_mensais_pct": juros_mes,
            "juros": Decimal("0.00"),
            "total": valor,
        }

    # Correção pelo IPCA
    cfg_idx = INDICES[indice_key]
    last_day = monthrange(data_pag.year, data_pag.month)[1]
    di_str = f"01/{data_venc.month:02d}/{data_venc.year}"
    df_str = f"{last_day:02d}/{data_pag.month:02d}/{data_pag.year}"
    if progress_cb:
        progress_cb(f"Buscando {cfg_idx['name']}...")
    dados = fetch_serie_em_lotes(cfg_idx["serie"], di_str, df_str,
                                progress_cb=progress_cb)
    if not dados:
        raise ValueError(f"A API do BCB não retornou dados de {cfg_idx['name']}.")

    fator, qtd_meses = _produtorio_indice_meses(
        dados,
        data_venc.month, data_venc.year,
        data_pag.month, data_pag.year,
    )
    valor_atualizado = (valor * fator).quantize(Decimal("0.01"))

    # Multa 10% sobre valor atualizado
    multa = (valor_atualizado * multa_pct).quantize(Decimal("0.01"))

    # Juros 1% a.m. × meses de atraso, sobre valor atualizado
    meses_atraso = Decimal(dias_atraso) / Decimal("30")
    juros = (valor_atualizado * juros_mes * meses_atraso).quantize(Decimal("0.01"))

    total = (valor_atualizado + multa + juros).quantize(Decimal("0.01"))

    return {
        "config": config,
        "indice_nome": cfg_idx["name"],
        "data_vencimento": data_venc,
        "data_pagamento": data_pag,
        "valor_parcela": valor,
        "fator": fator,
        "meses_corrigidos": qtd_meses,
        "meses_atraso": meses_atraso,
        "dias_atraso": dias_atraso,
        "valor_atualizado": valor_atualizado,
        "multa_pct": multa_pct,
        "multa": multa,
        "juros_mensais_pct": juros_mes,
        "juros": juros,
        "total": total,
    }


# ===================== Simulador IRPF (Receita Federal) ===================== #
#
# Tabelas progressivas mensais oficiais da Receita Federal.
# Cada entrada é uma vigência com a lista de faixas:
# (limite_superior, aliquota_decimal, parcela_a_deduzir)
# A última faixa usa limite_superior=None (sem teto).
# Fonte: Lei 11.482/2007, MP 1.171/2023, Lei 14.663/2023, MP 1.294/2025.

IRPF_TABELAS = [
    # (data_inicio_vigencia, faixas, desconto_simplificado_mensal)
    (date(2015, 4, 1), [
        (Decimal("1903.98"),  Decimal("0.000"), Decimal("0.00")),
        (Decimal("2826.65"),  Decimal("0.075"), Decimal("142.80")),
        (Decimal("3751.05"),  Decimal("0.150"), Decimal("354.80")),
        (Decimal("4664.68"),  Decimal("0.225"), Decimal("636.13")),
        (None,                Decimal("0.275"), Decimal("869.36")),
    ], None),  # antes de mai/2023 não havia desconto simplificado mensal

    (date(2023, 5, 1), [
        (Decimal("2112.00"),  Decimal("0.000"), Decimal("0.00")),
        (Decimal("2826.65"),  Decimal("0.075"), Decimal("158.40")),
        (Decimal("3751.05"),  Decimal("0.150"), Decimal("370.40")),
        (Decimal("4664.68"),  Decimal("0.225"), Decimal("651.73")),
        (None,                Decimal("0.275"), Decimal("884.96")),
    ], Decimal("528.00")),

    (date(2024, 2, 1), [
        (Decimal("2259.20"),  Decimal("0.000"), Decimal("0.00")),
        (Decimal("2826.65"),  Decimal("0.075"), Decimal("169.44")),
        (Decimal("3751.05"),  Decimal("0.150"), Decimal("381.44")),
        (Decimal("4664.68"),  Decimal("0.225"), Decimal("662.77")),
        (None,                Decimal("0.275"), Decimal("896.00")),
    ], Decimal("564.80")),

    (date(2025, 5, 1), [
        (Decimal("2428.80"),  Decimal("0.000"), Decimal("0.00")),
        (Decimal("2826.65"),  Decimal("0.075"), Decimal("182.16")),
        (Decimal("3751.05"),  Decimal("0.150"), Decimal("394.16")),
        (Decimal("4664.68"),  Decimal("0.225"), Decimal("675.49")),
        (None,                Decimal("0.275"), Decimal("908.73")),
    ], Decimal("607.20")),

    # Tabela 2026: mesma da MP 1.294/2025 (alíquotas e faixas mantidas pela
    # Lei 15.270/2025). Os redutores de IR para faixas até R$ 7.350 da
    # Reforma da Renda NÃO estão aplicados nesta calculadora — quem precisar
    # deve usar o simulador oficial da Receita.
    (date(2026, 1, 1), [
        (Decimal("2428.80"),  Decimal("0.000"), Decimal("0.00")),
        (Decimal("2826.65"),  Decimal("0.075"), Decimal("182.16")),
        (Decimal("3751.05"),  Decimal("0.150"), Decimal("394.16")),
        (Decimal("4664.68"),  Decimal("0.225"), Decimal("675.49")),
        (None,                Decimal("0.275"), Decimal("908.73")),
    ], Decimal("607.20")),
]

# Dedução mensal por dependente (fixa desde 2015)
IRPF_DEDUCAO_DEPENDENTE = Decimal("189.59")


def _irpf_tabela_vigente(competencia):
    """Retorna (faixas, desconto_simplificado) vigentes no mês/ano."""
    vigente = IRPF_TABELAS[0]
    for entry in IRPF_TABELAS:
        if competencia >= entry[0]:
            vigente = entry
        else:
            break
    return vigente[1], vigente[2]


def _irpf_imposto_pela_tabela(base_calculo, faixas):
    """Aplica a tabela progressiva sobre a base de cálculo.
    Usa fórmula simplificada: alíquota da faixa × base − parcela a deduzir."""
    if base_calculo <= 0:
        return Decimal("0.00"), Decimal("0.000"), Decimal("0.00")
    for limite, aliquota, deducao in faixas:
        if limite is None or base_calculo <= limite:
            imposto = (base_calculo * aliquota - deducao).quantize(Decimal("0.01"))
            if imposto < 0:
                imposto = Decimal("0.00")
            return imposto, aliquota, deducao
    return Decimal("0.00"), Decimal("0.000"), Decimal("0.00")


def _irpf_redutor_lei_15270(base_calculo, competencia):
    """Redutor da Lei 15.270/2025 (Reforma da Renda), válido a partir de
    janeiro/2026. Aplicado sobre o imposto devido na base mensal.

    Fórmula (Receita Federal):
    - Base ≤ R$ 5.000,00: redutor fixo de R$ 312,89 (limitado ao imposto)
    - Base entre R$ 5.000,01 e R$ 7.350,00: R$ 978,62 − (0,133145 × base)
    - Base > R$ 7.350,00: sem redutor (R$ 0,00)

    O redutor nunca pode ser maior que o imposto devido (não há restituição).
    """
    if competencia < date(2026, 1, 1):
        return Decimal("0.00")

    if base_calculo <= Decimal("5000.00"):
        return Decimal("312.89")
    elif base_calculo <= Decimal("7350.00"):
        redutor = (Decimal("978.62") - Decimal("0.133145") * base_calculo)
        if redutor < 0:
            redutor = Decimal("0.00")
        return redutor.quantize(Decimal("0.01"))
    else:
        return Decimal("0.00")


def calcular_irpf(config):
    """
    config:
      - ano: int (ano-calendário, 2015..2026)
      - mes: int (1..12)
      - rendimentos: Decimal (rendimentos tributáveis no mês)
      - previdencia_oficial: Decimal (INSS / Funprev / Funfin)
      - dependentes: int (quantidade)
      - pensao_alimenticia: Decimal
      - outras_deducoes: Decimal (Funpresp, FAPI, Carnê-Leão, Livro Caixa)

    Aplica a regra "mais benéfico ao contribuinte" entre deduções totais e
    desconto simplificado (quando disponível na vigência).
    Retorna dict com detalhes e imposto devido.
    """
    ano = int(config["ano"])
    mes = int(config["mes"])
    rendimentos = Decimal(str(config.get("rendimentos", "0")))
    prev_oficial = Decimal(str(config.get("previdencia_oficial", "0")))
    deps = int(config.get("dependentes", 0))
    pensao = Decimal(str(config.get("pensao_alimenticia", "0")))
    outras = Decimal(str(config.get("outras_deducoes", "0")))

    if ano < 2015 or ano > 2026:
        raise ValueError("Ano-calendário deve estar entre 2015 e 2026.")
    if mes < 1 or mes > 12:
        raise ValueError("Mês deve estar entre 1 e 12.")
    if rendimentos < 0:
        raise ValueError("Rendimentos não podem ser negativos.")

    competencia = date(ano, mes, 1)
    faixas, desc_simpl = _irpf_tabela_vigente(competencia)

    # Deduções legais
    ded_dep = (IRPF_DEDUCAO_DEPENDENTE * deps).quantize(Decimal("0.01"))
    ded_total = (prev_oficial + ded_dep + pensao + outras).quantize(Decimal("0.01"))

    # Desconto simplificado (se disponível)
    if desc_simpl is not None:
        # No regime atual o desconto simplificado limita o teto da isenção,
        # equivalente a 25% do limite da faixa isenta. Aplica-se diretamente
        # como dedução alternativa, comparando com ded_total.
        if desc_simpl > ded_total:
            deducao_utilizada = desc_simpl
            regime = "simplificado"
        else:
            deducao_utilizada = ded_total
            regime = "completo"
    else:
        deducao_utilizada = ded_total
        regime = "completo"

    base_calculo = (rendimentos - deducao_utilizada).quantize(Decimal("0.01"))
    if base_calculo < 0:
        base_calculo = Decimal("0.00")

    imposto_bruto, aliquota_faixa, parcela_deduzir = _irpf_imposto_pela_tabela(
        base_calculo, faixas)

    # Redutor da Lei 15.270/2025 (somente a partir de jan/2026)
    redutor = _irpf_redutor_lei_15270(base_calculo, competencia)
    if redutor > imposto_bruto:
        redutor = imposto_bruto  # nunca devolve mais do que devido

    imposto = (imposto_bruto - redutor).quantize(Decimal("0.01"))
    if imposto < 0:
        imposto = Decimal("0.00")

    # Demonstrativo por faixa: para cada uma das 5 faixas, mostra a parcela
    # da base que se encaixa nela, a alíquota e o valor calculado.
    demonstrativo_faixas = []
    limite_anterior = Decimal("0.00")
    for limite, aliquota, _ded in faixas:
        if limite is None:
            base_faixa = (base_calculo - limite_anterior).quantize(Decimal("0.01"))
        else:
            limite_efetivo = min(base_calculo, limite)
            base_faixa = (limite_efetivo - limite_anterior).quantize(Decimal("0.01"))
        if base_faixa < 0:
            base_faixa = Decimal("0.00")
        valor_faixa = (base_faixa * aliquota).quantize(Decimal("0.0001"))
        demonstrativo_faixas.append({
            "limite": limite,
            "limite_anterior": limite_anterior,
            "base_faixa": base_faixa,
            "aliquota": aliquota,
            "valor_imposto": valor_faixa,
        })
        if limite is not None:
            limite_anterior = limite

    if rendimentos > 0:
        # Simulador da Receita trunca em 2 casas (ROUND_DOWN), não arredonda
        from decimal import ROUND_DOWN as _RD
        aliquota_efetiva = (imposto / rendimentos * Decimal("100")).quantize(
            Decimal("0.01"), rounding=_RD)
    else:
        aliquota_efetiva = Decimal("0.00")

    return {
        "ano": ano,
        "mes": mes,
        "competencia": competencia,
        "rendimentos": rendimentos,
        "previdencia_oficial": prev_oficial,
        "dependentes": deps,
        "deducao_dependentes": ded_dep,
        "pensao_alimenticia": pensao,
        "outras_deducoes": outras,
        "deducoes_completas": ded_total,
        "desconto_simplificado_disponivel": desc_simpl or Decimal("0.00"),
        "regime": regime,
        "deducao_utilizada": deducao_utilizada,
        "base_calculo": base_calculo,
        "aliquota_faixa": aliquota_faixa,
        "parcela_a_deduzir": parcela_deduzir,
        "imposto_bruto": imposto_bruto,
        "redutor_lei_15270": redutor,
        "imposto": imposto,
        "aliquota_efetiva_pct": aliquota_efetiva,
        "faixas": faixas,
        "demonstrativo_faixas": demonstrativo_faixas,
    }


# ===================== Cálculo em Lote Simples ===================== #

def calcular_lote(linhas, progress_cb=None):
    """
    linhas: list of dicts:
      - indice_key
      - mes_ini, ano_ini, mes_fim, ano_fim
      - valor: Decimal ou None
      - descricao (opcional)

    Retorna list of dicts com resultado por linha.
    """
    resultados = []
    total = len(linhas)
    for i, linha in enumerate(linhas):
        if progress_cb:
            progress_cb(f"Calculando linha {i+1}/{total}...")
        try:
            res = calcular_unificado(
                linha["indice_key"],
                linha["mes_ini"], linha["ano_ini"],
                linha["mes_fim"], linha["ano_fim"],
                linha.get("valor"),
                progress_cb=None,  # silencia para não inundar UI
            )
            resultados.append({
                "ok": True,
                "descricao": linha.get("descricao", ""),
                "indice_key": linha["indice_key"],
                "indice_nome": res["indice_nome"],
                "periodo": res["periodo"],
                "meses": res["meses"],
                "fator": res["fator"],
                "variacao": res["variacao"],
                "valor_informado": res["valor_informado"],
                "valor_corrigido": res["valor_corrigido"],
                "erro": None,
            })
        except Exception as e:
            resultados.append({
                "ok": False,
                "descricao": linha.get("descricao", ""),
                "indice_key": linha.get("indice_key"),
                "indice_nome": "",
                "periodo": "",
                "meses": 0,
                "fator": None,
                "variacao": None,
                "valor_informado": linha.get("valor"),
                "valor_corrigido": None,
                "erro": str(e),
            })
    return resultados


# ===================== Import/Export CSV ===================== #

def exportar_csv_lote(filepath, resultados):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Descrição", "Índice", "Período", "Meses/Dias",
                   "Fator", "Variação (%)", "Valor Informado", "Valor Corrigido",
                   "Erro"])
        for r in resultados:
            w.writerow([
                r.get("descricao", ""),
                r.get("indice_nome", "") or r.get("indice_key", ""),
                r.get("periodo", ""),
                str(r.get("meses", "")),
                fmt_fator(r.get("fator")) if r.get("fator") else "",
                fmt_percent(r.get("variacao")) if r.get("variacao") else "",
                fmt_brl(r.get("valor_informado")) if r.get("valor_informado") else "",
                fmt_brl(r.get("valor_corrigido")) if r.get("valor_corrigido") else "",
                r.get("erro") or "",
            ])


def importar_csv_lote(filepath):
    """Lê CSV no formato (delimitador ; ou ,): Descrição; Índice; DataIni; DataFim; Valor"""
    linhas = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        # autodetect delim
        sample = f.read(2048)
        f.seek(0)
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.reader(f, delimiter=delim)
        header = None
        for row in reader:
            if not row or all(not c.strip() for c in row):
                continue
            if header is None:
                header = [c.strip().lower() for c in row]
                continue
            # row: tentar mapear
            d = {}
            for i, val in enumerate(row):
                if i < len(header):
                    d[header[i]] = val.strip()
            # encontrar campos por nome aproximado
            desc = d.get("descrição") or d.get("descricao") or ""
            indice = (d.get("índice") or d.get("indice") or "IPCA").upper().strip()
            data_ini = d.get("data inicial") or d.get("data ini") or d.get("inicial") or ""
            data_fim = d.get("data final") or d.get("data fim") or d.get("final") or ""
            valor = d.get("valor") or ""
            linhas.append({
                "descricao": desc, "indice": indice,
                "data_ini": data_ini, "data_fim": data_fim, "valor": valor,
            })
    return linhas


def exportar_csv_demo(filepath, resultado):
    """Exporta o demonstrativo para CSV no formato estendido."""
    linhas = resultado["linhas"]
    totais = resultado["totais"]
    config = resultado["config"]
    eh_selic = config["indice_key"] == "SELIC"

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([f"Demonstrativo - Correção por {resultado['indice_nome']}"])
        w.writerow([f"Data de Atualização: {config['data_atualizacao'].strftime('%d/%m/%Y')}"])
        w.writerow([f"Alíquota Segurado: {config['aliquota_seg']*100:.0f}% | Patronal: {config['aliquota_pat']*100:.0f}%"])
        w.writerow([])

        if eh_selic:
            w.writerow([
                "Competência", "Descrição", "Base de Cálculo", "Vencimento",
                "Fator Selic",
                "Devido Seg.", "Atualizado Seg.", "1% Seg.", "Multa Seg.", "Total Seg.",
                "Devido Pat.", "Atualizado Pat.", "1% Pat.", "Multa Pat.", "Total Pat.",
                "Total Geral", "Situação",
            ])
            for l in linhas:
                w.writerow([
                    l["competencia"], l["descricao"], fmt_brl(l["base"]),
                    l["vencimento"].strftime("%d/%m/%Y"),
                    fmt_fator(l["fator"]),
                    fmt_brl(l["valor_devido_seg"]), fmt_brl(l["valor_atual_seg"]),
                    fmt_brl(l["juros_seg"]), fmt_brl(l["multa_seg"]),
                    fmt_brl(l["total_seg"]),
                    fmt_brl(l["valor_devido_pat"]), fmt_brl(l["valor_atual_pat"]),
                    fmt_brl(l["juros_pat"]), fmt_brl(l["multa_pat"]),
                    fmt_brl(l["total_pat"]),
                    fmt_brl(l["total_geral"]), l["situacao"],
                ])
        else:
            w.writerow([
                "Competência", "Descrição", "Base de Cálculo", "Vencimento",
                "Fator Índice", "Meses",
                "Devido Seg.", "Atualizado Seg.", "Juros Seg.", "Total Seg.",
                "Devido Pat.", "Atualizado Pat.", "Juros Pat.", "Total Pat.",
                "Total Geral", "Situação",
            ])
            for l in linhas:
                w.writerow([
                    l["competencia"], l["descricao"], fmt_brl(l["base"]),
                    l["vencimento"].strftime("%d/%m/%Y"),
                    fmt_fator(l["fator"]),
                    f"{float(l['meses_atraso']):.2f}".replace(".", ","),
                    fmt_brl(l["valor_devido_seg"]), fmt_brl(l["valor_atual_seg"]),
                    fmt_brl(l["juros_seg"]), fmt_brl(l["total_seg"]),
                    fmt_brl(l["valor_devido_pat"]), fmt_brl(l["valor_atual_pat"]),
                    fmt_brl(l["juros_pat"]), fmt_brl(l["total_pat"]),
                    fmt_brl(l["total_geral"]), l["situacao"],
                ])

        w.writerow([])
        w.writerow([
            "TOTAIS:", "", fmt_brl(totais["base_total"]), "", "", "",
            fmt_brl(totais["valor_devido_seg"]),
            fmt_brl(totais["valor_atual_seg"]),
            fmt_brl(totais["juros_seg"]),
            fmt_brl(totais["total_seg"]),
            fmt_brl(totais["valor_devido_pat"]),
            fmt_brl(totais["valor_atual_pat"]),
            fmt_brl(totais["juros_pat"]),
            fmt_brl(totais["total_pat"]),
            fmt_brl(totais["total_geral"]),
        ])
        # Honorários (se aplicável)
        if totais.get("aplicar_honorarios") and totais.get("honorarios", Decimal("0")) > 0:
            pct_str = f"{float(totais['honorarios_pct'])*100:.1f}".replace(".", ",")
            w.writerow([])
            w.writerow([f"HONORÁRIOS ({pct_str}% sobre o total):", "", "", "", "", "",
                       "", "", "", "", "", "", "", "",
                       fmt_brl(totais["honorarios"])])
            w.writerow(["TOTAL + HONORÁRIOS:", "", "", "", "", "",
                       "", "", "", "", "", "", "", "",
                       fmt_brl(totais["total_com_honorarios"])])


_MESES_PT = {
    "janeiro": "01", "fevereiro": "02", "março": "03", "marco": "03",
    "abril": "04", "maio": "05", "junho": "06", "julho": "07",
    "agosto": "08", "setembro": "09", "outubro": "10",
    "novembro": "11", "dezembro": "12",
    "jan": "01", "fev": "02", "mar": "03", "abr": "04",
    "mai": "05", "jun": "06", "jul": "07", "ago": "08",
    "set": "09", "out": "10", "nov": "11", "dez": "12",
}

def _normalizar_competencia_import(s):
    """Converte competência importada para MM/AAAA.
    Aceita: 'Março/2015', 'março/2015', 'Mar/2015', '3/2015', '03/2015', '2015-03'.
    Retorna a string original se não reconhecer.
    """
    if not s:
        return s
    s = str(s).strip()
    # Já está no formato MM/AAAA
    if re.match(r"^\d{2}/\d{4}$", s):
        return s
    # Formato AAAA-MM (ISO)
    m = re.match(r"^(\d{4})[-/](\d{1,2})$", s)
    if m:
        return f"{int(m.group(2)):02d}/{m.group(1)}"
    # Formato M/AAAA ou MM/AAAA numérico
    m = re.match(r"^(\d{1,2})[/\-](\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{m.group(2)}"
    # Formato "Nome/AAAA" ou "Nome AAAA"
    m = re.match(r"^([A-Za-zçÇãÃõÕêÊéÉáÁóÓúÚíÍ]+)[/\s\-]+(\d{2,4})$", s)
    if m:
        nome = m.group(1).lower()
        ano = m.group(2)
        # Expande ano com 2 dígitos: 00-99 → 2000-2099
        if len(ano) == 2:
            ano = "20" + ano
        # Remove acentos simples para lookup
        nome = nome.replace("ç", "c").replace("ã", "a").replace("õ", "o")
        num = _MESES_PT.get(nome) or _MESES_PT.get(nome[:3])
        if num:
            return f"{num}/{ano}"
    return s


def _normalizar_valor_import(v):
    """Normaliza valores numéricos para o formato brasileiro 'x.xxx,xx'.
    - float/int do openpyxl → formata diretamente
    - '1.234,56' → '1.234,56' (já OK)
    - '1234.56'  → '1.234,56' (ponto decimal internacional)
    - '1234,56'  → '1.234,56'
    - '1.234'    → '1.234,00' (ponto como milhar)
    """
    if v is None or v == "":
        return ""
    # Valor numérico do openpyxl
    if isinstance(v, (int, float)):
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if not s:
        return ""
    # Já tem vírgula → formato brasileiro
    if "," in s:
        # remove pontos de milhar e deixa a vírgula decimal
        s = s.replace(".", "").replace(",", ",")
        return s
    # Só tem ponto
    if "." in s:
        partes = s.split(".")
        decimais = partes[-1]
        if len(decimais) == 2:
            # ponto decimal: "1234.56" → "1234,56"
            inteiro = "".join(partes[:-1])
            return inteiro + "," + decimais
        else:
            # ponto de milhar: "1.234" → "1234"
            return s.replace(".", "")
    return s


def importar_csv_demo(filepath):
    """Lê CSV de competências: Competência (MM/AAAA); Descrição; Base"""
    linhas = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        sample = f.read(2048)
        f.seek(0)
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.reader(f, delimiter=delim)
        header = None
        for row in reader:
            if not row or all(not c.strip() for c in row):
                continue
            if header is None:
                header = [c.strip().lower() for c in row]
                continue
            d = {}
            for i, val in enumerate(row):
                if i < len(header):
                    d[header[i]] = val.strip()
            comp = d.get("competência") or d.get("competencia") or ""
            desc = d.get("descrição") or d.get("descricao") or ""
            base = d.get("base de cálculo") or d.get("base de calculo") or d.get("base") or ""
            linhas.append({
                "competencia": _normalizar_competencia_import(comp),
                "descricao": desc,
                "base": _normalizar_valor_import(base),
            })
    return linhas


# ===================== Import/Export XLSX (opcional) ===================== #
#
# Estilização profissional: cabeçalho azul institucional, zebra colorida,
# totais destacados, situação colorida (verde/vermelho), formato de moeda
# brasileira, freeze panes, larguras otimizadas, altura generosa.

# Paleta XLSX
_XL_AZUL = "003366"
_XL_AZUL_CLARO = "DDE7F0"
_XL_AZUL_MEDIO = "5A85B5"
_XL_VERDE = "1F6F3F"
_XL_VERDE_CLARO = "E8F4ED"
_XL_VERMELHO = "9E2C1E"
_XL_VERMELHO_CLARO = "FCEAE7"
_XL_CINZA = "F4F6F8"
_XL_BORDA = "B8C2CC"
_XL_AMARELO = "FFF6D5"
_XL_TEXTO = "1A2733"

_MONEY_FMT_BR = '_-"R$ "* #,##0.00_-;-"R$ "* #,##0.00_-;_-"R$ "* "-"??_-;_-@_-'
_FATOR_FMT = "#,##0.0000000"
_PERCENT_FMT = '#,##0.0000"%"'
_NUM_FMT = "#,##0.00"


def _xl_border_thin(color=_XL_BORDA):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_apply_borders(ws, row_ini, row_fim, col_ini, col_fim, color=_XL_BORDA):
    b = _xl_border_thin(color)
    for r in range(row_ini, row_fim + 1):
        for c in range(col_ini, col_fim + 1):
            ws.cell(row=r, column=c).border = b


def _xl_title_block(ws, num_cols, title, subtitle="", info="", start_row=1):
    """Bloco de título estilizado em 3 linhas (mais 1 vazia).
    start_row permite posicionar o bloco após o cabeçalho institucional."""
    last_col_letter = get_column_letter(num_cols)
    r = start_row

    # Linha r - faixa azul com título
    ws.row_dimensions[r].height = 36
    ws.merge_cells(f"A{r}:{last_col_letter}{r}")
    c = ws.cell(row=r, column=1)
    c.value = title
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_XL_AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Linha r+1 - subtítulo
    if subtitle:
        ws.row_dimensions[r + 1].height = 22
        ws.merge_cells(f"A{r+1}:{last_col_letter}{r+1}")
        c2 = ws.cell(row=r + 1, column=1)
        c2.value = subtitle
        c2.font = Font(name="Calibri", size=11, bold=True, color=_XL_AZUL)
        c2.fill = PatternFill("solid", fgColor=_XL_AZUL_CLARO)
        c2.alignment = Alignment(horizontal="center", vertical="center")

    # Linha r+2 - info
    if info:
        ws.row_dimensions[r + 2].height = 20
        ws.merge_cells(f"A{r+2}:{last_col_letter}{r+2}")
        c3 = ws.cell(row=r + 2, column=1)
        c3.value = info
        c3.font = Font(name="Calibri", size=9, italic=True, color="555555")
        c3.alignment = Alignment(horizontal="center", vertical="center")

    # Linha r+3 - faixa fina para "respirar"
    ws.row_dimensions[r + 3].height = 6

    return r + 4  # próxima linha disponível


def _xl_processo_block(ws, num_cols, dados, start_row):
    """Cabeçalho institucional do IPREM + dados do processo em formato de
    tabela profissional (label | valor | label | valor), com faixas
    coloridas para o cabeçalho institucional."""
    if not dados:
        return start_row

    last_col_letter = get_column_letter(num_cols)

    # Divide as colunas em 4 regiões: label-e, valor-e, label-d, valor-d.
    # As proporções dependem de num_cols (9 no Lote, 16/17 no Demo).
    if num_cols >= 12:
        label_e_end = 2          # A:B
        valor_e_end = num_cols // 2     # C:meio
        label_d_start = (num_cols // 2) + 1
        label_d_end = (num_cols // 2) + 2
        valor_d_start = (num_cols // 2) + 3
    else:
        # Tabelas mais estreitas (Lote, 9 cols)
        label_e_end = 2          # A:B
        valor_e_end = 4          # C:D
        label_d_start = 5        # E
        label_d_end = 6          # E:F
        valor_d_start = 7        # G

    LE_END = get_column_letter(label_e_end)
    VE_END = get_column_letter(valor_e_end)
    LD_START = get_column_letter(label_d_start)
    LD_END = get_column_letter(label_d_end)
    VD_START = get_column_letter(valor_d_start)

    # Cores e bordas
    BG_BRANCO = "FFFFFF"
    BG_LINHA  = "F8F9FB"  # branco ligeiramente cinzento para alternar
    side_top  = Side(border_style="thin", color="C5CDD6")
    side_sep  = Side(border_style="thin", color="C5CDD6")
    side_strong = Side(border_style="medium", color="003366")

    def _aplicar_estilo_row(row, fill_color):
        """Pinta o fundo de todas as células da row com a cor dada, sem
        bordas internas. Isso garante que o XLSX não renderize 'gaps'
        cinzas em colunas vazias dentro do merge."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = Border()

    def _faixa(row, text, *, fill, font_color, font_size=11, height=22, bold=True):
        """Faixa inteira A:last, com fundo colorido e texto centralizado."""
        ws.row_dimensions[row].height = height
        _aplicar_estilo_row(row, fill)
        # Texto e fonte só na primeira célula (Excel respeita merge)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Calibri", size=font_size, bold=bold, color=font_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Aplica fonte/alignment também na linha toda (defensivo)
        for col in range(2, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=font_size, bold=bold,
                            color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")

    def _campo_duplo(row, label_e, valor_e, label_d, valor_d, *,
                     alt=False, height=20):
        """Linha de tabela com 4 colunas mescladas:
           [LABEL_E] [VALOR_E]  [LABEL_D] [VALOR_D]
        Labels em negrito azul, valores em texto normal preto."""
        ws.row_dimensions[row].height = height
        bg = BG_LINHA if alt else BG_BRANCO
        _aplicar_estilo_row(row, bg)

        thin_bord = Border(top=side_sep, bottom=side_sep)
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_bord

        cL1 = ws.cell(row=row, column=1, value=label_e)
        cL1.font = Font(name="Calibri", size=10, bold=True, color="003366")
        cL1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cV1 = ws.cell(row=row, column=3, value=valor_e)
        cV1.font = Font(name="Calibri", size=10, color="1A2733")
        cV1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cL2 = ws.cell(row=row, column=label_d_start, value=label_d)
        cL2.font = Font(name="Calibri", size=10, bold=True, color="003366")
        cL2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cV2 = ws.cell(row=row, column=valor_d_start, value=valor_d)
        cV2.font = Font(name="Calibri", size=10, color="1A2733")
        cV2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"A{row}:{LE_END}{row}")
        ws.merge_cells(f"C{row}:{VE_END}{row}")
        ws.merge_cells(f"{LD_START}{row}:{LD_END}{row}")
        ws.merge_cells(f"{VD_START}{row}:{last_col_letter}{row}")

    def _campo_triplo(row, lbl1, val1, lbl2, val2, lbl3, val3, *,
                      alt=False, height=20):
        """Linha com 3 grupos: [LBL1 2col][VAL1] | [LBL2 2col][VAL2] | [LBL3 2col][VAL3]
        Cada label recebe 2 colunas (igual ao _campo_duplo), resto é valor."""
        ws.row_dimensions[row].height = height
        bg = BG_LINHA if alt else BG_BRANCO
        _aplicar_estilo_row(row, bg)

        thin_bord = Border(top=side_sep, bottom=side_sep)
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_bord

        # Divide num_cols em 3 terços; cada terço: label=2 cols, valor=resto
        third = num_cols // 3          # tamanho de cada terço
        # Seção 1 (cols 1 .. third)
        L1s, L1e = 1, 2
        V1s, V1e = 3, third
        # Seção 2 (cols third+1 .. 2*third)
        L2s, L2e = third + 1, third + 2
        V2s, V2e = third + 3, 2 * third
        # Seção 3 (cols 2*third+1 .. num_cols)
        L3s, L3e = 2 * third + 1, 2 * third + 2
        V3s, V3e = 2 * third + 3, num_cols

        def _celula(col_s, col_e, texto, bold=False):
            c = ws.cell(row=row, column=col_s, value=texto)
            c.font = Font(name="Calibri", size=10, bold=bold,
                          color="003366" if bold else "1A2733")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            if col_e > col_s:
                ws.merge_cells(
                    f"{get_column_letter(col_s)}{row}:{get_column_letter(col_e)}{row}"
                )

        _celula(L1s, L1e, lbl1, bold=True)
        _celula(V1s, V1e, val1)
        _celula(L2s, L2e, lbl2, bold=True)
        _celula(V2s, V2e, val2)
        _celula(L3s, L3e, lbl3, bold=True)
        _celula(V3s, V3e, val3)

    r = start_row

    # 1) Cabeçalho azul escuro - INSTITUTO
    _faixa(r,
           "INSTITUTO DE PREVIDÊNCIA MUNICIPAL DE SÃO PAULO - IPREM  |  "
           "CNPJ: 47.109.087/0001-01",
           fill=_XL_AZUL, font_color="FFFFFF", font_size=12, height=26)
    r += 1
    # 2) Subfaixa azul clara - divisão
    _faixa(r, "IPREM/CGB/DRFPB/Divisão de Arrecadação",
           fill=_XL_AZUL_CLARO, font_color="003366",
           font_size=10, height=20)
    r += 1
    # 3) Servidor(a) | Registro Funcional | Data de Nascimento  ← mesma linha
    _campo_triplo(r,
                  "Servidor(a):",        dados.get("servidor", ""),
                  "Registro Funcional:", dados.get("registro_funcional", ""),
                  "Data de Nascimento:", dados.get("data_nascimento", ""))
    r += 1
    # 4) RG | CPF
    _campo_duplo(r,
                 "RG:",  dados.get("rg", ""),
                 "CPF:", dados.get("cpf", ""), alt=True)
    r += 1
    # 5) Faixa FUNFIN
    _faixa(r, "Fundo Financeiro (FUNFIN) - CNPJ: 46.252.639/0001-65",
           fill=_XL_AZUL_CLARO, font_color="003366",
           font_size=10, height=20)
    r += 1
    # 6) Órgão de Origem | Órgão Cessionário
    _campo_duplo(r,
                 "Órgão de Origem:",    dados.get("orgao_origem", ""),
                 "Órgão Cessionário:", dados.get("orgao_cessionario", ""))
    r += 1
    # 7) Período de Licença | Processo SEI
    _campo_duplo(r,
                 "Período de Licença:", dados.get("periodo_afastamento", ""),
                 "Processo SEI nº:",    dados.get("processo_sei", ""), alt=True)
    r += 1
    # Linha em branco fina pra separar da tabela
    ws.row_dimensions[r].height = 8
    _aplicar_estilo_row(r, BG_BRANCO)
    return r + 1


def _xl_table_header(ws, row, headers):
    """Cabeçalho da tabela estilizado com gradiente visual (cor + borda inferior mais escura)."""
    ws.row_dimensions[row].height = 38
    head_fill = PatternFill("solid", fgColor=_XL_AZUL)
    head_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_white = Side(border_style="thin", color="FFFFFF")
    thick_bot = Side(border_style="medium", color="001F3F")
    head_border = Border(left=thin_white, right=thin_white,
                        top=thin_white, bottom=thick_bot)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = head_align
        cell.border = head_border


def _xl_zebra(ws, row_ini, row_fim, num_cols):
    """Aplica zebra colorida nas linhas (alternando branco e cinza claro)."""
    for r in range(row_ini, row_fim + 1):
        if (r - row_ini) % 2 == 1:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=_XL_CINZA)


def _xl_totais_row(ws, row, num_cols, label_col=1):
    """Estiliza linha de totais (fundo destacado, fonte maior, borda dupla)."""
    ws.row_dimensions[row].height = 28
    fill = PatternFill("solid", fgColor=_XL_AZUL_CLARO)
    font = Font(name="Calibri", size=11, bold=True, color=_XL_AZUL)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    top_thick = Side(border_style="medium", color=_XL_AZUL)
    bot_thick = Side(border_style="medium", color=_XL_AZUL)
    side_thin = Side(border_style="thin", color=_XL_AZUL_MEDIO)
    border = Border(left=side_thin, right=side_thin, top=top_thick, bottom=bot_thick)

    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = align_left if c == label_col else align_right


def _xl_status_cell(cell, status_text):
    """Colore célula de status (A REPASSAR = verde, EM ATRASO = vermelho)."""
    if not status_text:
        return
    s = status_text.upper().strip()
    if "ATRASO" in s:
        cell.fill = PatternFill("solid", fgColor=_XL_VERMELHO_CLARO)
        cell.font = Font(name="Calibri", size=9, bold=True, color=_XL_VERMELHO)
    elif "REPASSAR" in s or "PAGAR" in s:
        cell.fill = PatternFill("solid", fgColor=_XL_VERDE_CLARO)
        cell.font = Font(name="Calibri", size=9, bold=True, color=_XL_VERDE)
    else:
        cell.font = Font(name="Calibri", size=9, color=_XL_TEXTO)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _xl_set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _xl_print_setup(ws, landscape_mode=False, fit_to_width=True):
    """Configurações de impressão / página."""
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE if landscape_mode
                                 else ws.ORIENTATION_PORTRAIT)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    if fit_to_width:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6


def exportar_xlsx_lote(filepath, resultados, dados_processo=None):
    if not HAS_XLSX:
        raise RuntimeError("openpyxl não está instalado. Use: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lote"

    headers = ["Descrição", "Índice", "Período", "Meses/Dias",
               "Fator", "Variação", "Valor Informado",
               "Valor Corrigido", "Status"]
    num_cols = len(headers)

    # 1) Cabeçalho institucional IPREM (topo)
    proc_end = _xl_processo_block(ws, num_cols, dados_processo, start_row=1)

    # 2) Título do documento (abaixo do bloco IPREM)
    title_end = _xl_title_block(
        ws, num_cols,
        title="Calculadora do Cidadão · Correção em Lote",
        subtitle="Banco Central do Brasil — Reprodução não-oficial",
        info=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
             f"{len(resultados)} cálculo(s)",
        start_row=proc_end if dados_processo else 1,
    )
    HEADER_ROW = title_end
    _xl_table_header(ws, HEADER_ROW, headers)

    row = HEADER_ROW + 1
    total_corrigido = Decimal("0")
    ok_count = 0
    for r in resultados:
        ws.row_dimensions[row].height = 22
        # Descrição
        ws.cell(row=row, column=1, value=r.get("descricao", "")).alignment = \
            Alignment(horizontal="left", vertical="center", indent=1)
        # Índice
        ws.cell(row=row, column=2, value=r.get("indice_nome", "")
                or r.get("indice_key", "")).alignment = \
            Alignment(horizontal="center", vertical="center")
        # Período
        ws.cell(row=row, column=3, value=r.get("periodo", "")).alignment = \
            Alignment(horizontal="center", vertical="center")
        # Qtd
        if r.get("meses"):
            c = ws.cell(row=row, column=4, value=int(r["meses"]))
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Fator
        if r.get("fator"):
            c = ws.cell(row=row, column=5, value=float(r["fator"]))
            c.number_format = _FATOR_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")
        # Variação
        if r.get("variacao"):
            c = ws.cell(row=row, column=6, value=float(r["variacao"]))
            c.number_format = _PERCENT_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")
        # Valor informado
        if r.get("valor_informado"):
            c = ws.cell(row=row, column=7, value=float(r["valor_informado"]))
            c.number_format = _MONEY_FMT_BR
            c.alignment = Alignment(horizontal="right", vertical="center")
        # Valor corrigido (em destaque)
        if r.get("valor_corrigido"):
            total_corrigido += r["valor_corrigido"]
            c = ws.cell(row=row, column=8, value=float(r["valor_corrigido"]))
            c.number_format = _MONEY_FMT_BR
            c.font = Font(name="Calibri", size=10, bold=True, color=_XL_VERDE)
            c.alignment = Alignment(horizontal="right", vertical="center")
        # Status
        status = "OK" if r.get("ok") else f"ERRO"
        sc = ws.cell(row=row, column=9, value=status)
        if r.get("ok"):
            sc.fill = PatternFill("solid", fgColor=_XL_VERDE_CLARO)
            sc.font = Font(name="Calibri", size=9, bold=True, color=_XL_VERDE)
            ok_count += 1
        else:
            sc.fill = PatternFill("solid", fgColor=_XL_VERMELHO_CLARO)
            sc.font = Font(name="Calibri", size=9, bold=True, color=_XL_VERMELHO)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # Zebra (sem mexer na cor do status que já foi colorido)
    for r in range(HEADER_ROW + 1, row):
        if (r - HEADER_ROW - 1) % 2 == 1:
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=r, column=c)
                # preserva preenchimento colorido (status) sobrepondo apenas branco
                if not cell.fill or cell.fill.fgColor is None or \
                   cell.fill.fgColor.value in (None, "00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=_XL_CINZA)

    _xl_apply_borders(ws, HEADER_ROW, row - 1, 1, num_cols)

    # Linha em branco e total
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=7, value="Total corrigido:").alignment = \
        Alignment(horizontal="right", vertical="center")
    if total_corrigido > 0:
        c = ws.cell(row=row, column=8, value=float(total_corrigido))
        c.number_format = _MONEY_FMT_BR
    _xl_totais_row(ws, row, num_cols, label_col=1)

    # Larguras
    _xl_set_widths(ws, [32, 22, 26, 12, 14, 14, 18, 20, 14])

    # Freeze e filtro
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(num_cols)}{row - 2}"

    _xl_print_setup(ws, landscape_mode=False)

    wb.save(filepath)


def importar_xlsx_lote(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    linhas = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        if header is None:
            joined = " ".join(str(c).lower() for c in row if c is not None)
            if any(k in joined for k in ["descri", "índic", "indic", "data", "valor"]):
                header = [str(c).lower().strip() if c is not None else "" for c in row]
                continue
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(header):
                d[header[i]] = "" if val is None else str(val).strip()
        desc = d.get("descrição") or d.get("descricao") or ""
        indice = (d.get("índice") or d.get("indice") or "IPCA").upper().strip()
        data_ini = d.get("data inicial") or d.get("data ini") or d.get("inicial") or ""
        data_fim = d.get("data final") or d.get("data fim") or d.get("final") or ""
        valor = d.get("valor") or ""
        if desc or indice or data_ini or valor:
            linhas.append({
                "descricao": desc, "indice": indice,
                "data_ini": data_ini, "data_fim": data_fim, "valor": valor,
            })
    return linhas


def exportar_xlsx_demo(filepath, resultado, dados_processo=None):
    if not HAS_XLSX:
        raise RuntimeError("openpyxl não está instalado. Use: pip install openpyxl")

    config = resultado["config"]
    linhas = resultado["linhas"]
    totais = resultado["totais"]
    eh_selic = config["indice_key"] == "SELIC"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demonstrativo"

    if eh_selic:
        headers = ["Competência", "Descrição", "Base de Cálculo", "Vencimento",
                   "Fator Selic",
                   "Devido Seg.", "Atualizado Seg.", "1% Seg.", "Multa Seg.", "Total Seg.",
                   "Devido Pat.", "Atualizado Pat.", "1% Pat.", "Multa Pat.", "Total Pat.",
                   "Total Geral", "Situação"]
        # mapeamento de colunas
        col_total_seg = 10
        col_total_pat = 15
        col_total_geral = 16
        col_situacao = 17
    else:
        headers = ["Competência", "Descrição", "Base de Cálculo", "Vencimento",
                   "Fator Índice", "Meses Atraso",
                   "Devido Seg.", "Atualizado Seg.", "Juros Seg.", "Total Seg.",
                   "Devido Pat.", "Atualizado Pat.", "Juros Pat.", "Total Pat.",
                   "Total Geral", "Situação"]
        col_total_seg = 10
        col_total_pat = 14
        col_total_geral = 15
        col_situacao = 16

    num_cols = len(headers)

    # Cabeçalho
    aliquotas_pct = (f"Segurado {float(config['aliquota_seg'])*100:.0f}%  ·  "
                    f"Patronal {float(config['aliquota_pat'])*100:.0f}%")
    info = (f"Atualização em {config['data_atualizacao'].strftime('%d/%m/%Y')}  ·  "
            f"Alíquotas: {aliquotas_pct}  ·  "
            f"Vencimento: dia {config['dia_vencimento']} do mês seguinte"
            + (f"  ·  Limite multa: {float(config['multa_limite_pct'])*100:.0f}%"
               if eh_selic else "")
            + f"  ·  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # 1) Cabeçalho institucional IPREM (topo)
    proc_end = _xl_processo_block(ws, num_cols, dados_processo, start_row=1)

    # 2) Título do documento (abaixo do bloco IPREM)
    title_end = _xl_title_block(
        ws, num_cols,
        title="Demonstrativo das Contribuições Previdenciárias a Recolher",
        subtitle=f"Correção monetária por {resultado['indice_nome']}",
        info=info,
        start_row=proc_end if dados_processo else 1,
    )
    HEADER_ROW = title_end
    _xl_table_header(ws, HEADER_ROW, headers)

    # Dados
    row = HEADER_ROW + 1
    for idx, l in enumerate(linhas):
        ws.row_dimensions[row].height = 22

        # Competência
        ws.cell(row=row, column=1, value=l["competencia"]).alignment = \
            Alignment(horizontal="center", vertical="center")
        # Descrição
        ws.cell(row=row, column=2, value=l["descricao"]).alignment = \
            Alignment(horizontal="left", vertical="center", indent=1)
        # Base
        c = ws.cell(row=row, column=3, value=float(l["base"]))
        c.number_format = _MONEY_FMT_BR
        c.alignment = Alignment(horizontal="right", vertical="center")
        # Vencimento
        ws.cell(row=row, column=4, value=l["vencimento"].strftime("%d/%m/%Y"))\
            .alignment = Alignment(horizontal="center", vertical="center")
        # Fator
        c = ws.cell(row=row, column=5, value=float(l["fator"]))
        c.number_format = _FATOR_FMT
        c.alignment = Alignment(horizontal="right", vertical="center")

        if eh_selic:
            base_col_seg = 6
            base_col_pat = 11
            campos_seg = [("valor_devido_seg", base_col_seg),
                         ("valor_atual_seg", base_col_seg + 1),
                         ("juros_seg", base_col_seg + 2),
                         ("multa_seg", base_col_seg + 3),
                         ("total_seg", base_col_seg + 4)]
            campos_pat = [("valor_devido_pat", base_col_pat),
                         ("valor_atual_pat", base_col_pat + 1),
                         ("juros_pat", base_col_pat + 2),
                         ("multa_pat", base_col_pat + 3),
                         ("total_pat", base_col_pat + 4)]
        else:
            # Meses
            c = ws.cell(row=row, column=6, value=float(l["meses_atraso"]))
            c.number_format = "0.00"
            c.alignment = Alignment(horizontal="right", vertical="center")
            base_col_seg = 7
            base_col_pat = 11
            campos_seg = [("valor_devido_seg", base_col_seg),
                         ("valor_atual_seg", base_col_seg + 1),
                         ("juros_seg", base_col_seg + 2),
                         ("total_seg", base_col_seg + 3)]
            campos_pat = [("valor_devido_pat", base_col_pat),
                         ("valor_atual_pat", base_col_pat + 1),
                         ("juros_pat", base_col_pat + 2),
                         ("total_pat", base_col_pat + 3)]

        for key, col in campos_seg + campos_pat:
            c = ws.cell(row=row, column=col, value=float(l[key]))
            c.number_format = _MONEY_FMT_BR
            c.alignment = Alignment(horizontal="right", vertical="center")

        # Totais segurado/patronal em destaque (verde, negrito)
        for col_total in (col_total_seg, col_total_pat):
            ws.cell(row=row, column=col_total).font = \
                Font(name="Calibri", size=10, bold=True, color=_XL_VERDE)

        # Total Geral (azul, negrito, fundo levemente azulado)
        c = ws.cell(row=row, column=col_total_geral, value=float(l["total_geral"]))
        c.number_format = _MONEY_FMT_BR
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name="Calibri", size=10, bold=True, color=_XL_AZUL)
        c.fill = PatternFill("solid", fgColor=_XL_AZUL_CLARO)

        # Situação (colorida)
        sc = ws.cell(row=row, column=col_situacao, value=l["situacao"])
        _xl_status_cell(sc, l["situacao"])

        row += 1

    # Zebra preservando células coloridas
    for r in range(HEADER_ROW + 1, row):
        if (r - HEADER_ROW - 1) % 2 == 1:
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=r, column=c)
                cur_color = (cell.fill.fgColor.value
                            if cell.fill and cell.fill.fgColor else None)
                # só pinta de cinza se ainda estiver "vazio" ou branco
                if cur_color in (None, "00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", fgColor=_XL_CINZA)

    _xl_apply_borders(ws, HEADER_ROW, row - 1, 1, num_cols)

    # Linha de totais
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAIS")
    ws.cell(row=total_row, column=3, value=float(totais["base_total"]))\
        .number_format = _MONEY_FMT_BR

    if eh_selic:
        cols_map = {6: "valor_devido_seg", 7: "valor_atual_seg",
                   8: "juros_seg", 9: "multa_seg", 10: "total_seg",
                   11: "valor_devido_pat", 12: "valor_atual_pat",
                   13: "juros_pat", 14: "multa_pat", 15: "total_pat",
                   16: "total_geral"}
    else:
        cols_map = {7: "valor_devido_seg", 8: "valor_atual_seg",
                   9: "juros_seg", 10: "total_seg",
                   11: "valor_devido_pat", 12: "valor_atual_pat",
                   13: "juros_pat", 14: "total_pat",
                   15: "total_geral"}
    for col, key in cols_map.items():
        c = ws.cell(row=total_row, column=col, value=float(totais[key]))
        c.number_format = _MONEY_FMT_BR

    _xl_totais_row(ws, total_row, num_cols, label_col=1)

    # Realce especial pro Total Geral na linha de totais
    geral_cell = ws.cell(row=total_row, column=col_total_geral)
    geral_cell.fill = PatternFill("solid", fgColor=_XL_AZUL)
    geral_cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    # Honorários (linhas extras)
    if (totais.get("aplicar_honorarios")
            and totais.get("honorarios", Decimal("0")) > 0):
        honor_row = total_row + 1
        pct = float(totais["honorarios_pct"]) * 100
        label_honor = f"HONORÁRIOS ({pct:.1f}% sobre o total geral)".replace(".", ",")

        ws.row_dimensions[honor_row].height = 24
        ws.cell(row=honor_row, column=1, value=label_honor)
        c_honor = ws.cell(row=honor_row, column=col_total_geral,
                         value=float(totais["honorarios"]))
        c_honor.number_format = _MONEY_FMT_BR
        # Estilo da linha de honorários (amarelo suave)
        honor_fill = PatternFill("solid", fgColor="FFF6D5")
        honor_font = Font(name="Calibri", size=11, bold=True, color="9C7700")
        honor_side = Side(border_style="thin", color="C4A93B")
        honor_border = Border(left=honor_side, right=honor_side,
                             top=honor_side, bottom=honor_side)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=honor_row, column=c)
            cell.fill = honor_fill
            cell.font = honor_font
            cell.border = honor_border
            cell.alignment = (Alignment(horizontal="left", vertical="center")
                             if c == 1
                             else Alignment(horizontal="right", vertical="center"))

        # Linha Total + Honorários
        thh_row = honor_row + 1
        ws.row_dimensions[thh_row].height = 30
        ws.cell(row=thh_row, column=1, value="TOTAL + HONORÁRIOS")
        c_thh = ws.cell(row=thh_row, column=col_total_geral,
                       value=float(totais["total_com_honorarios"]))
        c_thh.number_format = _MONEY_FMT_BR
        # Estilo do total final (azul escuro forte)
        thh_fill = PatternFill("solid", fgColor=_XL_AZUL)
        thh_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        thh_border = Border(
            left=Side(border_style="medium", color=_XL_AZUL),
            right=Side(border_style="medium", color=_XL_AZUL),
            top=Side(border_style="medium", color=_XL_AZUL),
            bottom=Side(border_style="medium", color=_XL_AZUL))
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=thh_row, column=c)
            cell.fill = thh_fill
            cell.font = thh_font
            cell.border = thh_border
            cell.alignment = (Alignment(horizontal="left", vertical="center")
                             if c == 1
                             else Alignment(horizontal="right", vertical="center"))

    # Larguras
    if eh_selic:
        widths = [14, 24, 16, 13, 13,
                 14, 15, 12, 13, 16,
                 14, 15, 12, 13, 16,
                 17, 14]
    else:
        widths = [14, 24, 16, 13, 14, 10,
                 14, 15, 13, 16,
                 14, 15, 13, 16,
                 17, 14]
    _xl_set_widths(ws, widths)

    # Freeze pane — só linhas, sem congelar colunas
    # (congelar colunas quebra a renderização das células mescladas do cabeçalho)
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # Auto-filter
    ws.auto_filter.ref = (f"A{HEADER_ROW}:"
                        f"{get_column_letter(num_cols)}{total_row - 1}")

    _xl_print_setup(ws, landscape_mode=True)

    wb.save(filepath)


def importar_xlsx_demo(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    linhas = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        if header is None:
            joined = " ".join(str(c).lower() for c in row if c is not None)
            if any(k in joined for k in ["compet", "base"]):
                header = [str(c).lower().strip() if c is not None else "" for c in row]
                continue
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(header):
                d[header[i]] = "" if val is None else val  # preserva tipo numérico
        comp = d.get("competência") or d.get("competencia") or ""
        desc = d.get("descrição") or d.get("descricao") or ""
        base = d.get("base de cálculo") or d.get("base de calculo") or d.get("base") or ""
        comp = str(comp).strip()
        if "total" in comp.lower():
            continue
        if comp or base:
            linhas.append({
                "competencia": _normalizar_competencia_import(comp),
                "descricao": str(desc).strip() if desc else "",
                "base": _normalizar_valor_import(base),
            })
    return linhas


# ===================== Export PDF (opcional) ===================== #
#
# Usa reportlab. PDFs estilizados com cabeçalho institucional, tabela
# colorida, totais em destaque e rodapé com paginação. O do demonstrativo
# sai em A4 paisagem (muitas colunas); o do lote em A4 retrato.

# Cores reportlab (RGB)
_PDF_AZUL = colors.HexColor("#003366") if HAS_PDF else None
_PDF_AZUL_CLARO = colors.HexColor("#DDE7F0") if HAS_PDF else None
_PDF_AZUL_MEDIO = colors.HexColor("#5A85B5") if HAS_PDF else None
_PDF_VERDE = colors.HexColor("#1F6F3F") if HAS_PDF else None
_PDF_VERDE_CLARO = colors.HexColor("#E8F4ED") if HAS_PDF else None
_PDF_VERMELHO = colors.HexColor("#9E2C1E") if HAS_PDF else None
_PDF_VERMELHO_CLARO = colors.HexColor("#FCEAE7") if HAS_PDF else None
_PDF_CINZA = colors.HexColor("#F4F6F8") if HAS_PDF else None
_PDF_BORDA = colors.HexColor("#B8C2CC") if HAS_PDF else None
_PDF_TEXTO = colors.HexColor("#1A2733") if HAS_PDF else None


def _pdf_styles():
    """Cria estilos de parágrafo padrão."""
    base = getSampleStyleSheet()
    styles = {}
    styles["title"] = ParagraphStyle(
        "Title", parent=base["Heading1"],
        fontName="Helvetica-Bold", fontSize=15,
        textColor=colors.white, alignment=TA_CENTER, leading=18,
    )
    styles["subtitle"] = ParagraphStyle(
        "Subtitle", parent=base["Heading2"],
        fontName="Helvetica-Bold", fontSize=11,
        textColor=_PDF_AZUL, alignment=TA_CENTER, leading=14,
    )
    styles["info"] = ParagraphStyle(
        "Info", parent=base["Normal"],
        fontName="Helvetica", fontSize=8,
        textColor=colors.HexColor("#555555"),
        alignment=TA_CENTER, leading=11,
    )
    styles["foot"] = ParagraphStyle(
        "Foot", parent=base["Normal"],
        fontName="Helvetica-Oblique", fontSize=7,
        textColor=colors.HexColor("#888888"),
        alignment=TA_CENTER,
    )
    styles["small_right"] = ParagraphStyle(
        "SmallR", parent=base["Normal"],
        fontName="Helvetica", fontSize=7,
        textColor=_PDF_TEXTO, alignment=TA_RIGHT,
    )
    styles["small_left"] = ParagraphStyle(
        "SmallL", parent=base["Normal"],
        fontName="Helvetica", fontSize=7,
        textColor=_PDF_TEXTO, alignment=TA_LEFT,
    )
    styles["small_center"] = ParagraphStyle(
        "SmallC", parent=base["Normal"],
        fontName="Helvetica", fontSize=7,
        textColor=_PDF_TEXTO, alignment=TA_CENTER,
    )
    styles["status_ok"] = ParagraphStyle(
        "StatusOk", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=7,
        textColor=_PDF_VERDE, alignment=TA_CENTER,
    )
    styles["status_err"] = ParagraphStyle(
        "StatusErr", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=7,
        textColor=_PDF_VERMELHO, alignment=TA_CENTER,
    )
    return styles


def _pdf_header_block(num_cols_visual_width, title, subtitle, info, page_w):
    """Cria header com faixa azul + subtítulo + info, ocupando page_w."""
    styles = _pdf_styles()
    items = []

    # Faixa azul com título (tabela de 1 célula)
    title_tbl = Table(
        [[Paragraph(title, styles["title"])]],
        colWidths=[page_w],
        rowHeights=[0.9 * cm],
    )
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _PDF_AZUL),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    items.append(title_tbl)

    # Subtítulo com fundo azul claro
    if subtitle:
        sub_tbl = Table(
            [[Paragraph(subtitle, styles["subtitle"])]],
            colWidths=[page_w],
            rowHeights=[0.6 * cm],
        )
        sub_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _PDF_AZUL_CLARO),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        items.append(sub_tbl)

    if info:
        info_tbl = Table(
            [[Paragraph(info, styles["info"])]],
            colWidths=[page_w],
            rowHeights=[0.55 * cm],
        )
        info_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F6F8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        items.append(info_tbl)

    items.append(Spacer(1, 0.3 * cm))
    return items


def _pdf_processo_block(dados, page_w):
    """Bloco com cabeçalho institucional IPREM + dados do processo
    (servidor, RG, CPF etc), para inserir no topo dos PDFs do Lote e
    Demonstrativo. Retorna lista de Flowables. Não emite nada se `dados`
    estiver vazio."""
    if not dados:
        return []
    styles = _pdf_styles()
    items = []

    # Estilos auxiliares
    head_white = ParagraphStyle(
        "ProcHeadW", parent=styles["title"], fontSize=10,
        textColor=colors.white, alignment=TA_CENTER, leading=13,
    )
    head_blue = ParagraphStyle(
        "ProcHeadB", parent=styles["subtitle"], fontSize=9,
        textColor=_PDF_AZUL, alignment=TA_CENTER, leading=12,
    )
    cell_l = ParagraphStyle(
        "ProcCellL", parent=styles["small_left"], fontSize=8.5,
        textColor=_PDF_TEXTO, alignment=TA_LEFT, leading=11,
    )

    h = 0.52 * cm

    def _row1(para, bg=None):
        """Linha simples de largura total."""
        ts = [
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDA),
        ]
        if bg:
            ts.append(("BACKGROUND", (0, 0), (-1, -1), bg))
        t = Table([[para]], colWidths=[page_w], rowHeights=[h])
        t.setStyle(TableStyle(ts))
        return t

    def _row_cols(cells, widths, bg=None):
        """Linha multi-coluna."""
        ts = [
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDA),
            ("LINEAFTER", (0, 0), (-2, -1), 0.5, _PDF_BORDA),
        ]
        if bg:
            ts.append(("BACKGROUND", (0, 0), (-1, -1), bg))
        t = Table([cells], colWidths=widths, rowHeights=[h])
        t.setStyle(TableStyle(ts))
        return t

    # 1) INSTITUTO (azul)
    items.append(_row1(
        Paragraph("INSTITUTO DE PREVIDÊNCIA MUNICIPAL DE SÃO PAULO - IPREM  |  "
                  "CNPJ: 47.109.087/0001-01", head_white), _PDF_AZUL))

    # 2) Divisão (azul claro)
    items.append(_row1(
        Paragraph("IPREM/CGB/DRFPB/Divisão de Arrecadação", head_blue),
        _PDF_AZUL_CLARO))

    # 3) Servidora | Registro Funcional | Data de Nascimento  (3 colunas)
    w3 = [page_w * 0.45, page_w * 0.30, page_w * 0.25]
    items.append(_row_cols([
        Paragraph(f"<b>Servidor(a):</b> {dados.get('servidor','')}", cell_l),
        Paragraph(f"<b>Reg. Funcional:</b> {dados.get('registro_funcional','')}", cell_l),
        Paragraph(f"<b>Dt. Nascimento:</b> {dados.get('data_nascimento','')}", cell_l),
    ], w3))

    # 4) RG | CPF  (2 colunas)
    items.append(_row_cols([
        Paragraph(f"<b>RG:</b> {dados.get('rg','')}", cell_l),
        Paragraph(f"<b>CPF:</b> {dados.get('cpf','')}", cell_l),
    ], [page_w * 0.5, page_w * 0.5]))

    # 5) FUNFIN (azul claro)
    items.append(_row1(
        Paragraph("Fundo Financeiro (FUNFIN) - CNPJ: 46.252.639/0001-65", head_blue),
        _PDF_AZUL_CLARO))

    # 6) Órgão de Origem | Órgão Cessionário  (2 colunas)
    items.append(_row_cols([
        Paragraph(f"<b>Órgão de Origem:</b> {dados.get('orgao_origem','')}", cell_l),
        Paragraph(f"<b>Órgão Cessionário:</b> {dados.get('orgao_cessionario','')}", cell_l),
    ], [page_w * 0.5, page_w * 0.5]))

    # 7) Período de Licença | Processo SEI  (2 colunas)
    items.append(_row_cols([
        Paragraph(f"<b>Período de Licença:</b> {dados.get('periodo_afastamento','')}", cell_l),
        Paragraph(f"<b>Processo SEI nº:</b> {dados.get('processo_sei','')}", cell_l),
    ], [page_w * 0.5, page_w * 0.5]))

    items.append(Spacer(1, 0.3 * cm))
    return items


def _pdf_footer(canvas_obj, doc):
    """Callback de rodapé com data e número de página."""
    canvas_obj.saveState()
    canvas_obj.setFont("Helvetica-Oblique", 7)
    canvas_obj.setFillColor(colors.HexColor("#888888"))
    page_w, page_h = doc.pagesize
    # linha fina superior
    canvas_obj.setStrokeColor(colors.HexColor("#cccccc"))
    canvas_obj.setLineWidth(0.3)
    canvas_obj.line(doc.leftMargin, 1.0 * cm,
                   page_w - doc.rightMargin, 1.0 * cm)
    # texto
    canvas_obj.drawString(
        doc.leftMargin, 0.6 * cm,
        f"Calculadora do Cidadão v{APP_VERSION}  ·  Reprodução não-oficial  ·  "
        f"Dados: api.bcb.gov.br/dados/serie",
    )
    canvas_obj.drawRightString(
        page_w - doc.rightMargin, 0.6 * cm,
        f"Página {doc.page}",
    )
    canvas_obj.restoreState()


def exportar_pdf_lote(filepath, resultados, dados_processo=None):
    if not HAS_PDF:
        raise RuntimeError("reportlab não está instalado. Use: pip install reportlab")

    page_size = portrait(A4)
    doc = SimpleDocTemplate(
        filepath, pagesize=page_size,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.5 * cm,
        title="Lote de Correções - Calculadora do Cidadão",
        author="Calculadora do Cidadão (não-oficial)",
    )
    page_w = page_size[0] - 2.4 * cm

    styles = _pdf_styles()

    story = []
    story.extend(_pdf_header_block(
        num_cols_visual_width=9,
        title="Correção em Lote de Valores",
        subtitle="Banco Central do Brasil — Reprodução não-oficial",
        info=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
             f"{len(resultados)} cálculo(s)",
        page_w=page_w,
    ))
    # Cabeçalho institucional do processo (se preenchido)
    story.extend(_pdf_processo_block(dados_processo, page_w))

    # Tabela
    headers = ["Descrição", "Índice", "Período", "Qtd",
               "Fator", "Var. %", "Valor Inf.", "Valor Corrigido", "Status"]
    data = [headers]

    total_corrigido = Decimal("0")
    for r in resultados:
        if r["ok"]:
            if r.get("valor_corrigido"):
                total_corrigido += r["valor_corrigido"]
            row = [
                Paragraph(r.get("descricao", "")[:40], styles["small_left"]),
                Paragraph(r.get("indice_key", ""), styles["small_center"]),
                Paragraph(r.get("periodo", ""), styles["small_center"]),
                str(r.get("meses", "")),
                fmt_fator(r.get("fator")) if r.get("fator") else "",
                fmt_percent(r.get("variacao")) if r.get("variacao") else "",
                "R$ " + fmt_brl(r.get("valor_informado")) if r.get("valor_informado") else "—",
                "R$ " + fmt_brl(r.get("valor_corrigido")) if r.get("valor_corrigido") else "—",
                Paragraph("OK", styles["status_ok"]),
            ]
        else:
            row = [
                Paragraph(r.get("descricao", "")[:40], styles["small_left"]),
                Paragraph(r.get("indice_key", ""), styles["small_center"]),
                "", "", "", "", "", "",
                Paragraph("ERRO", styles["status_err"]),
            ]
        data.append(row)

    # Linha de total
    if total_corrigido > 0:
        data.append([
            "TOTAL", "", "", "", "", "", "Total corrigido:",
            "R$ " + fmt_brl(total_corrigido), "",
        ])

    col_widths = [4.2 * cm, 1.5 * cm, 2.5 * cm, 0.9 * cm,
                  1.8 * cm, 1.5 * cm, 2.2 * cm, 2.8 * cm, 1.4 * cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    n_data = len(resultados)
    has_total = total_corrigido > 0
    last_data_row = n_data  # 1-indexed seria n_data+1; mas o reportlab usa 0-indexed
    total_row_idx = n_data + 1 if has_total else None

    style_cmds = [
        # Cabeçalho
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#001F3F")),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ALIGN", (3, 1), (7, -1), "RIGHT"),
        ("ALIGN", (1, 1), (2, -1), "CENTER"),
        ("ALIGN", (8, 1), (8, -1), "CENTER"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        # Grid
        ("LINEBELOW", (0, 1), (-1, last_data_row), 0.25, _PDF_BORDA),
        ("LINEAFTER", (0, 0), (-2, last_data_row), 0.25, _PDF_BORDA),
        ("BOX", (0, 0), (-1, last_data_row), 0.6, _PDF_AZUL_MEDIO),
        # Valor corrigido em verde
        ("TEXTCOLOR", (7, 1), (7, last_data_row), _PDF_VERDE),
        ("FONTNAME", (7, 1), (7, last_data_row), "Helvetica-Bold"),
    ]

    # Zebra
    for i in range(1, n_data + 1):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), _PDF_CINZA))

    # Linha de totais
    if total_row_idx is not None:
        style_cmds.extend([
            ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), _PDF_AZUL_CLARO),
            ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
            ("FONTSIZE", (0, total_row_idx), (-1, total_row_idx), 9),
            ("TEXTCOLOR", (0, total_row_idx), (-1, total_row_idx), _PDF_AZUL),
            ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 1.2, _PDF_AZUL),
            ("LINEBELOW", (0, total_row_idx), (-1, total_row_idx), 1.2, _PDF_AZUL),
            ("BACKGROUND", (7, total_row_idx), (7, total_row_idx), _PDF_AZUL),
            ("TEXTCOLOR", (7, total_row_idx), (7, total_row_idx), colors.white),
            ("FONTSIZE", (7, total_row_idx), (7, total_row_idx), 10),
            ("ALIGN", (7, total_row_idx), (7, total_row_idx), "RIGHT"),
        ])

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)


def exportar_pdf_demo(filepath, resultado, dados_processo=None):
    if not HAS_PDF:
        raise RuntimeError("reportlab não está instalado. Use: pip install reportlab")

    config = resultado["config"]
    linhas = resultado["linhas"]
    totais = resultado["totais"]
    eh_selic = config["indice_key"] == "SELIC"

    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        filepath, pagesize=page_size,
        leftMargin=0.8 * cm, rightMargin=0.8 * cm,
        topMargin=1.0 * cm, bottomMargin=1.4 * cm,
        title="Demonstrativo Previdenciário",
        author="Calculadora do Cidadão (não-oficial)",
    )
    page_w = page_size[0] - 1.6 * cm

    styles = _pdf_styles()

    aliquotas_pct = (f"Segurado {float(config['aliquota_seg'])*100:.0f}% · "
                    f"Patronal {float(config['aliquota_pat'])*100:.0f}%")
    info = (f"Atualização: {config['data_atualizacao'].strftime('%d/%m/%Y')} · "
            f"Alíquotas: {aliquotas_pct} · "
            f"Vencimento: dia {config['dia_vencimento']} do mês seguinte"
            + (f" · Limite multa: {float(config['multa_limite_pct'])*100:.0f}%"
               if eh_selic else "")
            + f" · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    story = []
    story.extend(_pdf_header_block(
        num_cols_visual_width=17,
        title="Demonstrativo de Contribuições Previdenciárias a Recolher",
        subtitle=f"Correção monetária por {resultado['indice_nome']}",
        info=info,
        page_w=page_w,
    ))
    # Cabeçalho institucional do processo (se preenchido)
    story.extend(_pdf_processo_block(dados_processo, page_w))

    # Cabeçalhos / colunas
    if eh_selic:
        headers = ["Compet.", "Descrição", "Base (R$)", "Venc.",
                  "Fator Selic",
                  "Dev. Seg.", "Atu. Seg.", "1% Seg.", "Multa Seg.", "Total Seg.",
                  "Dev. Pat.", "Atu. Pat.", "1% Pat.", "Multa Pat.", "Total Pat.",
                  "TOTAL (R$)", "Situação"]
        col_total_seg = 9
        col_total_pat = 14
        col_total_geral = 15
        col_situacao = 16
    else:
        headers = ["Compet.", "Descrição", "Base (R$)", "Venc.",
                  "Fator Índice", "Meses",
                  "Dev. Seg.", "Atu. Seg.", "Juros Seg.", "Total Seg.",
                  "Dev. Pat.", "Atu. Pat.", "Juros Pat.", "Total Pat.",
                  "TOTAL (R$)", "Situação"]
        col_total_seg = 9
        col_total_pat = 13
        col_total_geral = 14
        col_situacao = 15

    n_cols = len(headers)
    data = [headers]

    for l in linhas:
        situacao_style = (styles["status_err"] if "ATRASO" in l["situacao"].upper()
                         else styles["status_ok"])

        if eh_selic:
            row = [
                Paragraph(l["competencia"], styles["small_center"]),
                Paragraph(l["descricao"][:20], styles["small_left"]),
                fmt_brl(l["base"]),
                l["vencimento"].strftime("%d/%m/%y"),
                fmt_fator(l["fator"]),
                fmt_brl(l["valor_devido_seg"]),
                fmt_brl(l["valor_atual_seg"]),
                fmt_brl(l["juros_seg"]),
                fmt_brl(l["multa_seg"]),
                fmt_brl(l["total_seg"]),
                fmt_brl(l["valor_devido_pat"]),
                fmt_brl(l["valor_atual_pat"]),
                fmt_brl(l["juros_pat"]),
                fmt_brl(l["multa_pat"]),
                fmt_brl(l["total_pat"]),
                fmt_brl(l["total_geral"]),
                Paragraph(l["situacao"], situacao_style),
            ]
        else:
            row = [
                Paragraph(l["competencia"], styles["small_center"]),
                Paragraph(l["descricao"][:20], styles["small_left"]),
                fmt_brl(l["base"]),
                l["vencimento"].strftime("%d/%m/%y"),
                fmt_fator(l["fator"]),
                f"{float(l['meses_atraso']):.1f}".replace(".", ","),
                fmt_brl(l["valor_devido_seg"]),
                fmt_brl(l["valor_atual_seg"]),
                fmt_brl(l["juros_seg"]),
                fmt_brl(l["total_seg"]),
                fmt_brl(l["valor_devido_pat"]),
                fmt_brl(l["valor_atual_pat"]),
                fmt_brl(l["juros_pat"]),
                fmt_brl(l["total_pat"]),
                fmt_brl(l["total_geral"]),
                Paragraph(l["situacao"], situacao_style),
            ]
        data.append(row)

    # Linha de totais
    if eh_selic:
        tot_row = [
            "TOTAIS", "", fmt_brl(totais["base_total"]), "", "",
            fmt_brl(totais["valor_devido_seg"]),
            fmt_brl(totais["valor_atual_seg"]),
            fmt_brl(totais["juros_seg"]),
            fmt_brl(totais["multa_seg"]),
            fmt_brl(totais["total_seg"]),
            fmt_brl(totais["valor_devido_pat"]),
            fmt_brl(totais["valor_atual_pat"]),
            fmt_brl(totais["juros_pat"]),
            fmt_brl(totais["multa_pat"]),
            fmt_brl(totais["total_pat"]),
            fmt_brl(totais["total_geral"]),
            "",
        ]
    else:
        tot_row = [
            "TOTAIS", "", fmt_brl(totais["base_total"]), "", "", "",
            fmt_brl(totais["valor_devido_seg"]),
            fmt_brl(totais["valor_atual_seg"]),
            fmt_brl(totais["juros_seg"]),
            fmt_brl(totais["total_seg"]),
            fmt_brl(totais["valor_devido_pat"]),
            fmt_brl(totais["valor_atual_pat"]),
            fmt_brl(totais["juros_pat"]),
            fmt_brl(totais["total_pat"]),
            fmt_brl(totais["total_geral"]),
            "",
        ]
    data.append(tot_row)

    # Larguras (A4 paisagem ~ 27.7cm úteis)
    if eh_selic:
        # 17 colunas
        col_widths = [1.7 * cm, 2.2 * cm, 1.7 * cm, 1.2 * cm, 1.4 * cm,
                     1.5 * cm, 1.5 * cm, 1.0 * cm, 1.4 * cm, 1.7 * cm,
                     1.5 * cm, 1.5 * cm, 1.0 * cm, 1.4 * cm, 1.7 * cm,
                     2.0 * cm, 1.4 * cm]
    else:
        # 16 colunas
        col_widths = [1.7 * cm, 2.4 * cm, 1.8 * cm, 1.3 * cm, 1.5 * cm, 1.1 * cm,
                     1.6 * cm, 1.6 * cm, 1.5 * cm, 1.8 * cm,
                     1.6 * cm, 1.6 * cm, 1.5 * cm, 1.8 * cm,
                     2.0 * cm, 1.4 * cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    n_data = len(linhas)
    total_row_idx = n_data + 1

    style_cmds = [
        # Cabeçalho
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#001F3F")),
        # Body
        ("FONTNAME", (0, 1), (-1, n_data), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, n_data), 6.5),
        ("ALIGN", (2, 1), (col_total_geral, n_data), "RIGHT"),
        ("ALIGN", (3, 1), (3, n_data), "CENTER"),
        ("ALIGN", (4, 1), (4, n_data), "RIGHT"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        # Grid
        ("LINEBELOW", (0, 1), (-1, n_data), 0.25, _PDF_BORDA),
        ("LINEAFTER", (0, 0), (-2, n_data), 0.25, _PDF_BORDA),
        ("BOX", (0, 0), (-1, n_data), 0.6, _PDF_AZUL_MEDIO),
        # Totais Seg/Pat coloridos em verde, bold
        ("TEXTCOLOR", (col_total_seg, 1), (col_total_seg, n_data), _PDF_VERDE),
        ("FONTNAME", (col_total_seg, 1), (col_total_seg, n_data), "Helvetica-Bold"),
        ("TEXTCOLOR", (col_total_pat, 1), (col_total_pat, n_data), _PDF_VERDE),
        ("FONTNAME", (col_total_pat, 1), (col_total_pat, n_data), "Helvetica-Bold"),
        # Total Geral com fundo azul claro
        ("BACKGROUND", (col_total_geral, 1), (col_total_geral, n_data), _PDF_AZUL_CLARO),
        ("TEXTCOLOR", (col_total_geral, 1), (col_total_geral, n_data), _PDF_AZUL),
        ("FONTNAME", (col_total_geral, 1), (col_total_geral, n_data), "Helvetica-Bold"),
    ]

    # Zebra
    for i in range(1, n_data + 1):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (col_total_geral, i), _PDF_CINZA))

    # Linha de totais
    style_cmds.extend([
        ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), _PDF_AZUL_CLARO),
        ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
        ("FONTSIZE", (0, total_row_idx), (-1, total_row_idx), 7.5),
        ("TEXTCOLOR", (0, total_row_idx), (-1, total_row_idx), _PDF_AZUL),
        ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 1.5, _PDF_AZUL),
        ("LINEBELOW", (0, total_row_idx), (-1, total_row_idx), 1.5, _PDF_AZUL),
        ("ALIGN", (2, total_row_idx), (col_total_geral, total_row_idx), "RIGHT"),
        # Total Geral em destaque máximo
        ("BACKGROUND", (col_total_geral, total_row_idx),
         (col_total_geral, total_row_idx), _PDF_AZUL),
        ("TEXTCOLOR", (col_total_geral, total_row_idx),
         (col_total_geral, total_row_idx), colors.white),
        ("FONTSIZE", (col_total_geral, total_row_idx),
         (col_total_geral, total_row_idx), 9),
        ("BOTTOMPADDING", (0, total_row_idx), (-1, total_row_idx), 6),
        ("TOPPADDING", (0, total_row_idx), (-1, total_row_idx), 6),
    ])

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # Resumo final
    story.append(Spacer(1, 0.5 * cm))
    resumo_data = [[
        Paragraph(
            f"<b>Total Segurado:</b> R$ {fmt_brl(totais['total_seg'])}",
            styles["small_center"]),
        Paragraph(
            f"<b>Total Patronal:</b> R$ {fmt_brl(totais['total_pat'])}",
            styles["small_center"]),
        Paragraph(
            f"<b>TOTAL GERAL: R$ {fmt_brl(totais['total_geral'])}</b>",
            ParagraphStyle("BigTotal", fontName="Helvetica-Bold", fontSize=12,
                          textColor=colors.white, alignment=TA_CENTER)),
    ]]
    resumo = Table(resumo_data, colWidths=[page_w / 3] * 3, rowHeights=[1 * cm])
    resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), _PDF_VERDE_CLARO),
        ("BACKGROUND", (1, 0), (1, 0), _PDF_VERDE_CLARO),
        ("BACKGROUND", (2, 0), (2, 0), _PDF_AZUL),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, _PDF_AZUL_MEDIO),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, _PDF_AZUL_MEDIO),
    ]))
    story.append(resumo)

    # Resumo de honorários (se aplicável)
    if (totais.get("aplicar_honorarios")
            and totais.get("honorarios", Decimal("0")) > 0):
        story.append(Spacer(1, 0.4 * cm))
        pct = float(totais["honorarios_pct"]) * 100
        pct_str = f"{pct:.1f}".replace(".", ",")

        # cor "honorários" (amarelo suave)
        _PDF_HONOR_BG = colors.HexColor("#FFF6D5")
        _PDF_HONOR_BORDER = colors.HexColor("#C4A93B")
        _PDF_HONOR_FG = colors.HexColor("#9C7700")

        honor_data = [[
            Paragraph(
                f"<b>Honorários ({pct_str}% sobre o total geral):</b> "
                f"R$ {fmt_brl(totais['honorarios'])}",
                ParagraphStyle("HonorLine", fontName="Helvetica-Bold",
                              fontSize=10, textColor=_PDF_HONOR_FG,
                              alignment=TA_CENTER)),
            Paragraph(
                f"<b>TOTAL + HONORÁRIOS: R$ {fmt_brl(totais['total_com_honorarios'])}</b>",
                ParagraphStyle("BigTotalH", fontName="Helvetica-Bold",
                              fontSize=13, textColor=colors.white,
                              alignment=TA_CENTER)),
        ]]
        honor_tbl = Table(honor_data,
                         colWidths=[page_w * 2 / 3, page_w / 3],
                         rowHeights=[1.1 * cm])
        honor_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), _PDF_HONOR_BG),
            ("BACKGROUND", (1, 0), (1, 0), _PDF_AZUL),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.6, _PDF_HONOR_BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, _PDF_HONOR_BORDER),
        ]))
        story.append(honor_tbl)

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)


# ============ Export Cobrança Amigável (CSV / XLSX / PDF) ============ #

def _cob_linhas_demonstrativo(res):
    """Monta a lista de (rótulo, valor_str, is_total, is_credito) com as
    linhas do demonstrativo da Cobrança Amigável, na ordem de exibição.
    Reaproveitado por CSV, XLSX e PDF para manter consistência."""
    linhas = []
    linhas.append(("Período",
                   f"{res['data_origem'].strftime('%d/%m/%Y')} a "
                   f"{res['data_atualizacao'].strftime('%d/%m/%Y')}", False, False))
    linhas.append(("Índice de correção", res["indice_nome"], False, False))
    linhas.append(("Fator de correção", fmt_fator(res["fator"]), False, False))
    linhas.append(("Valor original", "R$ " + fmt_brl(res["valor_origem"]), False, False))
    linhas.append(("Valor atualizado", "R$ " + fmt_brl(res["valor_atualizado"]), True, False))
    if res.get("aplicar_multa"):
        sobre = "valor atualizado" if res.get("multa_sobre") == "atualizado" else "valor original"
        pct = f"{float(res['multa_pct'])*100:.1f}".replace(".", ",")
        linhas.append((f"Multa ({pct}% sobre {sobre})",
                      "R$ " + fmt_brl(res["multa"]), False, False))
    if res.get("aplicar_juros"):
        pct = f"{float(res['juros_mensais_pct'])*100:.1f}".replace(".", ",")
        meses = f"{float(res['meses_atraso_juros']):.2f}".replace(".", ",")
        linhas.append((f"Juros de mora ({pct}% a.m. x {meses} meses)",
                      "R$ " + fmt_brl(res["juros"]), False, False))
    linhas.append(("Subtotal", "R$ " + fmt_brl(res["subtotal"]), True, False))
    if res.get("aplicar_honorarios"):
        pct = f"{float(res['honorarios_pct'])*100:.1f}".replace(".", ",")
        linhas.append((f"Honorários ({pct}% sobre subtotal)",
                      "R$ " + fmt_brl(res["honorarios"]), False, False))
    if res.get("parcelas_pagas"):
        linhas.append((f"(-) Parcelas pagas ({len(res['parcelas_pagas'])})",
                      "R$ " + fmt_brl(res["total_parcelas_corrigido"]), False, True))
        for i, p in enumerate(res["parcelas_pagas"], start=1):
            linhas.append((f"     #{i} {p['data'].strftime('%d/%m/%Y')} "
                          f"(orig. R$ {fmt_brl(p['valor'])}, fator {fmt_fator(p['fator'])})",
                          "R$ " + fmt_brl(p["valor_corrigido"]), False, True))
    if res.get("credito_13"):
        c13 = res["credito_13"]
        op = c13.get("operacao", "subtrair")
        sinal_txt = "(+)" if op == "somar" else "(−)"
        linhas.append((f"{sinal_txt} 13º salário (IPC-FIPE, fator {fmt_fator(c13['fator'])})",
                      "R$ " + fmt_brl(c13["valor_corrigido"]), False, op == "subtrair"))
    linhas.append(("TOTAL FINAL", "R$ " + fmt_brl(res["total"]), True, False))
    return linhas


def exportar_csv_cobranca(filepath, res):
    """Exporta o resultado da Cobrança Amigável para CSV."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Cobrança Amigável - Demonstrativo de Atualização"])
        w.writerow([f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        w.writerow([])
        w.writerow(["Descrição", "Valor"])
        for rotulo, valor, _is_total, _is_cred in _cob_linhas_demonstrativo(res):
            w.writerow([rotulo, valor])


def exportar_xlsx_cobranca(filepath, res):
    """Exporta o resultado da Cobrança Amigável para XLSX estilizado."""
    if not HAS_XLSX:
        raise RuntimeError("openpyxl não está instalado. Use: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobrança Amigável"

    azul = PatternFill("solid", fgColor="003366")
    azul_claro = PatternFill("solid", fgColor="DDE7F0")
    verde_claro = PatternFill("solid", fgColor="E8F4ED")
    verm_claro = PatternFill("solid", fgColor="FCEAE7")
    f_branco_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    f_bold = Font(name="Calibri", bold=True, size=10)
    f_normal = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="B8C2CC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Cobrança Amigável - Demonstrativo de Atualização"
    c.fill = azul
    c.font = f_branco_bold
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=8, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4
    ws.cell(r, 1, "Descrição").font = f_bold
    ws.cell(r, 2, "Valor").font = f_bold
    ws.cell(r, 1).fill = azul_claro
    ws.cell(r, 2).fill = azul_claro
    ws.cell(r, 1).border = borda
    ws.cell(r, 2).border = borda
    r += 1

    for rotulo, valor, is_total, is_cred in _cob_linhas_demonstrativo(res):
        cr = ws.cell(r, 1, rotulo)
        cv = ws.cell(r, 2, valor)
        cr.border = borda
        cv.border = borda
        cv.alignment = Alignment(horizontal="right")
        if is_total:
            cr.font = f_bold
            cv.font = f_bold
            cr.fill = verde_claro
            cv.fill = verde_claro
        elif is_cred:
            cr.font = f_normal
            cv.font = f_normal
            cr.fill = verm_claro
            cv.fill = verm_claro
        else:
            cr.font = f_normal
            cv.font = f_normal
        r += 1

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 22

    # Segunda aba: Detalhamento mês a mês (só quando há competências)
    competencias = res.get("competencias") or []
    if competencias:
        ws2 = wb.create_sheet("Detalhamento Mês a Mês")
        aplica_juros = res.get("aplicar_juros", False)

        headers_det = ["Competência", "Descrição", "Valor Original (R$)",
                       "Fator", "Valor Corrigido (R$)"]
        if aplica_juros:
            headers_det += ["Meses Juros", "Juros (R$)"]

        # Cabeçalho
        for ci, h in enumerate(headers_det, start=1):
            c = ws2.cell(1, ci, h)
            c.fill = azul
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
        ws2.row_dimensions[1].height = 18

        # Dados
        for ri, det in enumerate(competencias, start=2):
            bg_fill = PatternFill("solid", fgColor="F0F4F8") if ri % 2 == 0                       else PatternFill("solid", fgColor="FFFFFF")
            vals = [
                det["data"].strftime("%m/%Y"),
                det.get("descricao", "") or "",
                float(det["valor"]),
                float(det["fator"]),
                float(det["valor_corrigido"]),
            ]
            if aplica_juros:
                vals += [
                    float(det.get("meses_juros", 0)),
                    float(det.get("juros", 0)),
                ]
            for ci, val in enumerate(vals, start=1):
                c = ws2.cell(ri, ci, val)
                c.border = borda
                c.fill = bg_fill
                c.font = f_normal
                if ci == 1:  # Competência
                    c.alignment = Alignment(horizontal="center")
                elif ci >= 3:  # Valores numéricos
                    if ci == 4:  # Fator
                        c.number_format = "0.000000"
                    else:
                        c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")

        # Linha de totais
        r_tot = len(competencias) + 2
        total_corr = sum(float(d["valor_corrigido"]) for d in competencias)
        total_orig = sum(float(d["valor"]) for d in competencias)
        tot_vals = ["TOTAL", "", total_orig, "", total_corr]
        if aplica_juros:
            tot_juros = sum(float(d.get("juros", 0)) for d in competencias)
            tot_vals += ["", tot_juros]
        for ci, val in enumerate(tot_vals, start=1):
            c = ws2.cell(r_tot, ci, val)
            c.border = borda
            c.font = Font(name="Calibri", bold=True, size=10)
            c.fill = verde_claro
            if ci >= 3 and val != "":
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")

        # Larguras
        col_widths = [14, 28, 20, 14, 20, 14, 16]
        for ci, w in enumerate(col_widths[:len(headers_det)], start=1):
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(filepath)


def exportar_pdf_cobranca(filepath, res):
    """Exporta o resultado da Cobrança Amigável para PDF institucional."""
    if not HAS_PDF:
        raise RuntimeError("reportlab não está instalado. Use: pip install reportlab")

    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        topMargin=1.4 * cm, bottomMargin=1.6 * cm,
    )
    page_w = doc.width
    story = []
    story += _pdf_header_block(
        2, "Cobrança Amigável",
        "Demonstrativo de Atualização de Débito",
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Índice: {res['indice_nome']}",
        page_w)

    data = [[Paragraph("<b>Descrição</b>", styles["small_left"]),
            Paragraph("<b>Valor</b>", styles["small_right"])]]
    estilo_linhas = []
    for idx, (rotulo, valor, is_total, is_cred) in enumerate(
            _cob_linhas_demonstrativo(res), start=1):
        data.append([Paragraph(rotulo, styles["small_left"]),
                    Paragraph(valor, styles["small_right"])])
        if is_total:
            estilo_linhas.append(("BACKGROUND", (0, idx), (-1, idx), _PDF_VERDE_CLARO))
            estilo_linhas.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))
        elif is_cred:
            estilo_linhas.append(("TEXTCOLOR", (1, idx), (1, idx), _PDF_VERMELHO))

    tbl = Table(data, colWidths=[page_w * 0.68, page_w * 0.32])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, _PDF_BORDA),
        ("LINEAFTER", (0, 0), (-2, -1), 0.25, _PDF_BORDA),
        ("BOX", (0, 0), (-1, -1), 0.6, _PDF_AZUL_MEDIO),
    ] + estilo_linhas))
    story.append(tbl)

    if res["total"] < 0:
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(
            "Atenção: saldo credor — parcelas pagas e abatimentos superam o "
            "débito atualizado. Verifique os dados.", styles["small_left"]))

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)


# ============ Export Atraso de Parcela (CSV / XLSX / PDF) ============ #

def _atr_linhas_demonstrativo(res):
    """Linhas (rótulo, valor, is_total) do resultado de Atraso de Parcela."""
    pct_multa = f"{float(res['multa_pct'])*100:.0f}"
    pct_juros = f"{float(res['juros_mensais_pct'])*100:.1f}".replace(".", ",")
    meses = f"{float(res['meses_atraso']):.2f}".replace(".", ",")
    return [
        ("Vencimento", res["data_vencimento"].strftime("%d/%m/%Y"), False),
        ("Pagamento", res["data_pagamento"].strftime("%d/%m/%Y"), False),
        ("Índice de correção", res["indice_nome"], False),
        ("Dias de atraso", str(res["dias_atraso"]), False),
        ("Meses de atraso (~30d)", meses, False),
        ("Fator de correção", fmt_fator(res["fator"]), False),
        ("Valor da parcela", "R$ " + fmt_brl(res["valor_parcela"]), False),
        ("Valor atualizado", "R$ " + fmt_brl(res["valor_atualizado"]), True),
        (f"Multa ({pct_multa}%)", "R$ " + fmt_brl(res["multa"]), False),
        (f"Juros ({pct_juros}% a.m. x {meses} meses)",
         "R$ " + fmt_brl(res["juros"]), False),
        ("TOTAL A PAGAR", "R$ " + fmt_brl(res["total"]), True),
    ]


def exportar_csv_atraso(filepath, res):
    """Exporta o resultado de Atraso de Parcela para CSV."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Atraso de Parcela - Demonstrativo de Atualização"])
        w.writerow([f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        w.writerow([])
        w.writerow(["Descrição", "Valor"])
        for rotulo, valor, _is_total in _atr_linhas_demonstrativo(res):
            w.writerow([rotulo, valor])


def exportar_xlsx_atraso(filepath, res):
    """Exporta o resultado de Atraso de Parcela para XLSX estilizado."""
    if not HAS_XLSX:
        raise RuntimeError("openpyxl não está instalado. Use: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atraso de Parcela"

    azul = PatternFill("solid", fgColor="003366")
    azul_claro = PatternFill("solid", fgColor="DDE7F0")
    verde_claro = PatternFill("solid", fgColor="E8F4ED")
    f_branco_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    f_bold = Font(name="Calibri", bold=True, size=10)
    f_normal = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="B8C2CC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Atraso de Parcela - Demonstrativo de Atualização"
    c.fill = azul
    c.font = f_branco_bold
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=8, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4
    ws.cell(r, 1, "Descrição").font = f_bold
    ws.cell(r, 2, "Valor").font = f_bold
    ws.cell(r, 1).fill = azul_claro
    ws.cell(r, 2).fill = azul_claro
    ws.cell(r, 1).border = borda
    ws.cell(r, 2).border = borda
    r += 1

    for rotulo, valor, is_total in _atr_linhas_demonstrativo(res):
        cr = ws.cell(r, 1, rotulo)
        cv = ws.cell(r, 2, valor)
        cr.border = borda
        cv.border = borda
        cv.alignment = Alignment(horizontal="right")
        if is_total:
            cr.font = f_bold
            cv.font = f_bold
            cr.fill = verde_claro
            cv.fill = verde_claro
        else:
            cr.font = f_normal
            cv.font = f_normal
        r += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    wb.save(filepath)


def exportar_pdf_atraso(filepath, res):
    """Exporta o resultado de Atraso de Parcela para PDF institucional."""
    if not HAS_PDF:
        raise RuntimeError("reportlab não está instalado. Use: pip install reportlab")

    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        topMargin=1.4 * cm, bottomMargin=1.6 * cm,
    )
    page_w = doc.width
    story = []
    story += _pdf_header_block(
        2, "Atraso de Parcela",
        "Demonstrativo de Atualização (Lei 13.275/2002)",
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Índice: {res['indice_nome']}",
        page_w)

    data = [[Paragraph("<b>Descrição</b>", styles["small_left"]),
            Paragraph("<b>Valor</b>", styles["small_right"])]]
    estilo_linhas = []
    for idx, (rotulo, valor, is_total) in enumerate(
            _atr_linhas_demonstrativo(res), start=1):
        data.append([Paragraph(rotulo, styles["small_left"]),
                    Paragraph(valor, styles["small_right"])])
        if is_total:
            estilo_linhas.append(("BACKGROUND", (0, idx), (-1, idx), _PDF_VERDE_CLARO))
            estilo_linhas.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))

    tbl = Table(data, colWidths=[page_w * 0.62, page_w * 0.38])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, _PDF_BORDA),
        ("LINEAFTER", (0, 0), (-2, -1), 0.25, _PDF_BORDA),
        ("BOX", (0, 0), (-1, -1), 0.6, _PDF_AZUL_MEDIO),
    ] + estilo_linhas))
    story.append(tbl)

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)


# ===================== UI - Estilos e Helpers ===================== #

def _entry_make(parent, width=14, placeholder=""):
    """Cria Entry com placeholder em cinza. Usa flag _is_placeholder para
    distinguir 'placeholder visível' de 'valor real igual ao placeholder'."""
    e = tk.Entry(parent, width=width, font=("Verdana", 10),
                 relief="solid", bd=1, bg="white", fg=COLOR_TEXT,
                 highlightthickness=1, highlightbackground="#999",
                 highlightcolor=COLOR_BCB_BLUE)
    e._is_placeholder = False  # type: ignore
    if placeholder:
        e.insert(0, placeholder)
        e.configure(fg="#aaa")
        e._placeholder = placeholder  # type: ignore
        e._is_placeholder = True  # type: ignore

        def on_focus_in(_):
            if getattr(e, "_is_placeholder", False):
                e.delete(0, "end")
                e.configure(fg=COLOR_TEXT)
                e._is_placeholder = False

        def on_focus_out(_):
            if not e.get():
                e.insert(0, placeholder)
                e.configure(fg="#aaa")
                e._is_placeholder = True

        # Qualquer digitação invalida o estado placeholder
        def on_key_press(_):
            if getattr(e, "_is_placeholder", False):
                e._is_placeholder = False
                e.configure(fg=COLOR_TEXT)

        e.bind("<FocusIn>", on_focus_in)
        e.bind("<FocusOut>", on_focus_out)
        e.bind("<KeyPress>", on_key_press, add="+")
    return e


def _entry_value(entry):
    """Retorna o valor real do Entry; '' se for só placeholder."""
    if getattr(entry, "_is_placeholder", False):
        return ""
    return entry.get()


def _entry_set(entry, value):
    """Define valor no Entry, removendo placeholder se necessário."""
    ph = getattr(entry, "_placeholder", None)
    entry.delete(0, "end")
    if value:
        entry.insert(0, str(value))
        entry.configure(fg=COLOR_TEXT)
        entry._is_placeholder = False
    elif ph:
        entry.insert(0, ph)
        entry.configure(fg="#aaa")
        entry._is_placeholder = True


def _mask_month_year(entry):
    def on_key(event):
        entry.after(1, lambda: _apply_mask(entry, [2], max_len=7))
    entry.bind("<KeyRelease>", on_key)


def _mask_full_date(entry):
    def on_key(event):
        entry.after(1, lambda: _apply_mask(entry, [2, 4], max_len=10))
    entry.bind("<KeyRelease>", on_key)


def _apply_mask(entry, positions, max_len=10):
    ph = getattr(entry, "_placeholder", None)
    v = entry.get()
    if ph and v == ph:
        return
    digits = re.sub(r"\D", "", v)[:8]
    if not digits:
        return
    out = ""
    for i, ch in enumerate(digits):
        if i in positions and i > 0:
            out += "/"
        out += ch
    out = out[:max_len]
    if out != v:
        cursor = entry.index("insert")
        entry.delete(0, "end")
        entry.insert(0, out)
        try:
            entry.icursor(min(len(out), cursor + (1 if len(out) > len(v) else 0)))
        except tk.TclError:
            pass


def _bind_valor_format(entry):
    def on_focus_out(_):
        v = _entry_value(entry).strip()
        if not v:
            return
        try:
            num = parse_valor_br(v)
            if num is not None:
                entry.delete(0, "end")
                entry.insert(0, fmt_brl(num))
                entry.configure(fg=COLOR_TEXT)
        except ValueError:
            pass
    entry.bind("<FocusOut>", on_focus_out, add="+")


# ===================== Aplicação principal ===================== #


class CalculadoraApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        # Dimensionamento simples (Tkinter puro, sem mexer em DPI). As abas
        # maiores (Cobrança/Atraso) têm rolagem própria, então o conteúdo
        # continua acessível mesmo em telas menores. O tamanho da fonte é
        # ajustável pelo seletor "Tamanho" (Pequeno/Médio/Grande), lembrado
        # entre sessões.
        self.geometry("1100x780")
        self.minsize(820, 600)

        self.configure(bg=COLOR_BG)
        # Carrega a preferência de tamanho ANTES de montar a UI, para o
        # seletor já nascer marcado na opção correta.
        self._tamanho_atual = self._carregar_pref_tamanho()
        self._build_styles()
        self._build_ui()
        self._update_clock()

        # Aplica o tamanho salvo da última sessão (ou 'auto' na 1ª vez).
        self._aplicar_tamanho_inicial()

        # Confirmação ao fechar a janela (botão X, Alt+F4), para evitar
        # fechar o programa sem querer e perder um cálculo na tela.
        self.protocol("WM_DELETE_WINDOW", self._confirmar_saida)

        # Tutorial de boas-vindas (aparece apenas na primeira vez)
        self.after(400, self._verificar_tutorial)

        # Verifica atualizações em background (após 3 s para não atrasar a abertura)
        self.after(3000, self._verificar_atualizacao)

    def _confirmar_saida(self):
        if messagebox.askyesno(
                "Sair", "Deseja realmente sair?",
                icon="question", default="no", parent=self):
            self.destroy()

    # ---------- Auto-atualização via GitHub Releases ------------------------- #

    def _verificar_atualizacao(self):
        """Verifica em background se há nova versão disponível no GitHub."""
        if GITHUB_REPO.startswith("SEU_USUARIO"):
            return  # repositório não configurado, pula silenciosamente
        def _worker():
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            try:
                req = urllib.request.Request(
                    url,
                    headers={"User-Agent": f"calculadora-bcb/{APP_VERSION}",
                             "Accept": "application/vnd.github+json"},
                )
                with urllib.request.urlopen(req, timeout=8) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                nova_versao = data.get("tag_name", "").lstrip("v")
                if not nova_versao or not _versao_maior(nova_versao, APP_VERSION):
                    return
                # Procura o asset .zip
                asset_url = None
                for asset in data.get("assets", []):
                    if asset.get("name", "").lower().endswith(".zip"):
                        asset_url = asset["browser_download_url"]
                        break
                notes = data.get("body", "").strip()[:400] or "Sem notas de versão."
                # Notifica a UI (thread-safe via after)
                self.after(0, lambda: self._mostrar_atualizacao(
                    nova_versao, asset_url, notes))
            except Exception:
                pass  # falha silenciosa — sem internet, API off etc.
        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    def _mostrar_atualizacao(self, versao: str, asset_url, notes: str):
        """Dialog informando que há nova versão disponível."""
        win = tk.Toplevel(self)
        win.title("Atualização disponível")
        win.resizable(False, False)
        win.grab_set()
        win.configure(bg=COLOR_BG)
        win.geometry("480x320")
        win.transient(self)

        # Cabeçalho
        hdr = tk.Frame(win, bg=COLOR_BCB_BLUE, padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔄  Nova versão disponível!",
                 font=("Segoe UI", 13, "bold"),
                 fg="white", bg=COLOR_BCB_BLUE).pack(anchor="w")
        tk.Label(hdr,
                 text=f"Versão atual: {APP_VERSION}   →   Nova versão: {versao}",
                 font=("Segoe UI", 10),
                 fg="#c8d8e8", bg=COLOR_BCB_BLUE).pack(anchor="w", pady=(2, 0))

        # Notas
        body = tk.Frame(win, bg=COLOR_BG, padx=12, pady=10)
        body.pack(fill="both", expand=True)
        tk.Label(body, text="O que há de novo:", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_BG, fg=COLOR_TEXT).pack(anchor="w")
        txt = tk.Text(body, height=7, wrap="word",
                      font=("Segoe UI", 9), relief="flat",
                      bg="#f0f4f8", fg=COLOR_TEXT, bd=1)
        txt.pack(fill="both", expand=True, pady=(4, 0))
        txt.insert("1.0", notes)
        txt.configure(state="disabled")

        # Botões
        btn_frame = tk.Frame(win, bg=COLOR_BG, padx=12, pady=8)
        btn_frame.pack(fill="x")

        def _fechar():
            win.destroy()

        def _atualizar():
            if not asset_url:
                messagebox.showwarning(
                    "Sem arquivo",
                    "Nenhum arquivo .zip encontrado nessa release.\n"
                    "Acesse o GitHub para baixar manualmente.",
                    parent=win)
                return
            win.destroy()
            self._baixar_e_atualizar(asset_url)

        tk.Button(btn_frame, text="Agora não", command=_fechar,
                  font=("Segoe UI", 9), relief="flat",
                  bg="#cccccc", fg=COLOR_TEXT, padx=10, pady=4).pack(
                      side="right", padx=(6, 0))
        if asset_url:
            tk.Button(btn_frame, text="⬇  Atualizar agora",
                      command=_atualizar,
                      font=("Segoe UI", 9, "bold"), relief="flat",
                      bg=COLOR_BCB_BLUE, fg="white", padx=10, pady=4).pack(
                          side="right")

    def _baixar_e_atualizar(self, asset_url: str):
        """Baixa o novo .zip e cria updater.bat para substituição da pasta."""
        import sys

        win = tk.Toplevel(self)
        win.title("Baixando atualização…")
        win.resizable(False, False)
        win.grab_set()
        win.configure(bg=COLOR_BG)
        win.geometry("380x130")
        win.transient(self)

        tk.Label(win, text="Baixando nova versão, aguarde…",
                 font=("Segoe UI", 10), bg=COLOR_BG, fg=COLOR_TEXT).pack(
                     pady=(18, 6))
        bar = ttk.Progressbar(win, mode="indeterminate", length=340)
        bar.pack(pady=4)
        bar.start(10)
        status_lbl = tk.Label(win, text="Conectando…",
                              font=("Segoe UI", 8), bg=COLOR_BG,
                              fg="#666666")
        status_lbl.pack()

        def _worker():
            try:
                # Só atualiza automaticamente na versão compilada (frozen)
                if not getattr(sys, "frozen", False):
                    self.after(0, lambda: (
                        bar.stop(), win.destroy(),
                        messagebox.showinfo(
                            "Desenvolvimento",
                            "Auto-update disponível apenas na versão compilada.",
                            parent=self)
                    ))
                    return

                current_exe  = sys.executable
                exe_dir      = os.path.dirname(current_exe)   # pasta da app
                app_parent   = os.path.dirname(exe_dir)       # pasta PAI
                app_dir_name = os.path.basename(exe_dir)      # ex.: "CalculadoraBCB"
                zip_path     = os.path.join(exe_dir, "_update.zip")
                temp_dir     = os.path.join(app_parent, "_update_temp")
                new_app_dir  = os.path.join(temp_dir, app_dir_name)
                bat_path     = os.path.join(app_parent, "updater.bat")

                # Download do ZIP
                self.after(0, lambda: status_lbl.configure(
                    text="Baixando arquivo…"))
                req = urllib.request.Request(
                    asset_url,
                    headers={"User-Agent": f"calculadora-bcb/{APP_VERSION}"})
                with urllib.request.urlopen(req, timeout=120) as resp, \
                        open(zip_path, "wb") as f:
                    total = int(resp.headers.get("Content-Length", 0))
                    downloaded = 0
                    while True:
                        chunk = resp.read(65536)
                        if not chunk:
                            break
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total:
                            pct = int(downloaded * 100 / total)
                            self.after(0, lambda p=pct: status_lbl.configure(
                                text=f"Baixando… {p}%"))

                # Cria updater.bat na pasta PAI (fora da app, para não ser deletado)
                # O bat: extrai o ZIP → copia os arquivos novos → reinicia o app
                ps_cmd = (
                    f"Expand-Archive -LiteralPath '{zip_path}' "
                    f"-DestinationPath '{temp_dir}' -Force"
                )
                bat_content = (
                    "@echo off\r\n"
                    "timeout /t 2 /nobreak >nul\r\n"
                    f'powershell -NoProfile -ExecutionPolicy Bypass -Command "{ps_cmd}"\r\n'
                    f'robocopy "{new_app_dir}" "{exe_dir}" /E /IS /IT /NFL /NDL /NJH /NJS /NP\r\n'
                    f'rmdir /S /Q "{temp_dir}"\r\n'
                    f'del "{zip_path}"\r\n'
                    f'start "" "{current_exe}"\r\n'
                    'del "%~f0"\r\n'
                )
                with open(bat_path, "w", encoding="utf-8") as f:
                    f.write(bat_content)

                self.after(0, lambda: self._finalizar_atualizacao(
                    bat_path, win))
            except Exception as err:
                self.after(0, lambda e=err: (
                    bar.stop(),
                    win.destroy(),
                    messagebox.showerror(
                        "Erro ao atualizar",
                        f"Não foi possível baixar a atualização:\n{e}",
                        parent=self)
                ))

        threading.Thread(target=_worker, daemon=True).start()

    def _finalizar_atualizacao(self, bat_path: str, win: tk.Toplevel):
        """Confirma reinício e executa o bat de substituição."""
        win.destroy()
        if not messagebox.askyesno(
                "Atualização pronta",
                "O arquivo foi baixado com sucesso!\n\n"
                "O aplicativo será fechado e reiniciado automaticamente "
                "com a nova versão. Deseja continuar?",
                parent=self):
            return
        import subprocess
        subprocess.Popen(
            ["cmd.exe", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW
            if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        self.destroy()

    # ---------- Tutorial de boas-vindas -------------------------------------- #

    _TUTORIAL_PASSOS = [
        {
            "icone": "⚠️",
            "titulo": "VERSÃO BETA TEST",
            "texto": (
                "Esta é uma versão de TESTE do aplicativo.\n\n"
                "Os cálculos estão sendo validados e podem conter erros. "
                "NÃO utilize os resultados para fins oficiais sem conferir "
                "com a Calculadora do Cidadão do Banco Central "
                "(bcb.gov.br/calcidadao) ou outra fonte confiável.\n\n"
                "Correções e melhorias serão feitas ao longo dos testes. "
                "Agradecemos a sua colaboração!"
            ),
        },
        {
            "icone": "📋",
            "titulo": "Índices de Preços",
            "texto": (
                "Corrige um valor entre duas datas usando IPCA, IGP-M, "
                "INPC, IPCA-E, IPC-BR ou IPC-SP.\n\n"
                "Como preencher:\n"
                "• Valor original → o montante na data de origem\n"
                "• Data de origem → quando o valor era aquele\n"
                "• Data de atualização → hoje (ou a data desejada)\n"
                "• Índice → escolha o índice de correção\n\n"
                "Clique em Calcular. O resultado mostra o fator e o valor corrigido."
            ),
        },
        {
            "icone": "📈",
            "titulo": "Selic",
            "texto": (
                "Corrige um valor pela taxa Selic acumulada no período.\n\n"
                "Como preencher:\n"
                "• Valor original → o montante a corrigir\n"
                "• Data de origem → data inicial\n"
                "• Data de atualização → data final\n\n"
                "Indicada para cálculos financeiros e correções de tributos "
                "federais que usam a Selic como referência."
            ),
        },
        {
            "icone": "📦",
            "titulo": "Lote (várias correções)",
            "texto": (
                "Processa múltiplas correções de uma só vez.\n\n"
                "Como usar:\n"
                "• Adicione linhas com: valor, data inicial, data final e índice\n"
                "• Clique em Calcular Lote\n"
                "• Exporte os resultados em CSV ou XLSX\n\n"
                "Útil quando há muitos valores a corrigir com diferentes "
                "datas e índices."
            ),
        },
        {
            "icone": "🏛️",
            "titulo": "Demonstrativo Previdenciário",
            "texto": (
                "Calcula contribuições previdenciárias em atraso "
                "(FUNFIN/SEI), com correção e juros por competência.\n\n"
                "Como preencher:\n"
                "• Base de cálculo → salário de contribuição do mês\n"
                "• Competência → mês/ano da contribuição atrasada\n"
                "• Alíquotas → Segurado e Patronal (%)\n"
                "• Índice de correção → IPCA (padrão) ou Selic\n"
                "• Data de atualização → data base do cálculo\n\n"
                "Gera o total de Segurado + Patronal com correção, juros e "
                "o detalhamento por competência. Exporta para CSV, XLSX e PDF."
            ),
        },
        {
            "icone": "🤝",
            "titulo": "Cobrança Amigável",
            "texto": (
                "Atualiza um débito único com correção + juros + multa + "
                "honorários, com abatimento de parcelas pagas e crédito de 13º.\n\n"
                "Como preencher:\n"
                "• Data de origem / Valor original → o débito inicial\n"
                "• Data de atualização → data base\n"
                "• Multa, Juros, Honorários → marque e informe % se quiser\n"
                "• Parcelas pagas → abatimento de valores já quitados\n\n"
                "Para várias competências, marque\n"
                "\"Lançar débito em várias competências\" e entre com cada "
                "mês separadamente — isso garante que o resultado bata com "
                "o Demonstrativo Previdenciário."
            ),
        },
        {
            "icone": "📅",
            "titulo": "Atraso de Parcela",
            "texto": (
                "Calcula o valor atualizado de uma parcela contratual "
                "em atraso (Lei 13.275/2002).\n\n"
                "Como preencher:\n"
                "• Valor da parcela → valor original no vencimento\n"
                "• Data de vencimento → quando era para ter sido pago\n"
                "• Data de pagamento → data efetiva (ou data de cálculo)\n\n"
                "O sistema aplica automaticamente:\n"
                "correção IPCA + multa 10% + juros 1% a.m."
            ),
        },
        {
            "icone": "💾",
            "titulo": "Exportações",
            "texto": (
                "Todos os resultados podem ser exportados.\n\n"
                "Formatos disponíveis:\n"
                "• CSV → para qualquer planilha (Excel, LibreOffice…)\n"
                "• XLSX → Excel formatado com layout institucional BCB\n"
                "• PDF → disponível no Demonstrativo e Cobrança Amigável\n\n"
                "Os botões de exportação aparecem logo acima do resultado "
                "após o cálculo.\n\n"
                "Lembre-se: os valores são obtidos em tempo real da API "
                "pública do Banco Central. É necessária conexão com a internet."
            ),
        },
    ]

    def _verificar_tutorial(self):
        """Exibe o tutorial se ainda não foi visto."""
        path = self._arquivo_config()
        if path and os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                if cfg.get("tutorial_visto"):
                    return  # já viu, não mostra de novo
            except Exception:
                pass
        self._mostrar_tutorial()

    def _marcar_tutorial_visto(self):
        """Salva no config que o tutorial já foi exibido."""
        path = self._arquivo_config()
        if not path:
            return
        cfg = {}
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
            except Exception:
                pass
        cfg["tutorial_visto"] = True
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cfg, f)
        except Exception:
            pass

    def _mostrar_tutorial(self):
        """Cria e exibe o overlay de tutorial estilo card flutuante."""
        passos = self._TUTORIAL_PASSOS
        total = len(passos)
        estado = {"passo": 0, "janela": None}

        def _criar_card(passo_idx):
            # Destrói card anterior se existir
            if estado["janela"] and estado["janela"].winfo_exists():
                estado["janela"].destroy()

            p = passos[passo_idx]
            num = passo_idx + 1

            # Janela-card (Toplevel sem decoração)
            card = tk.Toplevel(self)
            card.overrideredirect(True)          # sem barra de título
            card.attributes("-topmost", True)
            card.configure(bg="#e0e0e0")         # sombra simulada
            estado["janela"] = card

            # Sombra: borda levemente mais escura
            sombra = tk.Frame(card, bg="#b0b8c1", padx=2, pady=2)
            sombra.pack(fill="both", expand=True, padx=3, pady=3)

            # Card principal
            corpo = tk.Frame(sombra, bg=COLOR_PANEL, padx=22, pady=18,
                             relief="flat", bd=0)
            corpo.pack(fill="both", expand=True)

            # Linha 1: badge de passo
            badge_frame = tk.Frame(corpo, bg=COLOR_PANEL)
            badge_frame.pack(fill="x", anchor="w")

            # Badge colorido (ex: "1 / 8")
            badge_bg = COLOR_BCB_BLUE if passo_idx > 0 else "#c0392b"
            badge = tk.Label(
                badge_frame,
                text=f"  {num} / {total}  ",
                bg=badge_bg, fg="white",
                font=("Verdana", 8, "bold"),
                padx=4, pady=2,
            )
            badge.pack(side="left")

            # Linha 2: ícone + título
            titulo_frame = tk.Frame(corpo, bg=COLOR_PANEL)
            titulo_frame.pack(fill="x", pady=(10, 0))

            tk.Label(titulo_frame, text=p["icone"],
                     bg=COLOR_PANEL, fg=COLOR_TEXT,
                     font=("Segoe UI Emoji", 14)).pack(side="left", padx=(0, 8))
            tk.Label(titulo_frame, text=p["titulo"],
                     bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                     font=("Verdana", 13, "bold"),
                     anchor="w").pack(side="left", fill="x")

            # Separador
            tk.Frame(corpo, bg="#dde3ea", height=1).pack(fill="x", pady=(8, 0))

            # Corpo do texto
            tk.Label(corpo, text=p["texto"],
                     bg=COLOR_PANEL, fg=COLOR_TEXT,
                     font=("Verdana", 9),
                     justify="left", wraplength=380,
                     anchor="nw").pack(fill="x", pady=(10, 12))

            # Rodapé: "Pular" | dots | "Próximo / Concluir"
            rodape = tk.Frame(corpo, bg=COLOR_PANEL)
            rodape.pack(fill="x")

            def _pular():
                _marcar_e_fechar()

            def _proximo():
                if estado["passo"] < total - 1:
                    estado["passo"] += 1
                    _criar_card(estado["passo"])
                else:
                    _marcar_e_fechar()

            def _marcar_e_fechar():
                self._marcar_tutorial_visto()
                if estado["janela"] and estado["janela"].winfo_exists():
                    estado["janela"].destroy()

            # Botão "Pular tutorial"
            tk.Button(
                rodape, text="Pular tutorial",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                font=("Verdana", 8), relief="flat", cursor="hand2",
                bd=0, activeforeground=COLOR_BCB_BLUE,
                command=_pular
            ).pack(side="left")

            # Dots de progresso (centro)
            dots_frame = tk.Frame(rodape, bg=COLOR_PANEL)
            dots_frame.pack(side="left", expand=True)
            for i in range(total):
                cor = COLOR_BCB_BLUE if i == passo_idx else "#c8d0d8"
                tk.Label(dots_frame, text="●", bg=COLOR_PANEL,
                         fg=cor, font=("Verdana", 8)).pack(side="left", padx=1)

            # Botão "Próximo / Concluir"
            btn_txt = "Concluir ✓" if passo_idx == total - 1 else "Próximo ›"
            btn_bg = "#006400" if passo_idx == total - 1 else COLOR_BCB_BLUE
            tk.Button(
                rodape, text=btn_txt,
                bg=btn_bg, fg="white",
                font=("Verdana", 9, "bold"),
                relief="flat", cursor="hand2",
                padx=14, pady=5,
                activebackground=COLOR_BCB_BLUE_DARK,
                activeforeground="white",
                command=_proximo
            ).pack(side="right")

            # Posiciona o card no centro-inferior da janela principal
            self.update_idletasks()
            card.update_idletasks()
            w_card = card.winfo_reqwidth()
            h_card = card.winfo_reqheight()
            w_app  = self.winfo_width()
            h_app  = self.winfo_height()
            x_app  = self.winfo_rootx()
            y_app  = self.winfo_rooty()
            # Centra horizontalmente, posiciona um pouco abaixo do centro
            cx = x_app + (w_app - w_card) // 2
            cy = y_app + int(h_app * 0.20)
            card.geometry(f"+{cx}+{cy}")

            # Permite fechar com ESC
            card.bind("<Escape>", lambda e: _marcar_e_fechar())

        # Inicia no passo 0
        _criar_card(0)

    def _abrir_tutorial_manual(self):
        """Reabre o tutorial a qualquer momento (chamado pelo menu Ajuda)."""
        self._mostrar_tutorial()

    # ---------- Controle de tamanho da interface (Pequeno/Médio/Grande) -----
    #
    # Em vez de tentar adivinhar o tamanho ideal automaticamente (o que dá
    # errado com dois monitores de zoom diferente), o usuário escolhe o
    # tamanho num menu visível, e a escolha é lembrada entre sessões.
    #
    # Os valores são MULTIPLICADORES aplicados ao tamanho ORIGINAL das
    # fontes de cada widget (1.0 = tamanho de fábrica).
    _TAMANHOS = {
        "auto":    1.00,   # tamanho original do app
        "pequeno": 0.85,
        "medio":   1.15,
        "grande":  1.35,
    }

    def _arquivo_config(self):
        """Caminho do arquivo de configuração (na pasta do usuário)."""
        try:
            base = os.path.expanduser("~")
            return os.path.join(base, ".calculadora_bcb_config.json")
        except Exception:
            return None

    def _carregar_pref_tamanho(self):
        path = self._arquivo_config()
        if path and os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                t = cfg.get("tamanho")
                if t in self._TAMANHOS:
                    return t
            except Exception:
                pass
        return "auto"

    def _salvar_pref_tamanho(self, tamanho):
        path = self._arquivo_config()
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"tamanho": tamanho}, f)
        except Exception:
            pass

    def _aplicar_tamanho_inicial(self):
        self._tamanho_atual = self._carregar_pref_tamanho()
        if hasattr(self, "tamanho_var"):
            self.tamanho_var.set(self._tamanho_atual)
        # Aplica após a UI estar montada (garante que todos os widgets existem)
        self.after(80, lambda: self._aplicar_tamanho(self._tamanho_atual, salvar=False))

    def _coletar_fontes_originais(self, widget=None, apenas_novos=False):
        """Percorre os widgets e guarda o tamanho de fonte 'base' de cada um.
        Com apenas_novos=True, registra só widgets ainda não vistos (úteis
        para resultados/linhas criados após a 1ª aplicação)."""
        if not hasattr(self, "_fontes_orig"):
            self._fontes_orig = {}
        if widget is None:
            widget = self
        chave = str(widget)
        if not (apenas_novos and chave in self._fontes_orig):
            try:
                f = widget.cget("font")
            except Exception:
                f = None
            if f:
                try:
                    partes = self.tk.splitlist(self.tk.call("font", "actual", f))
                    d = {}
                    it = iter(partes)
                    for k in it:
                        v = next(it, "")
                        d[str(k)] = v
                    familia = d.get("-family", "Verdana")
                    tam = int(float(d.get("-size", "9")))
                    weight = d.get("-weight", "normal")
                    slant = d.get("-slant", "roman")
                    # Se já há um tamanho aplicado, dividimos para achar o base
                    mult_atual = self._TAMANHOS.get(
                        getattr(self, "_tamanho_atual", "auto"), 1.0)
                    base = abs(tam) / mult_atual if mult_atual else abs(tam)
                    self._fontes_orig[chave] = {
                        "widget": widget, "familia": familia,
                        "tam": max(6, int(round(base))) if tam else 9,
                        "neg": tam < 0, "weight": weight, "slant": slant,
                    }
                except Exception:
                    pass
        for c in widget.winfo_children():
            self._coletar_fontes_originais(c, apenas_novos=apenas_novos)

    def _aplicar_tamanho(self, tamanho, salvar=True):
        if tamanho not in self._TAMANHOS:
            tamanho = "auto"
        mult = self._TAMANHOS[tamanho]

        # 1ª vez: coleta todas as fontes originais (com o app ainda em 'auto').
        # Depois: registra apenas widgets novos (resultados, parcelas, etc.).
        if not hasattr(self, "_fontes_orig"):
            self._coletar_fontes_originais(self)
        else:
            self._coletar_fontes_originais(self, apenas_novos=True)

        self._tamanho_atual = tamanho

        for info in list(self._fontes_orig.values()):
            w = info["widget"]
            try:
                if not int(w.winfo_exists()):
                    continue
            except Exception:
                continue
            novo_tam = max(6, int(round(info["tam"] * mult)))
            if info["neg"]:
                novo_tam = -novo_tam
            spec = [info["familia"], novo_tam]
            if info["weight"] == "bold":
                spec.append("bold")
            if info["slant"] == "italic":
                spec.append("italic")
            try:
                w.configure(font=tuple(spec))
            except Exception:
                pass

        if hasattr(self, "tamanho_var"):
            try:
                self.tamanho_var.set(tamanho)
            except Exception:
                pass
        if salvar:
            self._salvar_pref_tamanho(tamanho)

    def reaplicar_tamanho(self):
        """Reaplica o tamanho atual — chamar após criar novos widgets
        (ex.: render de resultado) para que eles também sejam reescalados."""
        if getattr(self, "_tamanho_atual", "auto") != "auto":
            self._aplicar_tamanho(self._tamanho_atual, salvar=False)

    def _build_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure("BCB.TNotebook", background=COLOR_BG, borderwidth=0)
        style.configure(
            "BCB.TNotebook.Tab",
            background="#dcdcdc", foreground=COLOR_BCB_BLUE,
            padding=(16, 8), font=("Verdana", 9, "bold"),
        )
        style.map(
            "BCB.TNotebook.Tab",
            background=[("selected", COLOR_PANEL)],
            foreground=[("selected", COLOR_BCB_BLUE)],
            expand=[("selected", (1, 1, 1, 0))],
        )

        style.configure("BCB.TFrame", background=COLOR_PANEL)
        style.configure(
            "FieldLabel.TLabel",
            background=COLOR_PANEL, foreground=COLOR_BCB_BLUE,
            font=("Verdana", 9, "bold"),
        )
        style.configure(
            "Info.TLabel",
            background="#fffbeb", foreground="#5d4a10",
            font=("Verdana", 8), padding=(8, 6),
        )
        style.configure(
            "BCB.TButton",
            background=COLOR_BCB_BLUE, foreground="white",
            font=("Verdana", 9, "bold"), padding=(16, 6), borderwidth=0,
        )
        style.map("BCB.TButton",
                  background=[("active", COLOR_BCB_BLUE_LIGHT), ("disabled", "#999")])
        style.configure(
            "BCBSecondary.TButton",
            background="#cccccc", foreground="#333333",
            font=("Verdana", 9, "bold"), padding=(16, 6), borderwidth=0,
        )
        style.map("BCBSecondary.TButton", background=[("active", "#bbbbbb")])
        style.configure(
            "BCBSmall.TButton",
            background="#dde7f0", foreground=COLOR_BCB_BLUE,
            font=("Verdana", 8, "bold"), padding=(8, 3), borderwidth=0,
        )
        style.map("BCBSmall.TButton", background=[("active", "#c5d4e2")])

        # Treeview
        style.configure(
            "BCB.Treeview",
            background="white", fieldbackground="white",
            foreground=COLOR_TEXT, rowheight=22,
            font=("Verdana", 8),
        )
        style.configure(
            "BCB.Treeview.Heading",
            background=COLOR_BCB_BLUE, foreground="white",
            font=("Verdana", 8, "bold"), padding=(4, 4),
        )
        style.map("BCB.Treeview.Heading", background=[("active", COLOR_BCB_BLUE_LIGHT)])

    # ------- UI principal ------- #
    def _build_ui(self):
        # Header
        header = tk.Frame(self, bg=COLOR_HEADER_BG, height=72)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        logo = tk.Canvas(header, width=56, height=56, bg=COLOR_HEADER_BG,
                        highlightthickness=0)
        logo.create_rectangle(2, 2, 54, 54, fill=COLOR_BCB_BLUE,
                             outline=COLOR_BCB_BLUE_DARK, width=2)
        logo.create_text(28, 22, text="BCB", fill="white", font=("Verdana", 12, "bold"))
        logo.create_text(28, 40, text="BANCO\nCENTRAL", fill="white",
                        font=("Verdana", 5), justify="center")
        logo.pack(side="left", padx=(16, 12), pady=8)

        title_block = tk.Frame(header, bg=COLOR_HEADER_BG)
        title_block.pack(side="left", fill="y", pady=10)
        tk.Label(title_block, text="Calculadora do Cidadão",
                bg=COLOR_HEADER_BG, fg=COLOR_BCB_BLUE,
                font=("Verdana", 15, "bold")).pack(anchor="w")
        tk.Label(title_block,
                text=f"Banco Central do Brasil · Correção de Valores · v{APP_VERSION}",
                bg=COLOR_HEADER_BG, fg=COLOR_SUBTLE,
                font=("Verdana", 7)).pack(anchor="w")

        # Badge BETA + botão tutorial
        beta_frame = tk.Frame(header, bg=COLOR_HEADER_BG)
        beta_frame.pack(side="right", padx=(0, 10), pady=14)
        tk.Label(beta_frame,
                 text=" ⚠  BETA TEST ",
                 bg="#c0392b", fg="white",
                 font=("Verdana", 8, "bold"),
                 padx=4, pady=2).pack(anchor="e")
        tk.Button(beta_frame,
                  text="❓ Ver Tutorial",
                  bg=COLOR_HEADER_BG, fg=COLOR_BCB_BLUE,
                  font=("Verdana", 7, "bold"),
                  relief="flat", cursor="hand2", bd=0,
                  activeforeground=COLOR_BCB_BLUE_DARK,
                  command=self._abrir_tutorial_manual).pack(anchor="e", pady=(4, 0))

        self.clock_var = tk.StringVar()
        clock_frame = tk.Frame(header, bg=COLOR_HEADER_BG)
        clock_frame.pack(side="right", padx=18, pady=14)
        tk.Label(clock_frame, text="Acesso público",
                bg=COLOR_HEADER_BG, fg=COLOR_BCB_BLUE,
                font=("Verdana", 7, "bold")).pack(anchor="e")
        tk.Label(clock_frame, textvariable=self.clock_var,
                bg=COLOR_HEADER_BG, fg=COLOR_SUBTLE,
                font=("Verdana", 7)).pack(anchor="e")

        tk.Frame(self, bg=COLOR_BCB_BLUE, height=3).pack(fill="x")

        # Breadcrumb
        bc = tk.Frame(self, bg="#f0f0f0", height=26)
        bc.pack(fill="x")
        bc.pack_propagate(False)
        tk.Label(bc, text="  Início › Calculadora do Cidadão › Correção de valores",
                bg="#f0f0f0", fg=COLOR_SUBTLE, font=("Verdana", 7),
                anchor="w").pack(side="left", fill="x", expand=True)

        # Seletor de tamanho da interface (lembrado entre sessões)
        self.tamanho_var = tk.StringVar(value=getattr(self, "_tamanho_atual", "auto"))
        tam_frame = tk.Frame(bc, bg="#f0f0f0")
        tam_frame.pack(side="right", padx=(0, 10))
        tk.Label(tam_frame, text="Tamanho:", bg="#f0f0f0", fg=COLOR_SUBTLE,
                font=("Verdana", 7, "bold")).pack(side="left", padx=(0, 4))
        for rotulo, valor in [("Automático", "auto"), ("Pequeno", "pequeno"),
                             ("Médio", "medio"), ("Grande", "grande")]:
            tk.Radiobutton(
                tam_frame, text=rotulo, value=valor,
                variable=self.tamanho_var,
                command=lambda v=valor: self._aplicar_tamanho(v),
                bg="#f0f0f0", fg=COLOR_BCB_BLUE, font=("Verdana", 7),
                activebackground="#f0f0f0", selectcolor="white",
                indicatoron=True, padx=2).pack(side="left")

        tk.Label(bc, text="CALFW0301  ", bg="#f0f0f0", fg=COLOR_SUBTLE,
                font=("Verdana", 7)).pack(side="right")

        # Container + Notebook
        outer = tk.Frame(self, bg=COLOR_BG)
        outer.pack(fill="both", expand=True, padx=10, pady=8)

        panel = tk.Frame(outer, bg=COLOR_PANEL, highlightthickness=1,
                        highlightbackground="#b8b8b8")
        panel.pack(fill="both", expand=True)

        tk.Label(panel, text="Correção de valores",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 10, "bold"),
                padx=10, pady=6, anchor="w").pack(fill="x")

        self.notebook = ttk.Notebook(panel, style="BCB.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=6, pady=(4, 6))

        self._build_indices_tab()
        self._build_selic_tab()
        self._build_lote_tab()
        self._build_demo_tab()
        self._build_cobranca_tab()
        self._build_atraso_tab()

        # Footer
        footer = tk.Frame(self, bg=COLOR_BG)
        footer.pack(fill="x", pady=(0, 6))
        tk.Label(
            footer,
            text=("Reprodução não-oficial. Para cálculos oficiais consulte "
                  "www3.bcb.gov.br/CALCIDADAO  ·  Dados via api.bcb.gov.br/dados/serie"),
            bg=COLOR_BG, fg=COLOR_SUBTLE, font=("Verdana", 7),
        ).pack()

    # ----------------------------------------------------------------
    # Helpers visuais
    # ----------------------------------------------------------------
    def _clear_frame(self, frame):
        for w in frame.winfo_children():
            w.destroy()

    def _show_loading(self, frame, msg="Consultando dados do Banco Central..."):
        self._clear_frame(frame)
        box = tk.Frame(frame, bg="#f0f4f8", highlightbackground="#99aab8",
                      highlightthickness=1)
        box.pack(fill="x")
        self._loading_label = tk.Label(box, text="⟳  " + msg, bg="#f0f4f8",
                                       fg=COLOR_BCB_BLUE,
                                       font=("Verdana", 9, "italic"), pady=12)
        self._loading_label.pack()

    def _update_loading(self, frame, msg):
        try:
            self._loading_label.config(text="⟳  " + msg)
        except (AttributeError, tk.TclError):
            pass

    def _show_error(self, frame, msg):
        self._clear_frame(frame)
        box = tk.Frame(frame, bg="#fdecea", highlightbackground=COLOR_ERROR,
                      highlightthickness=1)
        box.pack(fill="x")
        tk.Label(box, text="⚠  " + msg, bg="#fdecea", fg="#8a1f1c",
                font=("Verdana", 9), wraplength=900, justify="left",
                padx=12, pady=10).pack(anchor="w")

    def _show_result_simple(self, frame, title, rows, highlight_row=None):
        self._clear_frame(frame)
        box = tk.Frame(frame, bg=COLOR_RESULT_BG,
                      highlightbackground=COLOR_BCB_BLUE, highlightthickness=2)
        box.pack(fill="x")

        head = tk.Frame(box, bg=COLOR_RESULT_BG)
        head.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(head, text="✓", bg=COLOR_RESULT_OK, fg="white",
                font=("Verdana", 10, "bold"), width=2).pack(side="left", padx=(0, 8))
        tk.Label(head, text=title, bg=COLOR_RESULT_BG, fg=COLOR_BCB_BLUE,
                font=("Verdana", 11, "bold")).pack(side="left")
        ttk.Separator(box, orient="horizontal").pack(fill="x", padx=14, pady=(2, 8))

        table = tk.Frame(box, bg=COLOR_RESULT_BG)
        table.pack(fill="x", padx=14, pady=(0, 14))
        table.columnconfigure(0, weight=1)
        table.columnconfigure(1, weight=1)

        for i, (label, value) in enumerate(rows):
            is_highlight = (highlight_row == i)
            row_bg = "#fff8dc" if is_highlight else COLOR_RESULT_BG
            val_fg = COLOR_RESULT_OK if is_highlight else COLOR_BCB_BLUE
            val_font = ("Consolas", 13, "bold") if is_highlight else ("Consolas", 10, "bold")

            tk.Label(table, text=label, bg=row_bg, fg=COLOR_SUBTLE,
                    font=("Verdana", 9), anchor="w").grid(
                row=i, column=0, sticky="ew", padx=6, pady=3)
            tk.Label(table, text=value, bg=row_bg, fg=val_fg,
                    font=val_font, anchor="e").grid(
                row=i, column=1, sticky="ew", padx=6, pady=3)

    def _field_label(self, parent, row, text, required=False, hint=None, col=0):
        cell = tk.Frame(parent, bg=COLOR_PANEL)
        cell.grid(row=row, column=col, sticky="ne", padx=(0, 4), pady=8)
        line = tk.Frame(cell, bg=COLOR_PANEL)
        line.pack(anchor="e")
        if required:
            tk.Label(line, text="* ", bg=COLOR_PANEL, fg=COLOR_ERROR,
                    font=("Verdana", 9, "bold")).pack(side="left")
        tk.Label(line, text=text, bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9, "bold")).pack(side="left")
        if hint:
            tk.Label(cell, text=hint, bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                    font=("Verdana", 7)).pack(anchor="e")

    def _build_processo_section(self, parent, scope):
        """Cria a seção colapsável 'Dados do Processo (cabeçalho do relatório)'
        que adiciona o cabeçalho institucional do IPREM + dados preenchíveis do
        servidor nos relatórios exportados (XLSX/PDF). `scope` é uma string
        única ('lote' ou 'demo') pra distinguir as duas instâncias da seção."""
        if not hasattr(self, "_processo_entries"):
            self._processo_entries = {}

        frame = tk.LabelFrame(
            parent, text=" Dados do Processo (cabeçalho do relatório) ",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            font=("Verdana", 8, "bold"), bd=1, relief="solid")
        frame.pack(fill="x", pady=(0, 6))

        # Aviso
        tk.Label(
            frame,
            text="Estes campos são opcionais. Quando preenchidos, aparecem como "
            "cabeçalho no Excel e PDF exportados desta aba.",
            bg=COLOR_PANEL, fg="#9C7700", font=("Verdana", 7),
            wraplength=900, justify="left", anchor="w"
        ).pack(fill="x", padx=8, pady=(4, 2))

        inner = tk.Frame(frame, bg=COLOR_PANEL)
        inner.pack(fill="x", padx=8, pady=(4, 8))
        inner.columnconfigure(1, weight=2)
        inner.columnconfigure(3, weight=1)
        inner.columnconfigure(5, weight=1)

        # Linha 0: Servidor(a) | Registro Funcional | Data de Nascimento
        # Linha 1: RG | CPF
        # Linha 2: Órgão de Origem | Órgão Cessionário
        # Linha 3: Período de Licença | Processo SEI nº
        campos = [
            (0, 0, "Servidor(a):",          "servidor",            30),
            (0, 2, "Registro Funcional:",   "registro_funcional",  16),
            (0, 4, "Data de Nascimento:",   "data_nascimento",     14),
            (1, 0, "RG:",                   "rg",                  20),
            (1, 2, "CPF:",                  "cpf",                 16),
            (2, 0, "Órgão de Origem:",      "orgao_origem",        30),
            (2, 2, "Órgão Cessionário:",    "orgao_cessionario",   20),
            (3, 0, "Período de Licença:",   "periodo_afastamento", 22),
            (3, 2, "Processo SEI nº:",      "processo_sei",        20),
        ]
        entries = {}
        for linha, col, label, chave, width in campos:
            tk.Label(inner, text=label, bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                    font=("Verdana", 8, "bold"), anchor="e").grid(
                row=linha, column=col, sticky="e", padx=(2, 4), pady=3)
            e = _entry_make(inner, width=width, placeholder="")
            e.grid(row=linha, column=col + 1, sticky="we", padx=2, pady=3)
            entries[chave] = e

        self._processo_entries[scope] = entries
        return frame

    def _coletar_dados_processo(self, scope):
        """Retorna dict com os 8 campos preenchidos para o scope. Se todos
        estiverem vazios, retorna None (assim a exportação omite o cabeçalho)."""
        ents = getattr(self, "_processo_entries", {}).get(scope)
        if not ents:
            return None
        dados = {chave: _entry_value(e).strip() for chave, e in ents.items()}
        if not any(dados.values()):
            return None
        return dados

    # ----------------------------------------------------------------
    # Tab: Índices (correção simples por mês)
    # ----------------------------------------------------------------
    def _build_indices_tab(self):
        tab = ttk.Frame(self.notebook, style="BCB.TFrame", padding=14)
        self.notebook.add(tab, text="Índices de preços")

        ttk.Label(tab, text=" Os campos com  *  são de preenchimento obrigatório.",
                 style="Info.TLabel").pack(fill="x", pady=(0, 12))

        tk.Label(tab, text="Correção de valor por índices de preços",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 10, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Separator(tab, orient="horizontal").pack(fill="x", pady=(0, 14))

        form = tk.Frame(tab, bg=COLOR_PANEL)
        form.pack(fill="x")
        form.columnconfigure(1, weight=1)

        self._field_label(form, 0, "Selecione o índice", required=True)
        opts = [
            ("IGP-M (FGV) — a partir de 06/1989", "IGP-M"),
            ("IGP-DI (FGV) — a partir de 02/1944", "IGP-DI"),
            ("INPC (IBGE) — a partir de 04/1979", "INPC"),
            ("IPCA (IBGE) — a partir de 01/1980", "IPCA"),
            ("IPCA-E (IBGE) — a partir de 01/1992", "IPCA-E"),
            ("IPC-BRASIL (FGV) — a partir de 01/1990", "IPC-BR"),
            ("IPC-SP (FIPE) — a partir de 11/1942", "IPC-SP"),
        ]
        self._indice_map = {label: key for label, key in opts}
        self.cb_indice = ttk.Combobox(form, values=[label for label, _ in opts],
                                      state="readonly", width=42)
        self.cb_indice.current(3)
        self.cb_indice.grid(row=0, column=1, sticky="w", padx=8, pady=6)

        self._field_label(form, 1, "Data inicial (MM/AAAA)", required=True,
                         hint="(inclui o mês inicial)")
        self.e_idx_di = _entry_make(form, width=12, placeholder="MM/AAAA")
        self.e_idx_di.grid(row=1, column=1, sticky="w", padx=8, pady=6)
        _mask_month_year(self.e_idx_di)

        self._field_label(form, 2, "Data final (MM/AAAA)", required=True)
        self.e_idx_df = _entry_make(form, width=12, placeholder="MM/AAAA")
        self.e_idx_df.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        _mask_month_year(self.e_idx_df)

        self._field_label(form, 3, "Valor a ser corrigido")
        self.e_idx_valor = _entry_make(form, width=18, placeholder="0,00")
        self.e_idx_valor.grid(row=3, column=1, sticky="w", padx=8, pady=6)
        _bind_valor_format(self.e_idx_valor)

        btn_bar = tk.Frame(tab, bg=COLOR_PANEL, pady=10)
        btn_bar.pack(fill="x", pady=(10, 0))
        self.btn_idx = ttk.Button(btn_bar, text="Corrigir valor",
                                  style="BCB.TButton",
                                  command=self._on_calc_indice)
        self.btn_idx.pack(side="left", padx=(0, 6))
        ttk.Button(btn_bar, text="Limpar", style="BCBSecondary.TButton",
                  command=self._on_clear_indice).pack(side="left")

        self.idx_result_frame = tk.Frame(tab, bg=COLOR_PANEL)
        self.idx_result_frame.pack(fill="both", expand=True, pady=(14, 0))

        # Enter
        for w in (self.e_idx_di, self.e_idx_df, self.e_idx_valor):
            w.bind("<Return>", lambda _e: self._on_calc_indice())

    def _on_calc_indice(self):
        di = _entry_value(self.e_idx_di).strip()
        df = _entry_value(self.e_idx_df).strip()
        v_raw = _entry_value(self.e_idx_valor).strip()
        indice_key = self._indice_map.get(self.cb_indice.get(), "IPCA")

        ini = parse_month_year(di)
        fim = parse_month_year(df)
        if not ini:
            return self._show_error(self.idx_result_frame, "Data inicial inválida. Use MM/AAAA.")
        if not fim:
            return self._show_error(self.idx_result_frame, "Data final inválida. Use MM/AAAA.")
        try:
            valor = parse_valor_br(v_raw) if v_raw else None
        except ValueError:
            return self._show_error(self.idx_result_frame,
                                   "Valor inválido. Use formato 1.234,56.")

        self.btn_idx.config(state="disabled")
        self._show_loading(self.idx_result_frame)

        # Limpar cache da API pra forçar nova consulta
        limpar_cache_api()

        def worker():
            try:
                res = calcular_indice(indice_key, ini, fim, valor)
                self.after(0, lambda: self._render_indice_result(res))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: self._show_error(self.idx_result_frame, err))
            finally:
                self.after(0, lambda: self.btn_idx.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _render_indice_result(self, res):
        rows = [
            ("Período:", res["periodo"]),
            ("Meses incluídos no cálculo:", str(res["meses"])),
            ("Índice de correção no período:", fmt_fator(res["fator"])),
            ("Variação acumulada no período:", fmt_percent(res["variacao"]) + " %"),
        ]
        highlight = None
        if res["valor_informado"] is not None:
            rows.append(("Valor informado:", "R$ " + fmt_brl(res["valor_informado"])))
            rows.append(("Valor corrigido:", "R$ " + fmt_brl(res["valor_corrigido"])))
            highlight = len(rows) - 1
        self._show_result_simple(
            self.idx_result_frame,
            f"Resultado da correção pelo {res['indice_nome']}",
            rows, highlight_row=highlight)

    def _on_clear_indice(self):
        for e, ph in [(self.e_idx_di, "MM/AAAA"),
                     (self.e_idx_df, "MM/AAAA"),
                     (self.e_idx_valor, "0,00")]:
            _entry_set(e, "")
        self._clear_frame(self.idx_result_frame)

    # ----------------------------------------------------------------
    # Tab: Selic
    # ----------------------------------------------------------------
    def _build_selic_tab(self):
        tab = ttk.Frame(self.notebook, style="BCB.TFrame", padding=14)
        self.notebook.add(tab, text="Selic")

        ttk.Label(tab,
                 text=(" Obs.: Para a Selic, períodos a partir de 04/06/1986. "
                       "Campos com  *  são obrigatórios."),
                 style="Info.TLabel").pack(fill="x", pady=(0, 12))

        tk.Label(tab, text="Correção de valor pela taxa Selic",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 10, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Separator(tab, orient="horizontal").pack(fill="x", pady=(0, 14))

        form = tk.Frame(tab, bg=COLOR_PANEL)
        form.pack(fill="x")
        form.columnconfigure(1, weight=1)

        self._field_label(form, 0, "Data inicial (DD/MM/AAAA)", required=True)
        self.e_sel_di = _entry_make(form, width=14, placeholder="DD/MM/AAAA")
        self.e_sel_di.grid(row=0, column=1, sticky="w", padx=8, pady=6)
        _mask_full_date(self.e_sel_di)

        self._field_label(form, 1, "Data final (DD/MM/AAAA)", required=True)
        self.e_sel_df = _entry_make(form, width=14, placeholder="DD/MM/AAAA")
        self.e_sel_df.grid(row=1, column=1, sticky="w", padx=8, pady=6)
        _mask_full_date(self.e_sel_df)

        self._field_label(form, 2, "Valor a ser corrigido")
        self.e_sel_valor = _entry_make(form, width=18, placeholder="0,00")
        self.e_sel_valor.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        _bind_valor_format(self.e_sel_valor)

        btn_bar = tk.Frame(tab, bg=COLOR_PANEL, pady=10)
        btn_bar.pack(fill="x", pady=(10, 0))
        self.btn_sel = ttk.Button(btn_bar, text="Corrigir valor",
                                  style="BCB.TButton",
                                  command=self._on_calc_selic)
        self.btn_sel.pack(side="left", padx=(0, 6))
        ttk.Button(btn_bar, text="Limpar", style="BCBSecondary.TButton",
                  command=self._on_clear_selic).pack(side="left")

        self.sel_result_frame = tk.Frame(tab, bg=COLOR_PANEL)
        self.sel_result_frame.pack(fill="both", expand=True, pady=(14, 0))

        for w in (self.e_sel_di, self.e_sel_df, self.e_sel_valor):
            w.bind("<Return>", lambda _e: self._on_calc_selic())

    def _on_calc_selic(self):
        di = _entry_value(self.e_sel_di).strip()
        df = _entry_value(self.e_sel_df).strip()
        v_raw = _entry_value(self.e_sel_valor).strip()
        ini = parse_date_br(di)
        fim = parse_date_br(df)
        if not ini:
            return self._show_error(self.sel_result_frame, "Data inicial inválida. Use DD/MM/AAAA.")
        if not fim:
            return self._show_error(self.sel_result_frame, "Data final inválida. Use DD/MM/AAAA.")
        try:
            valor = parse_valor_br(v_raw) if v_raw else None
        except ValueError:
            return self._show_error(self.sel_result_frame, "Valor inválido.")

        self.btn_sel.config(state="disabled")
        self._show_loading(self.sel_result_frame, "Consultando série Selic diária no BCB...")

        # Limpar cache da API pra forçar nova consulta
        limpar_cache_api()

        def worker():
            try:
                res = calcular_selic(ini, fim, valor)
                self.after(0, lambda: self._render_selic_result(res))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: self._show_error(self.sel_result_frame, err))
            finally:
                self.after(0, lambda: self.btn_sel.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _render_selic_result(self, res):
        rows = [
            ("Período:", res["periodo"]),
            ("Dias úteis considerados:", str(res["meses"])),
            ("Fator acumulado:", fmt_fator(res["fator"])),
            ("Variação acumulada:", fmt_percent(res["variacao"]) + " %"),
        ]
        highlight = None
        if res["valor_informado"] is not None:
            rows.append(("Valor informado:", "R$ " + fmt_brl(res["valor_informado"])))
            rows.append(("Valor corrigido:", "R$ " + fmt_brl(res["valor_corrigido"])))
            highlight = len(rows) - 1
        self._show_result_simple(
            self.sel_result_frame,
            "Resultado da correção pela Taxa Selic",
            rows, highlight_row=highlight)

    def _on_clear_selic(self):
        for e, ph in [(self.e_sel_di, "DD/MM/AAAA"),
                     (self.e_sel_df, "DD/MM/AAAA"),
                     (self.e_sel_valor, "0,00")]:
            _entry_set(e, "")
        self._clear_frame(self.sel_result_frame)

    # ----------------------------------------------------------------
    # Tab: Clock
    # ----------------------------------------------------------------
    def _update_clock(self):
        now = datetime.now()
        weekdays = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        meses_abrev = ["jan", "fev", "mar", "abr", "mai", "jun",
                       "jul", "ago", "set", "out", "nov", "dez"]
        self.clock_var.set(
            f"{weekdays[now.weekday()]} {now.day:02d} {meses_abrev[now.month-1]} "
            f"{now.year}  {now.strftime('%H:%M:%S')}"
        )
        self.after(1000, self._update_clock)

    # ================================================================
    # Tab: Lote Simples
    # ================================================================
    def _build_lote_tab(self):
        tab = ttk.Frame(self.notebook, style="BCB.TFrame", padding=10)
        self.notebook.add(tab, text="Lote (várias correções)")

        ttk.Label(
            tab,
            text=(" Calcula várias correções de uma só vez. Para Selic, usa o "
                  "primeiro dia do mês inicial até o último dia do mês final."),
            style="Info.TLabel"
        ).pack(fill="x", pady=(0, 8))

        # Barra de botões superior
        top_bar = tk.Frame(tab, bg=COLOR_PANEL)
        top_bar.pack(fill="x", pady=(0, 6))

        ttk.Button(top_bar, text="+ Adicionar linha",
                  style="BCBSmall.TButton",
                  command=self._lote_add_row).pack(side="left", padx=(0, 4))
        ttk.Button(top_bar, text="Importar CSV…",
                  style="BCBSmall.TButton",
                  command=lambda: self._lote_import("csv")).pack(side="left", padx=2)
        if HAS_XLSX:
            ttk.Button(top_bar, text="Importar XLSX…",
                      style="BCBSmall.TButton",
                      command=lambda: self._lote_import("xlsx")).pack(side="left", padx=2)
        ttk.Button(top_bar, text="Limpar tudo",
                  style="BCBSmall.TButton",
                  command=self._lote_clear).pack(side="left", padx=(8, 2))

        self.btn_lote_calc = ttk.Button(top_bar, text="▶ Calcular tudo",
                                        style="BCB.TButton",
                                        command=self._lote_calc)
        self.btn_lote_calc.pack(side="right", padx=2)

        # Seção: Dados do Processo (cabeçalho do relatório)
        self._build_processo_section(tab, scope="lote")

        # Área scrollável com as linhas
        list_frame = tk.LabelFrame(tab, text=" Linhas a calcular ",
                                  bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                  font=("Verdana", 8, "bold"),
                                  bd=1, relief="solid")
        list_frame.pack(fill="x", pady=(4, 6))

        # cabeçalho da tabela
        hdr = tk.Frame(list_frame, bg=COLOR_BCB_BLUE)
        hdr.pack(fill="x")
        cols = [("#", 30), ("Descrição", 200), ("Índice", 100),
                ("Data Inicial (MM/AAAA)", 150), ("Data Final (MM/AAAA)", 150),
                ("Valor (opcional)", 130), ("", 50)]
        for txt, w in cols:
            tk.Label(hdr, text=txt, bg=COLOR_BCB_BLUE, fg="white",
                    font=("Verdana", 8, "bold"), width=max(w//8, 4),
                    padx=4, pady=4).pack(side="left", padx=1)

        # canvas scrollável
        canvas_frame = tk.Frame(list_frame, bg=COLOR_PANEL)
        canvas_frame.pack(fill="x")

        self.lote_canvas = tk.Canvas(canvas_frame, bg=COLOR_PANEL,
                                    height=180, highlightthickness=0)
        self.lote_canvas.pack(side="left", fill="x", expand=True)
        vsb = ttk.Scrollbar(canvas_frame, orient="vertical",
                           command=self.lote_canvas.yview)
        vsb.pack(side="right", fill="y")
        self.lote_canvas.configure(yscrollcommand=vsb.set)

        self.lote_rows_frame = tk.Frame(self.lote_canvas, bg=COLOR_PANEL)
        self.lote_canvas.create_window((0, 0), window=self.lote_rows_frame,
                                       anchor="nw")
        self.lote_rows_frame.bind(
            "<Configure>",
            lambda e: self.lote_canvas.configure(scrollregion=self.lote_canvas.bbox("all"))
        )

        # mouse wheel
        def _on_mw(e):
            self.lote_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self.lote_canvas.bind("<MouseWheel>", _on_mw)
        self.lote_rows_frame.bind("<MouseWheel>", _on_mw)

        self.lote_rows = []  # list of dicts {desc, indice, di, df, val, btn_rm, frame}

        # Adiciona 3 linhas de exemplo
        for _ in range(3):
            self._lote_add_row()

        # Resultado
        result_frame = tk.LabelFrame(tab, text=" Resultados ",
                                    bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                    font=("Verdana", 8, "bold"),
                                    bd=1, relief="solid")
        result_frame.pack(fill="both", expand=True, pady=(4, 0))

        toolbar = tk.Frame(result_frame, bg=COLOR_PANEL)
        toolbar.pack(fill="x", padx=4, pady=4)

        self.lbl_lote_status = tk.Label(toolbar, text="",
                                        bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                                        font=("Verdana", 8))
        self.lbl_lote_status.pack(side="left", padx=4)

        ttk.Button(toolbar, text="Exportar CSV…", style="BCBSmall.TButton",
                  command=lambda: self._lote_export("csv")).pack(side="right", padx=2)
        if HAS_XLSX:
            ttk.Button(toolbar, text="Exportar XLSX…", style="BCBSmall.TButton",
                      command=lambda: self._lote_export("xlsx")).pack(side="right", padx=2)
        if HAS_PDF:
            ttk.Button(toolbar, text="Exportar PDF…", style="BCBSmall.TButton",
                      command=lambda: self._lote_export("pdf")).pack(side="right", padx=2)
        tv_frame = tk.Frame(result_frame, bg=COLOR_PANEL)
        tv_frame.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        columns = ("desc", "indice", "periodo", "qtd", "fator",
                   "var", "valor_in", "valor_out", "status")
        self.lote_tree = ttk.Treeview(tv_frame, columns=columns,
                                      show="headings", style="BCB.Treeview",
                                      height=8)
        headings = {
            "desc": "Descrição", "indice": "Índice", "periodo": "Período",
            "qtd": "Meses/Dias", "fator": "Fator", "var": "Variação %",
            "valor_in": "Valor Informado", "valor_out": "Valor Corrigido",
            "status": "Status",
        }
        widths = {"desc": 160, "indice": 90, "periodo": 130, "qtd": 70,
                 "fator": 100, "var": 90, "valor_in": 110,
                 "valor_out": 130, "status": 90}
        for c in columns:
            self.lote_tree.heading(c, text=headings[c])
            self.lote_tree.column(c, width=widths[c],
                                  anchor="e" if c not in ("desc", "indice", "status") else "w")

        vsb2 = ttk.Scrollbar(tv_frame, orient="vertical",
                            command=self.lote_tree.yview)
        self.lote_tree.configure(yscrollcommand=vsb2.set)
        self.lote_tree.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

        self.lote_tree.tag_configure("ok", foreground=COLOR_BCB_BLUE)
        self.lote_tree.tag_configure("err", background="#fdecea",
                                    foreground="#8a1f1c")
        self.lote_tree.tag_configure("highlight", font=("Verdana", 8, "bold"))

        self.lote_resultados = []

    def _lote_add_row(self, descricao="", indice="IPCA", di="", df="", valor=""):
        rowf = tk.Frame(self.lote_rows_frame, bg=COLOR_PANEL)
        rowf.pack(fill="x", padx=2, pady=1)

        num = len(self.lote_rows) + 1
        lbl = tk.Label(rowf, text=str(num), bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                      font=("Verdana", 8), width=4)
        lbl.pack(side="left", padx=2)

        e_desc = _entry_make(rowf, width=22, placeholder="(opcional)")
        e_desc.pack(side="left", padx=1)
        if descricao:
            _entry_set(e_desc, descricao)

        cb_idx = ttk.Combobox(rowf, values=list(INDICES.keys()),
                             state="readonly", width=10)
        cb_idx.set(indice if indice in INDICES else "IPCA")
        cb_idx.pack(side="left", padx=1)

        e_di = _entry_make(rowf, width=16, placeholder="MM/AAAA")
        e_di.pack(side="left", padx=1)
        _mask_month_year(e_di)
        if di:
            _entry_set(e_di, di)

        e_df = _entry_make(rowf, width=16, placeholder="MM/AAAA")
        e_df.pack(side="left", padx=1)
        _mask_month_year(e_df)
        if df:
            _entry_set(e_df, df)

        e_val = _entry_make(rowf, width=14, placeholder="0,00")
        e_val.pack(side="left", padx=1)
        _bind_valor_format(e_val)
        if valor:
            _entry_set(e_val, valor)

        row_data = {
            "frame": rowf, "lbl": lbl,
            "desc": e_desc, "indice": cb_idx,
            "di": e_di, "df": e_df, "val": e_val,
        }

        btn_rm = tk.Button(
            rowf, text="✕", bg="#fdecea", fg=COLOR_ERROR,
            font=("Verdana", 8, "bold"), bd=0, padx=6, pady=1,
            command=lambda r=row_data: self._lote_remove_row(r),
            cursor="hand2",
        )
        btn_rm.pack(side="left", padx=4)
        row_data["btn_rm"] = btn_rm

        self.lote_rows.append(row_data)

    def _lote_remove_row(self, row_data):
        row_data["frame"].destroy()
        self.lote_rows.remove(row_data)
        # renumerar
        for i, r in enumerate(self.lote_rows, start=1):
            r["lbl"].config(text=str(i))

    def _lote_clear(self):
        if not messagebox.askyesno("Limpar", "Remover todas as linhas?"):
            return
        for r in list(self.lote_rows):
            r["frame"].destroy()
        self.lote_rows.clear()
        for item in self.lote_tree.get_children():
            self.lote_tree.delete(item)
        self.lote_resultados = []
        self.lbl_lote_status.config(text="")

    def _lote_import(self, fmt):
        if fmt == "csv":
            path = filedialog.askopenfilename(
                title="Importar CSV",
                filetypes=[("CSV", "*.csv"), ("Todos", "*.*")],
            )
        else:
            path = filedialog.askopenfilename(
                title="Importar XLSX",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            )
        if not path:
            return
        try:
            if fmt == "csv":
                linhas = importar_csv_lote(path)
            else:
                linhas = importar_xlsx_lote(path)
        except Exception as e:
            messagebox.showerror("Erro ao importar", str(e))
            return

        if not linhas:
            messagebox.showwarning(
                "Importação",
                "Nenhuma linha encontrada. Verifique se o arquivo tem cabeçalho "
                "com colunas tipo: Descrição, Índice, Data Inicial, Data Final, Valor.")
            return

        # limpar linhas atuais
        for r in list(self.lote_rows):
            r["frame"].destroy()
        self.lote_rows.clear()

        for linha in linhas:
            self._lote_add_row(
                descricao=linha.get("descricao", ""),
                indice=linha.get("indice", "IPCA"),
                di=linha.get("data_ini", ""),
                df=linha.get("data_fim", ""),
                valor=linha.get("valor", ""),
            )
        self.lbl_lote_status.config(
            text=f"✓ {len(linhas)} linha(s) importada(s) de {os.path.basename(path)}")

    def _lote_export(self, fmt):
        if not self.lote_resultados:
            messagebox.showwarning("Exportar", "Calcule os resultados primeiro.")
            return
        if fmt == "csv":
            path = filedialog.asksaveasfilename(
                title="Salvar CSV", defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile="lote-correcoes.csv",
            )
        elif fmt == "xlsx":
            path = filedialog.asksaveasfilename(
                title="Salvar XLSX", defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="lote-correcoes.xlsx",
            )
        elif fmt == "pdf":
            path = filedialog.asksaveasfilename(
                title="Salvar PDF", defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile="lote-correcoes.pdf",
            )
        else:
            return
        if not path:
            return
        try:
            dp = self._coletar_dados_processo("lote")
            if fmt == "csv":
                exportar_csv_lote(path, self.lote_resultados)
            elif fmt == "xlsx":
                exportar_xlsx_lote(path, self.lote_resultados, dados_processo=dp)
            elif fmt == "pdf":
                exportar_pdf_lote(path, self.lote_resultados, dados_processo=dp)
            messagebox.showinfo("Exportar",
                              f"Arquivo salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _lote_calc(self):
        # Validar e coletar linhas
        linhas_input = []
        for i, r in enumerate(self.lote_rows, start=1):
            desc = _entry_value(r["desc"]).strip()
            idx_key = r["indice"].get()
            di = _entry_value(r["di"]).strip()
            df = _entry_value(r["df"]).strip()
            val_raw = _entry_value(r["val"]).strip()

            # ignorar linhas totalmente vazias
            if not any([desc, di, df, val_raw]):
                continue

            ini = parse_month_year(di)
            fim = parse_month_year(df)
            if not ini or not fim:
                messagebox.showerror(
                    "Validação",
                    f"Linha {i}: data inválida. Use MM/AAAA.")
                return
            try:
                valor = parse_valor_br(val_raw) if val_raw else None
            except ValueError:
                messagebox.showerror(
                    "Validação", f"Linha {i}: valor inválido.")
                return

            linhas_input.append({
                "descricao": desc, "indice_key": idx_key,
                "mes_ini": ini[0], "ano_ini": ini[1],
                "mes_fim": fim[0], "ano_fim": fim[1],
                "valor": valor,
            })

        if not linhas_input:
            messagebox.showwarning("Lote", "Nenhuma linha preenchida para calcular.")
            return

        self.btn_lote_calc.config(state="disabled")
        self.lbl_lote_status.config(text="Calculando...")
        for item in self.lote_tree.get_children():
            self.lote_tree.delete(item)

        # Limpar cache da API pra forçar nova consulta
        limpar_cache_api()

        def progress(msg):
            self.after(0, lambda: self.lbl_lote_status.config(text=msg))

        def worker():
            try:
                resultados = calcular_lote(linhas_input, progress_cb=progress)
                self.after(0, lambda: self._render_lote_results(resultados))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: messagebox.showerror("Erro no cálculo", err))
            finally:
                self.after(0, lambda: self.btn_lote_calc.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _render_lote_results(self, resultados):
        self.lote_resultados = resultados
        for item in self.lote_tree.get_children():
            self.lote_tree.delete(item)

        total_corrigido = Decimal("0")
        ok_count = 0
        err_count = 0
        for r in resultados:
            if r["ok"]:
                ok_count += 1
                if r.get("valor_corrigido"):
                    total_corrigido += r["valor_corrigido"]
                self.lote_tree.insert("", "end", values=(
                    r.get("descricao", ""),
                    r.get("indice_key", ""),
                    r.get("periodo", ""),
                    r.get("meses", ""),
                    fmt_fator(r.get("fator")) if r.get("fator") else "",
                    fmt_percent(r.get("variacao")) + " %" if r.get("variacao") else "",
                    fmt_brl(r.get("valor_informado")) if r.get("valor_informado") else "—",
                    fmt_brl(r.get("valor_corrigido")) if r.get("valor_corrigido") else "—",
                    "OK",
                ), tags=("ok",))
            else:
                err_count += 1
                self.lote_tree.insert("", "end", values=(
                    r.get("descricao", ""),
                    r.get("indice_key", ""),
                    "", "", "", "", "", "",
                    "ERRO: " + (r.get("erro") or "")[:50],
                ), tags=("err",))

        # Linha de total
        if total_corrigido > 0:
            self.lote_tree.insert("", "end", values=(
                "TOTAL", "", "", "", "", "", "",
                fmt_brl(total_corrigido), "",
            ), tags=("highlight",))

        status = f"✓ {ok_count} cálculo(s) realizado(s)"
        if err_count:
            status += f"  ·  {err_count} com erro"
        if total_corrigido > 0:
            status += f"  ·  Total corrigido: R$ {fmt_brl(total_corrigido)}"
        self.lbl_lote_status.config(text=status)

    # ================================================================
    # Tab: Demonstrativo Previdenciário
    # ================================================================
    def _build_demo_tab(self):
        tab = ttk.Frame(self.notebook, style="BCB.TFrame", padding=10)
        self.notebook.add(tab, text="Demonstrativo Previdenciário")

        ttk.Label(
            tab,
            text=(" Calcula demonstrativo de contribuições previdenciárias com "
                  "correção monetária, juros de mora e (Selic) multa."),
            style="Info.TLabel"
        ).pack(fill="x", pady=(0, 6))

        # ── Dividir aba: inputs (cima) / resultado (baixo) ────────────────
        _demo_paned = ttk.PanedWindow(tab, orient="vertical")
        _demo_paned.pack(fill="both", expand=True)
        inp = tk.Frame(_demo_paned, bg=COLOR_PANEL)
        _demo_paned.add(inp, weight=0)
        res = tk.Frame(_demo_paned, bg=COLOR_PANEL)
        _demo_paned.add(res, weight=1)
        # Define posição inicial do sash após o widget ser renderizado
        tab.after(200, lambda p=_demo_paned: p.sashpos(0, 450))

        # Seção: Dados do Processo (cabeçalho do relatório)
        self._build_processo_section(inp, scope="demo")

        # === Configurações ===
        cfg_frame = tk.LabelFrame(inp, text=" Configurações ",
                                 bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                 font=("Verdana", 8, "bold"),
                                 bd=1, relief="solid")
        cfg_frame.pack(fill="x", pady=(0, 6))

        cfg_inner = tk.Frame(cfg_frame, bg=COLOR_PANEL)
        cfg_inner.pack(fill="x", padx=8, pady=8)

        # linha 1
        tk.Label(cfg_inner, text="Índice:", bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9, "bold")).grid(row=0, column=0, sticky="e", padx=4)
        self.cb_demo_indice = ttk.Combobox(
            cfg_inner, values=["IPCA", "IGP-M", "IGP-DI", "INPC",
                              "IPCA-E", "IPC-BR", "IPC-SP", "SELIC"],
            state="readonly", width=10
        )
        self.cb_demo_indice.set("IPCA")
        self.cb_demo_indice.grid(row=0, column=1, columnspan=2, sticky="w", padx=4)
        self.cb_demo_indice.bind("<<ComboboxSelected>>", self._demo_toggle_selic_fields)

        tk.Label(cfg_inner, text="Data de Atualização:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=3, sticky="e", padx=(20, 4))
        self.e_demo_dataatu = _entry_make(cfg_inner, width=14, placeholder="DD/MM/AAAA")
        self.e_demo_dataatu.grid(row=0, column=4, columnspan=2, sticky="w", padx=4)
        _mask_full_date(self.e_demo_dataatu)
        # default = hoje
        _entry_set(self.e_demo_dataatu, date.today().strftime("%d/%m/%Y"))

        tk.Label(cfg_inner, text="Dia do Vencimento:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=6, sticky="e", padx=(20, 4))
        self.e_demo_diavenc = _entry_make(cfg_inner, width=6, placeholder="5")
        self.e_demo_diavenc.grid(row=0, column=7, sticky="w", padx=4)
        _entry_set(self.e_demo_diavenc, "5")
        tk.Label(cfg_inner, text="(do mês seguinte)", bg=COLOR_PANEL,
                fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
            row=0, column=8, sticky="w")

        # linha 2 - alíquotas
        tk.Label(cfg_inner, text="Alíquota Segurado:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=1, column=0, sticky="e", padx=4, pady=(8, 0))
        self.e_demo_aliqseg = _entry_make(cfg_inner, width=8, placeholder="14")
        self.e_demo_aliqseg.grid(row=1, column=1, sticky="w", padx=4, pady=(8, 0))
        _entry_set(self.e_demo_aliqseg, "14")
        tk.Label(cfg_inner, text="%", bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                font=("Verdana", 8)).grid(row=1, column=2,
                sticky="w", padx=(0, 8), pady=(8, 0))

        tk.Label(cfg_inner, text="Alíquota Patronal:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=1, column=3, sticky="e", padx=(20, 4), pady=(8, 0))
        self.e_demo_aliqpat = _entry_make(cfg_inner, width=8, placeholder="28")
        self.e_demo_aliqpat.grid(row=1, column=4, sticky="w", padx=4, pady=(8, 0))
        _entry_set(self.e_demo_aliqpat, "28")
        tk.Label(cfg_inner, text="%", bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                font=("Verdana", 8)).grid(row=1, column=5,
                sticky="w", padx=(0, 8), pady=(8, 0))

        # linha 3 - juros mensais
        tk.Label(cfg_inner, text="Juros Mensais:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=2, column=0, sticky="e", padx=4, pady=(8, 0))
        self.e_demo_juros_mes = _entry_make(cfg_inner, width=8, placeholder="1")
        self.e_demo_juros_mes.grid(row=2, column=1, sticky="w", padx=4, pady=(8, 0))
        _entry_set(self.e_demo_juros_mes, "1")
        tk.Label(cfg_inner, text="% a.m.", bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                font=("Verdana", 8)).grid(row=2, column=2,
                sticky="w", padx=(0, 8), pady=(8, 0))
        tk.Label(cfg_inner,
                text="(IPCA: × meses de atraso. Selic: 1% no mês do pagamento)",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
                row=2, column=3, columnspan=4, sticky="w",
                padx=(20, 4), pady=(8, 0))

        # linha 4 - multa Selic (diária + limite)
        tk.Label(cfg_inner, text="Multa Diária:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=3, column=0, sticky="e", padx=4, pady=(8, 0))
        self.e_demo_multa_dia = _entry_make(cfg_inner, width=8, placeholder="0,33")
        self.e_demo_multa_dia.grid(row=3, column=1, sticky="w", padx=4, pady=(8, 0))
        _entry_set(self.e_demo_multa_dia, "0,33")
        self.lbl_demo_multadia_hint = tk.Label(
            cfg_inner, text="% a.d.", bg=COLOR_PANEL,
            fg=COLOR_SUBTLE, font=("Verdana", 8))
        self.lbl_demo_multadia_hint.grid(row=3, column=2, sticky="w",
                padx=(0, 8), pady=(8, 0))

        tk.Label(cfg_inner, text="Limite Multa:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=3, column=3, sticky="e", padx=(20, 4), pady=(8, 0))
        self.e_demo_multa = _entry_make(cfg_inner, width=8, placeholder="20")
        self.e_demo_multa.grid(row=3, column=4, sticky="w", padx=4, pady=(8, 0))
        _entry_set(self.e_demo_multa, "20")
        self.lbl_demo_multa_hint = tk.Label(
            cfg_inner, text="% (só Selic — multa diária × dias, até este teto)",
            bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7))
        self.lbl_demo_multa_hint.grid(row=3, column=5, columnspan=3,
                sticky="w", padx=(0, 4), pady=(8, 0))

        # linha 5 - honorários
        self.honor_var = tk.BooleanVar(value=False)
        self.chk_honor = tk.Checkbutton(
            cfg_inner, text="Aplicar honorários sobre o débito atualizado",
            variable=self.honor_var, bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            activebackground=COLOR_PANEL, font=("Verdana", 9, "bold"),
            selectcolor="white",
            command=lambda: self._demo_toggle_honor_state())
        self.chk_honor.grid(row=4, column=0, columnspan=3, sticky="w",
                          padx=4, pady=(10, 0))

        self.e_demo_honor_pct = _entry_make(cfg_inner, width=8, placeholder="10")
        self.e_demo_honor_pct.grid(row=4, column=3, sticky="w",
                                  padx=(20, 4), pady=(10, 0))
        _entry_set(self.e_demo_honor_pct, "10")
        self.e_demo_honor_pct.configure(state="disabled")  # começa desabilitado
        tk.Label(cfg_inner, text="% sobre o total geral atualizado",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
                row=4, column=4, columnspan=4, sticky="w",
                padx=(0, 4), pady=(10, 0))


        # ===== Alíquotas por Período (opcional) =====
        self._periodos_frame = tk.LabelFrame(
            inp, text=" Alíquotas por Período (IPREM) ",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            font=("Verdana", 8, "bold"), bd=1, relief="solid")
        self._periodos_frame.pack(fill="x", pady=(0, 4))

        self.usar_periodos_var = tk.BooleanVar(value=False)
        chk_p = tk.Checkbutton(
            self._periodos_frame,
            text="Usar alíquotas diferenciadas por período",
            variable=self.usar_periodos_var,
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            activebackground=COLOR_PANEL, font=("Verdana", 9, "bold"),
            selectcolor="white",
            command=self._demo_toggle_periodos)
        chk_p.pack(side="left", padx=(8, 4), pady=6)

        ttk.Button(self._periodos_frame, text="Padrão IPREM/SP",
                   style="BCBSmall.TButton",
                   command=self._demo_reset_periodos_iprem).pack(
                       side="right", padx=8, pady=6)

        # Container interno (mostrado/oculto)
        self._periodos_inner = tk.Frame(self._periodos_frame, bg=COLOR_PANEL)
        # Header
        _ph = tk.Frame(self._periodos_inner, bg=COLOR_BCB_BLUE)
        _ph.pack(fill="x")
        for _txt, _w in [("Data Início", 12), ("Data Fim (vazio=aberto)", 18),
                          ("Alíq. Seg. %", 10), ("Alíq. Pat. %", 10), ("", 4)]:
            tk.Label(_ph, text=_txt, bg=COLOR_BCB_BLUE, fg="white",
                     font=("Verdana", 8, "bold"), width=_w,
                     padx=4, pady=3).pack(side="left", padx=1)
        self._periodos_rows_frame = tk.Frame(self._periodos_inner, bg=COLOR_PANEL)
        self._periodos_rows_frame.pack(fill="x")
        ttk.Button(self._periodos_inner, text="+ Adicionar período",
                   style="BCBSmall.TButton",
                   command=self._demo_add_periodo).pack(side="left", pady=(4, 4), padx=4)
        self.periodos_rows = []
        # começa oculto

        # Botões topo
        top_bar = tk.Frame(inp, bg=COLOR_PANEL)
        top_bar.pack(fill="x", pady=(0, 4))
        ttk.Button(top_bar, text="+ Competência",
                  style="BCBSmall.TButton",
                  command=self._demo_add_row).pack(side="left", padx=(0, 4))
        ttk.Button(top_bar, text="Importar CSV…",
                  style="BCBSmall.TButton",
                  command=lambda: self._demo_import("csv")).pack(side="left", padx=2)
        if HAS_XLSX:
            ttk.Button(top_bar, text="Importar XLSX…",
                      style="BCBSmall.TButton",
                      command=lambda: self._demo_import("xlsx")).pack(side="left", padx=2)
        ttk.Button(top_bar, text="Limpar tudo",
                  style="BCBSmall.TButton",
                  command=self._demo_clear).pack(side="left", padx=(8, 2))

        self.btn_demo_calc = ttk.Button(top_bar, text="▶ Calcular demonstrativo",
                                        style="BCB.TButton",
                                        command=self._demo_calc)
        self.btn_demo_calc.pack(side="right", padx=2)

        # Tabela competências
        comp_frame = tk.LabelFrame(inp, text=" Competências ",
                                  bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                  font=("Verdana", 8, "bold"),
                                  bd=1, relief="solid")
        comp_frame.pack(fill="x", pady=(0, 4))

        hdr = tk.Frame(comp_frame, bg=COLOR_BCB_BLUE)
        hdr.pack(fill="x")
        cols = [("#", 30), ("Competência (MM/AAAA)", 180),
                ("Descrição", 240), ("Base de Cálculo (R$)", 180), ("", 60)]
        for txt, w in cols:
            tk.Label(hdr, text=txt, bg=COLOR_BCB_BLUE, fg="white",
                    font=("Verdana", 8, "bold"), width=max(w//8, 4),
                    padx=4, pady=4).pack(side="left", padx=1)

        canvas_frame = tk.Frame(comp_frame, bg=COLOR_PANEL)
        canvas_frame.pack(fill="x")
        self.demo_canvas = tk.Canvas(canvas_frame, bg=COLOR_PANEL,
                                    height=140, highlightthickness=0)
        self.demo_canvas.pack(side="left", fill="x", expand=True)
        vsb = ttk.Scrollbar(canvas_frame, orient="vertical",
                           command=self.demo_canvas.yview)
        vsb.pack(side="right", fill="y")
        self.demo_canvas.configure(yscrollcommand=vsb.set)

        self.demo_rows_frame = tk.Frame(self.demo_canvas, bg=COLOR_PANEL)
        self.demo_canvas.create_window((0, 0), window=self.demo_rows_frame, anchor="nw")
        self.demo_rows_frame.bind(
            "<Configure>",
            lambda e: self.demo_canvas.configure(scrollregion=self.demo_canvas.bbox("all"))
        )

        def _on_mw(e):
            self.demo_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self.demo_canvas.bind("<MouseWheel>", _on_mw)
        self.demo_rows_frame.bind("<MouseWheel>", _on_mw)

        self.demo_rows = []
        # 3 linhas exemplo
        for _ in range(3):
            self._demo_add_row()

        # Resultado
        result_frame = tk.LabelFrame(res, text=" Resultado do Demonstrativo ",
                                    bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                    font=("Verdana", 8, "bold"),
                                    bd=1, relief="solid")
        result_frame.pack(fill="both", expand=True)

        toolbar = tk.Frame(result_frame, bg=COLOR_PANEL)
        toolbar.pack(fill="x", padx=4, pady=4)
        self.lbl_demo_status = tk.Label(toolbar, text="",
                                        bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                                        font=("Verdana", 8))
        self.lbl_demo_status.pack(side="left", padx=4)

        ttk.Button(toolbar, text="Exportar CSV…", style="BCBSmall.TButton",
                  command=lambda: self._demo_export("csv")).pack(side="right", padx=2)
        if HAS_XLSX:
            ttk.Button(toolbar, text="Exportar XLSX…", style="BCBSmall.TButton",
                      command=lambda: self._demo_export("xlsx")).pack(side="right", padx=2)
        if HAS_PDF:
            ttk.Button(toolbar, text="Exportar PDF…", style="BCBSmall.TButton",
                      command=lambda: self._demo_export("pdf")).pack(side="right", padx=2)

        tv_frame = tk.Frame(result_frame, bg=COLOR_PANEL)
        tv_frame.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        # Treeview com colunas adaptadas
        self.demo_tree_frame = tv_frame
        self._demo_build_tree(eh_selic=False)  # default IPCA

        self.demo_resultado = None

    def _demo_build_tree(self, eh_selic):
        # remover treeview anterior
        for w in self.demo_tree_frame.winfo_children():
            w.destroy()

        if eh_selic:
            columns = ("comp", "desc", "base", "venc", "fator",
                      "dev_seg", "atu_seg", "j_seg", "m_seg", "t_seg",
                      "dev_pat", "atu_pat", "j_pat", "m_pat", "t_pat",
                      "geral", "sit")
            headings = {
                "comp": "Competência", "desc": "Descr.", "base": "Base",
                "venc": "Venc.", "fator": "Fator Selic",
                "dev_seg": "Dev.Seg", "atu_seg": "Atu.Seg",
                "j_seg": "1%Seg", "m_seg": "Multa Seg", "t_seg": "Total Seg",
                "dev_pat": "Dev.Pat", "atu_pat": "Atu.Pat",
                "j_pat": "1%Pat", "m_pat": "Multa Pat", "t_pat": "Total Pat",
                "geral": "Total Geral", "sit": "Situação",
            }
            widths = {"comp": 90, "desc": 100, "base": 80, "venc": 80,
                     "fator": 80, "dev_seg": 70, "atu_seg": 75, "j_seg": 60,
                     "m_seg": 65, "t_seg": 80, "dev_pat": 70, "atu_pat": 75,
                     "j_pat": 60, "m_pat": 65, "t_pat": 80,
                     "geral": 90, "sit": 80}
        else:
            columns = ("comp", "desc", "base", "venc", "fator", "meses",
                      "dev_seg", "atu_seg", "j_seg", "t_seg",
                      "dev_pat", "atu_pat", "j_pat", "t_pat",
                      "geral", "sit")
            headings = {
                "comp": "Competência", "desc": "Descr.", "base": "Base",
                "venc": "Venc.", "fator": "Fator Índice", "meses": "Meses",
                "dev_seg": "Dev.Seg", "atu_seg": "Atu.Seg",
                "j_seg": "Juros Seg", "t_seg": "Total Seg",
                "dev_pat": "Dev.Pat", "atu_pat": "Atu.Pat",
                "j_pat": "Juros Pat", "t_pat": "Total Pat",
                "geral": "Total Geral", "sit": "Situação",
            }
            widths = {"comp": 100, "desc": 110, "base": 80, "venc": 80,
                     "fator": 85, "meses": 50, "dev_seg": 75, "atu_seg": 80,
                     "j_seg": 70, "t_seg": 85, "dev_pat": 75, "atu_pat": 80,
                     "j_pat": 70, "t_pat": 85, "geral": 95, "sit": 90}

        self.demo_tree = ttk.Treeview(
            self.demo_tree_frame, columns=columns,
            show="headings", style="BCB.Treeview", height=10,
        )
        for c in columns:
            self.demo_tree.heading(c, text=headings[c])
            anchor = "w" if c in ("desc", "sit") else "e"
            self.demo_tree.column(c, width=widths[c], anchor=anchor)

        # scroll H+V
        vsb = ttk.Scrollbar(self.demo_tree_frame, orient="vertical",
                           command=self.demo_tree.yview)
        hsb = ttk.Scrollbar(self.demo_tree_frame, orient="horizontal",
                           command=self.demo_tree.xview)
        self.demo_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.demo_tree.pack(side="left", fill="both", expand=True)

        self.demo_tree.tag_configure("totais", background="#dde7f0",
                                    foreground=COLOR_BCB_BLUE,
                                    font=("Verdana", 8, "bold"))
        self.demo_tree.tag_configure("alt", background=COLOR_TABLE_ALT)

    def _demo_toggle_honor_state(self):
        """Habilita/desabilita campo de percentual conforme checkbox."""
        if self.honor_var.get():
            self.e_demo_honor_pct.configure(state="normal")
        else:
            self.e_demo_honor_pct.configure(state="disabled")

    def _demo_toggle_periodos(self):
        if self.usar_periodos_var.get():
            self._periodos_inner.pack(fill="x", padx=8, pady=(0, 8))
            if not self.periodos_rows:
                self._demo_reset_periodos_iprem()
        else:
            self._periodos_inner.pack_forget()

    def _demo_reset_periodos_iprem(self):
        """Preenche com as alíquotas históricas padrão IPREM/SP."""
        for r in list(self.periodos_rows):
            r["frame"].destroy()
        self.periodos_rows.clear()
        padrao = [
            ("04/01/1990", "10/08/2005", "5", "2"),
            ("11/08/2005", "27/03/2019", "11", "22"),
            ("28/03/2019", "", "14", "28"),
        ]
        for ini, fim, seg, pat in padrao:
            self._demo_add_periodo(ini, fim, seg, pat)

    def _demo_add_periodo(self, ini="", fim="", seg="", pat=""):
        """Adiciona uma linha de período de alíquota."""
        row_frame = tk.Frame(self._periodos_rows_frame, bg=COLOR_PANEL)
        row_frame.pack(fill="x", pady=1)

        e_ini = _entry_make(row_frame, width=12, placeholder="DD/MM/AAAA")
        e_ini.pack(side="left", padx=(4, 2))
        _mask_full_date(e_ini)
        if ini:
            _entry_set(e_ini, ini)

        e_fim = _entry_make(row_frame, width=16, placeholder="DD/MM/AAAA")
        e_fim.pack(side="left", padx=2)
        _mask_full_date(e_fim)
        if fim:
            _entry_set(e_fim, fim)

        e_seg = _entry_make(row_frame, width=8, placeholder="14")
        e_seg.pack(side="left", padx=(8, 2))
        if seg:
            _entry_set(e_seg, seg)
        tk.Label(row_frame, text="%", bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 8)).pack(side="left")

        e_pat = _entry_make(row_frame, width=8, placeholder="28")
        e_pat.pack(side="left", padx=(10, 2))
        if pat:
            _entry_set(e_pat, pat)
        tk.Label(row_frame, text="%", bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 8)).pack(side="left")

        row_data = {"frame": row_frame, "ini": e_ini, "fim": e_fim,
                    "seg": e_seg, "pat": e_pat}

        def remove_row(rd=row_data):
            rd["frame"].destroy()
            if rd in self.periodos_rows:
                self.periodos_rows.remove(rd)

        ttk.Button(row_frame, text="✕", style="BCBSmall.TButton",
                   width=3, command=remove_row).pack(side="left", padx=6)
        self.periodos_rows.append(row_data)

    def _demo_toggle_selic_fields(self, event=None):
        is_selic = self.cb_demo_indice.get() == "SELIC"
        self.lbl_demo_multa_hint.config(
            text="% (teto da multa Selic)" if is_selic
            else "% (não usado p/ índice mensal)")
        self.lbl_demo_multadia_hint.config(
            text="% a.d.  (só Selic)" if is_selic
            else "% a.d.  (não usado p/ índice mensal)")

    def _demo_add_row(self, competencia="", descricao="", base=""):
        rowf = tk.Frame(self.demo_rows_frame, bg=COLOR_PANEL)
        rowf.pack(fill="x", padx=2, pady=1)

        num = len(self.demo_rows) + 1
        lbl = tk.Label(rowf, text=str(num), bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                      font=("Verdana", 8), width=4)
        lbl.pack(side="left", padx=2)

        e_comp = _entry_make(rowf, width=20, placeholder="MM/AAAA")
        e_comp.pack(side="left", padx=1)
        _mask_month_year(e_comp)
        if competencia:
            _entry_set(e_comp, competencia)

        e_desc = _entry_make(rowf, width=28, placeholder="(opcional)")
        e_desc.pack(side="left", padx=1)
        if descricao:
            _entry_set(e_desc, descricao)

        e_base = _entry_make(rowf, width=20, placeholder="0,00")
        e_base.pack(side="left", padx=1)
        _bind_valor_format(e_base)
        if base:
            _entry_set(e_base, base)

        row_data = {"frame": rowf, "lbl": lbl,
                   "comp": e_comp, "desc": e_desc, "base": e_base}

        btn_rm = tk.Button(
            rowf, text="✕", bg="#fdecea", fg=COLOR_ERROR,
            font=("Verdana", 8, "bold"), bd=0, padx=6, pady=1,
            command=lambda r=row_data: self._demo_remove_row(r),
            cursor="hand2",
        )
        btn_rm.pack(side="left", padx=4)
        row_data["btn_rm"] = btn_rm

        self.demo_rows.append(row_data)

    def _demo_remove_row(self, row_data):
        row_data["frame"].destroy()
        self.demo_rows.remove(row_data)
        for i, r in enumerate(self.demo_rows, start=1):
            r["lbl"].config(text=str(i))

    def _demo_clear(self):
        if not messagebox.askyesno("Limpar", "Remover todas as competências?"):
            return
        for r in list(self.demo_rows):
            r["frame"].destroy()
        self.demo_rows.clear()
        if hasattr(self, "demo_tree"):
            for item in self.demo_tree.get_children():
                self.demo_tree.delete(item)
        self.demo_resultado = None
        self.lbl_demo_status.config(text="")

    def _demo_import(self, fmt):
        if fmt == "csv":
            path = filedialog.askopenfilename(
                title="Importar Competências (CSV)",
                filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        else:
            path = filedialog.askopenfilename(
                title="Importar Competências (XLSX)",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if not path:
            return
        try:
            if fmt == "csv":
                linhas = importar_csv_demo(path)
            else:
                linhas = importar_xlsx_demo(path)
        except Exception as e:
            messagebox.showerror("Erro ao importar", str(e))
            return
        if not linhas:
            messagebox.showwarning(
                "Importação",
                "Nenhuma competência encontrada. Verifique se o arquivo tem cabeçalho "
                "com colunas: Competência, Descrição, Base de Cálculo.")
            return

        for r in list(self.demo_rows):
            r["frame"].destroy()
        self.demo_rows.clear()

        for linha in linhas:
            self._demo_add_row(
                competencia=linha.get("competencia", ""),
                descricao=linha.get("descricao", ""),
                base=linha.get("base", ""),
            )
        self.lbl_demo_status.config(
            text=f"✓ {len(linhas)} competência(s) importada(s) de {os.path.basename(path)}")

    def _demo_export(self, fmt):
        if not self.demo_resultado:
            messagebox.showwarning("Exportar", "Calcule o demonstrativo primeiro.")
            return
        if fmt == "csv":
            path = filedialog.asksaveasfilename(
                title="Salvar CSV", defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile="demonstrativo.csv")
        elif fmt == "xlsx":
            path = filedialog.asksaveasfilename(
                title="Salvar XLSX", defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="demonstrativo.xlsx")
        elif fmt == "pdf":
            path = filedialog.asksaveasfilename(
                title="Salvar PDF", defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile="demonstrativo.pdf")
        else:
            return
        if not path:
            return
        try:
            dp = self._coletar_dados_processo("demo")
            if fmt == "csv":
                exportar_csv_demo(path, self.demo_resultado)
            elif fmt == "xlsx":
                exportar_xlsx_demo(path, self.demo_resultado, dados_processo=dp)
            elif fmt == "pdf":
                exportar_pdf_demo(path, self.demo_resultado, dados_processo=dp)
            messagebox.showinfo("Exportar", f"Arquivo salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _demo_calc(self):
        # ===== Validação por campo (mensagens específicas) =====
        indice_key = self.cb_demo_indice.get()

        data_atu = parse_date_br(_entry_value(self.e_demo_dataatu))
        if not data_atu:
            return messagebox.showerror(
                "Validação",
                "Data de Atualização inválida. Use o formato DD/MM/AAAA.")
        if data_atu.year < 1990 or data_atu.year > 2100:
            return messagebox.showerror(
                "Validação",
                f"Data de Atualização tem ano implausível ({data_atu.year}). "
                "Verifique se digitou corretamente (use o formato DD/MM/AAAA).")

        def _pct(entry, nome, default=None):
            v = _entry_value(entry).strip()
            if not v:
                if default is not None:
                    return Decimal(default) / Decimal("100")
                raise ValueError(f"O campo '{nome}' está vazio.")
            try:
                return Decimal(v.replace(",", ".")) / Decimal("100")
            except Exception:
                raise ValueError(
                    f"O campo '{nome}' tem valor inválido: '{v}'. "
                    "Use apenas números (ex: 14 ou 0,33).")

        try:
            aliq_seg = _pct(self.e_demo_aliqseg, "Alíquota Segurado")
            aliq_pat = _pct(self.e_demo_aliqpat, "Alíquota Patronal")
            juros_mes = _pct(self.e_demo_juros_mes, "Juros Mensais", default="1")
            multa_dia = _pct(self.e_demo_multa_dia, "Multa Diária", default="0.33")
            multa_lim = _pct(self.e_demo_multa, "Limite Multa", default="20")
        except ValueError as e:
            return messagebox.showerror("Validação", str(e))

        try:
            dia_v_str = _entry_value(self.e_demo_diavenc).strip() or "5"
            dia_venc = int(dia_v_str)
        except Exception:
            return messagebox.showerror(
                "Validação",
                f"Dia do Vencimento inválido: '{dia_v_str}'. "
                "Use um número inteiro entre 1 e 28.")

        if dia_venc < 1 or dia_venc > 28:
            return messagebox.showerror(
                "Validação",
                "Dia do Vencimento deve ser entre 1 e 28.")

        # Honorários
        aplicar_honor = bool(self.honor_var.get())
        honor_pct = Decimal("0")
        if aplicar_honor:
            try:
                honor_pct = _pct(self.e_demo_honor_pct, "Honorários", default="10")
            except ValueError as e:
                return messagebox.showerror("Validação", str(e))

        # ===== Coletar competências =====
        competencias = []
        for i, r in enumerate(self.demo_rows, start=1):
            comp = _entry_value(r["comp"]).strip()
            desc = _entry_value(r["desc"]).strip()
            base_raw = _entry_value(r["base"]).strip()
            if not any([comp, desc, base_raw]):
                continue
            my = parse_month_year(comp)
            if not my:
                return messagebox.showerror(
                    "Validação",
                    f"Linha {i}: competência inválida ('{comp}'). Use MM/AAAA.")
            try:
                base = parse_valor_br(base_raw)
                if base is None:
                    raise ValueError()
            except Exception:
                return messagebox.showerror(
                    "Validação",
                    f"Linha {i}: base de cálculo inválida ('{base_raw}').")
            competencias.append({
                "mes": my[0], "ano": my[1],
                "descricao": desc,
                "base_calculo": base,
            })

        if not competencias:
            return messagebox.showwarning(
                "Demonstrativo", "Adicione ao menos uma competência.")

        # Coletar períodos de alíquota se habilitado
        periodos_aliquota_cfg = None
        if getattr(self, "usar_periodos_var", None) and self.usar_periodos_var.get():
            periodos_aliquota_cfg = []
            for r in self.periodos_rows:
                ini_str = _entry_value(r["ini"]).strip()
                fim_str = _entry_value(r["fim"]).strip()
                seg_str = _entry_value(r["seg"]).strip()
                pat_str = _entry_value(r["pat"]).strip()
                if not ini_str or not seg_str or not pat_str:
                    continue
                d_ini = parse_date_br(ini_str)
                d_fim = parse_date_br(fim_str) if fim_str else None
                if not d_ini:
                    return messagebox.showerror(
                        "Validação", f"Período: data início inválida '{ini_str}'.")
                try:
                    alq_s = Decimal(seg_str.replace(",", ".")) / Decimal("100")
                    alq_p = Decimal(pat_str.replace(",", ".")) / Decimal("100")
                except Exception:
                    return messagebox.showerror(
                        "Validação", "Período: alíquota inválida. Use números (ex: 14).")
                periodos_aliquota_cfg.append({
                    "data_ini": d_ini,
                    "data_fim": d_fim,
                    "aliq_seg": alq_s,
                    "aliq_pat": alq_p,
                })

        config = {
            "indice_key": indice_key,
            "data_atualizacao": data_atu,
            "aliquota_seg": aliq_seg,
            "aliquota_pat": aliq_pat,
            "dia_vencimento": dia_venc,
            "juros_mensais_pct": juros_mes,    # editável (IPCA e Selic)
            "multa_diaria_pct": multa_dia,     # editável (só Selic)
            "multa_limite_pct": multa_lim,     # editável (teto Selic)
            "aplicar_honorarios": aplicar_honor,
            "honorarios_pct": honor_pct,
            "periodos_aliquota": periodos_aliquota_cfg or [],
        }

        self.btn_demo_calc.config(state="disabled")
        self.lbl_demo_status.config(text="Calculando...")
        self._demo_build_tree(eh_selic=(indice_key == "SELIC"))

        # Limpar cache da API pra forçar nova consulta
        limpar_cache_api()

        def progress(msg):
            self.after(0, lambda: self.lbl_demo_status.config(text=msg))

        def worker():
            try:
                res = calcular_demonstrativo(config, competencias,
                                            progress_cb=progress)
                self.after(0, lambda: self._render_demo_results(res))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: messagebox.showerror("Erro no cálculo", err))
            finally:
                self.after(0, lambda: self.btn_demo_calc.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _render_demo_results(self, res):
        self.demo_resultado = res
        eh_selic = res["config"]["indice_key"] == "SELIC"

        for item in self.demo_tree.get_children():
            self.demo_tree.delete(item)

        for i, l in enumerate(res["linhas"]):
            tag = "alt" if i % 2 == 1 else ""
            if eh_selic:
                values = (
                    l["competencia"], l["descricao"][:18],
                    fmt_brl(l["base"]),
                    l["vencimento"].strftime("%d/%m/%y"),
                    fmt_fator(l["fator"]),
                    fmt_brl(l["valor_devido_seg"]),
                    fmt_brl(l["valor_atual_seg"]),
                    fmt_brl(l["juros_seg"]),
                    fmt_brl(l["multa_seg"]),
                    fmt_brl(l["total_seg"]),
                    fmt_brl(l["valor_devido_pat"]),
                    fmt_brl(l["valor_atual_pat"]),
                    fmt_brl(l["juros_pat"]),
                    fmt_brl(l["multa_pat"]),
                    fmt_brl(l["total_pat"]),
                    fmt_brl(l["total_geral"]),
                    l["situacao"],
                )
            else:
                values = (
                    l["competencia"], l["descricao"][:18],
                    fmt_brl(l["base"]),
                    l["vencimento"].strftime("%d/%m/%y"),
                    fmt_fator(l["fator"]),
                    f"{float(l['meses_atraso']):.2f}".replace(".", ","),
                    fmt_brl(l["valor_devido_seg"]),
                    fmt_brl(l["valor_atual_seg"]),
                    fmt_brl(l["juros_seg"]),
                    fmt_brl(l["total_seg"]),
                    fmt_brl(l["valor_devido_pat"]),
                    fmt_brl(l["valor_atual_pat"]),
                    fmt_brl(l["juros_pat"]),
                    fmt_brl(l["total_pat"]),
                    fmt_brl(l["total_geral"]),
                    l["situacao"],
                )
            self.demo_tree.insert("", "end", values=values, tags=(tag,))

        # Linha de totais
        t = res["totais"]
        if eh_selic:
            tot_values = (
                "TOTAIS", "", fmt_brl(t["base_total"]), "", "",
                fmt_brl(t["valor_devido_seg"]),
                fmt_brl(t["valor_atual_seg"]),
                fmt_brl(t["juros_seg"]),
                fmt_brl(t["multa_seg"]),
                fmt_brl(t["total_seg"]),
                fmt_brl(t["valor_devido_pat"]),
                fmt_brl(t["valor_atual_pat"]),
                fmt_brl(t["juros_pat"]),
                fmt_brl(t["multa_pat"]),
                fmt_brl(t["total_pat"]),
                fmt_brl(t["total_geral"]),
                "",
            )
        else:
            tot_values = (
                "TOTAIS", "", fmt_brl(t["base_total"]), "", "", "",
                fmt_brl(t["valor_devido_seg"]),
                fmt_brl(t["valor_atual_seg"]),
                fmt_brl(t["juros_seg"]),
                fmt_brl(t["total_seg"]),
                fmt_brl(t["valor_devido_pat"]),
                fmt_brl(t["valor_atual_pat"]),
                fmt_brl(t["juros_pat"]),
                fmt_brl(t["total_pat"]),
                fmt_brl(t["total_geral"]),
                "",
            )
        self.demo_tree.insert("", "end", values=tot_values, tags=("totais",))

        # Linhas extras de honorários
        if t.get("aplicar_honorarios") and t.get("honorarios", Decimal("0")) > 0:
            n_cols = len(self.demo_tree["columns"])
            pct = float(t["honorarios_pct"]) * 100
            label_honor = f"HONORÁRIOS ({pct:.1f}% sobre o total)".replace(".", ",")
            valores_honor = [""] * n_cols
            valores_honor[0] = label_honor
            # Penúltima coluna (Total Geral) recebe o valor
            valores_honor[-2] = fmt_brl(t["honorarios"])
            self.demo_tree.insert("", "end", values=valores_honor, tags=("honor",))

            valores_total = [""] * n_cols
            valores_total[0] = "TOTAL + HONORÁRIOS"
            valores_total[-2] = fmt_brl(t["total_com_honorarios"])
            self.demo_tree.insert("", "end", values=valores_total, tags=("totais",))

        # Tags de estilo
        self.demo_tree.tag_configure("honor", background="#FFF9E0",
                                    foreground="#9C7700",
                                    font=("Verdana", 8, "bold"))

        # Status final
        if t.get("aplicar_honorarios") and t.get("honorarios", Decimal("0")) > 0:
            status_txt = (f"✓ Demonstrativo calculado ({t['qtd_competencias']} competências)  ·  "
                         f"Total: R$ {fmt_brl(t['total_geral'])}  +  "
                         f"Honorários: R$ {fmt_brl(t['honorarios'])}  =  "
                         f"R$ {fmt_brl(t['total_com_honorarios'])}")
        else:
            status_txt = (f"✓ Demonstrativo calculado ({t['qtd_competencias']} competências)  ·  "
                         f"TOTAL GERAL: R$ {fmt_brl(t['total_geral'])}")
        self.lbl_demo_status.config(text=status_txt)


    # ================================================================
    # Tab: Cobrança Amigável
    # ================================================================
    def _build_cobranca_tab(self):
        tab_outer = ttk.Frame(self.notebook, style="BCB.TFrame")
        self.notebook.add(tab_outer, text="Cobrança Amigável")

        # Canvas rolável que ocupa a aba inteira, para que em telas menores
        # o usuário consiga rolar até o botão Calcular e o resultado.
        cob_canvas = tk.Canvas(tab_outer, bg=COLOR_PANEL, highlightthickness=0)
        cob_vsb = ttk.Scrollbar(tab_outer, orient="vertical",
                               command=cob_canvas.yview)
        cob_canvas.configure(yscrollcommand=cob_vsb.set)
        cob_vsb.pack(side="right", fill="y")
        cob_canvas.pack(side="left", fill="both", expand=True)

        # Frame interno com padding — este passa a ser o "tab" usado abaixo.
        tab = tk.Frame(cob_canvas, bg=COLOR_PANEL, padx=10, pady=10)
        _cob_window = cob_canvas.create_window((0, 0), window=tab, anchor="nw")

        # Ajustar a largura do frame interno à largura do canvas
        def _cob_resize(event):
            cob_canvas.itemconfig(_cob_window, width=event.width)
        cob_canvas.bind("<Configure>", _cob_resize)

        # Atualizar a região de rolagem quando o conteúdo mudar de tamanho
        tab.bind(
            "<Configure>",
            lambda e: cob_canvas.configure(scrollregion=cob_canvas.bbox("all"))
        )

        # Rolagem com a roda do mouse enquanto o ponteiro estiver sobre a aba
        def _cob_on_mw(e):
            cob_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        def _cob_bind_mw(_e):
            cob_canvas.bind_all("<MouseWheel>", _cob_on_mw)
        def _cob_unbind_mw(_e):
            cob_canvas.unbind_all("<MouseWheel>")
        cob_canvas.bind("<Enter>", _cob_bind_mw)
        cob_canvas.bind("<Leave>", _cob_unbind_mw)

        ttk.Label(
            tab,
            text=(" Atualização de débito único do termo (valor cheio) por IPCA, "
                  "com multa, juros, honorários e abatimentos opcionais."),
            style="Info.TLabel"
        ).pack(fill="x", pady=(0, 6))

        # === Bloco: Débito ===
        deb_frame = tk.LabelFrame(tab, text=" Débito Original ",
                                 bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                 font=("Verdana", 8, "bold"),
                                 bd=1, relief="solid")
        deb_frame.pack(fill="x", pady=(0, 6))

        deb_inner = tk.Frame(deb_frame, bg=COLOR_PANEL)
        deb_inner.pack(fill="x", padx=8, pady=8)

        tk.Label(deb_inner, text="Data de Origem:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=0, sticky="e", padx=4)
        self.e_cob_data_origem = _entry_make(deb_inner, width=14,
                                            placeholder="DD/MM/AAAA")
        self.e_cob_data_origem.grid(row=0, column=1, sticky="w", padx=4)
        _mask_full_date(self.e_cob_data_origem)

        tk.Label(deb_inner, text="Valor Original (R$):", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=2, sticky="e", padx=(20, 4))
        self.e_cob_valor_origem = _entry_make(deb_inner, width=16,
                                             placeholder="0,00")
        self.e_cob_valor_origem.grid(row=0, column=3, sticky="w", padx=4)
        _bind_valor_format(self.e_cob_valor_origem)

        tk.Label(deb_inner, text="Data de Atualização:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=1, column=0, sticky="e", padx=4, pady=(8, 0))
        self.e_cob_data_atual = _entry_make(deb_inner, width=14,
                                           placeholder="DD/MM/AAAA")
        self.e_cob_data_atual.grid(row=1, column=1, sticky="w", padx=4, pady=(8, 0))
        _mask_full_date(self.e_cob_data_atual)
        _entry_set(self.e_cob_data_atual, date.today().strftime("%d/%m/%Y"))

        tk.Label(deb_inner, text="Índice:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=1, column=2, sticky="e", padx=(20, 4), pady=(8, 0))
        self.cb_cob_indice = ttk.Combobox(
            deb_inner, values=["IPCA", "IGP-M", "IGP-DI", "INPC",
                              "IPCA-E", "IPC-BR", "IPC-SP"],
            state="readonly", width=10)
        self.cb_cob_indice.set("IPCA")
        self.cb_cob_indice.grid(row=1, column=3, sticky="w", padx=4, pady=(8, 0))

        # === Bloco: Várias competências (opcional) ===
        comp_frame = tk.LabelFrame(tab, text=" Débito em várias competências (opcional) ",
                                  bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                  font=("Verdana", 8, "bold"),
                                  bd=1, relief="solid")
        comp_frame.pack(fill="x", pady=(0, 6))
        comp_inner = tk.Frame(comp_frame, bg=COLOR_PANEL)
        comp_inner.pack(fill="x", padx=8, pady=6)

        self.cob_multimes_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            comp_inner, text="Lançar débito em várias competências (mês a mês)",
            variable=self.cob_multimes_var, bg=COLOR_PANEL,
            fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
            activebackground=COLOR_PANEL, selectcolor="white",
            command=self._cob_toggle_multimes).grid(
            row=0, column=0, sticky="w", padx=4, columnspan=3)

        tk.Label(comp_inner,
                text="Quando marcado, cada competência é tratada individualmente: "
                "corrigida pelo IPCA da sua própria data até a Data de Atualização "
                "e com juros calculados em MESES INTEIROS (igual ao SEI). "
                "Importante: digite na Data de Atualização a data exata do "
                "relatório/cálculo, pois é ela que define os meses de juros.",
                bg=COLOR_PANEL, fg="#9C7700", font=("Verdana", 7),
                wraplength=820, justify="left").grid(
            row=1, column=0, columnspan=3, sticky="w", padx=4, pady=(2, 4))

        # Cabeçalho da tabela
        self.cob_comp_header = tk.Frame(comp_inner, bg=COLOR_BCB_BLUE)
        self.cob_comp_header.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(4, 0))
        for txt, w in [("#", 4), ("Competência (MM/AAAA)", 22),
                      ("Descrição (opcional)", 26), ("Valor (R$)", 16), ("", 4)]:
            tk.Label(self.cob_comp_header, text=txt, bg=COLOR_BCB_BLUE, fg="white",
                    font=("Verdana", 8, "bold"), width=w, anchor="w",
                    padx=4, pady=3).pack(side="left", padx=1)

        self.cob_comp_rows_frame = tk.Frame(comp_inner, bg=COLOR_PANEL)
        self.cob_comp_rows_frame.grid(row=3, column=0, columnspan=3, sticky="ew")
        self.cob_comp_rows = []

        self.btn_cob_add_comp = ttk.Button(
            comp_inner, text="+ Adicionar competência",
            style="BCBSmall.TButton", command=self._cob_add_comp_row)
        self.btn_cob_add_comp.grid(row=4, column=0, sticky="w", padx=4, pady=(4, 0))

        # Estado inicial: tabela desabilitada (modo débito único)
        self._cob_toggle_multimes()

        # === Bloco: Multa ===
        multa_frame = tk.LabelFrame(tab, text=" Multa ",
                                   bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                   font=("Verdana", 8, "bold"),
                                   bd=1, relief="solid")
        multa_frame.pack(fill="x", pady=(0, 6))
        multa_inner = tk.Frame(multa_frame, bg=COLOR_PANEL)
        multa_inner.pack(fill="x", padx=8, pady=6)

        self.cob_multa_var = tk.BooleanVar(value=False)
        tk.Checkbutton(multa_inner, text="Aplicar multa",
                      variable=self.cob_multa_var, bg=COLOR_PANEL,
                      fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
                      activebackground=COLOR_PANEL, selectcolor="white",
                      command=self._cob_toggle_multa).grid(
            row=0, column=0, sticky="w", padx=4)

        self.e_cob_multa_pct = _entry_make(multa_inner, width=6, placeholder="10")
        self.e_cob_multa_pct.grid(row=0, column=1, sticky="w", padx=4)
        _entry_set(self.e_cob_multa_pct, "10")
        self.e_cob_multa_pct.configure(state="disabled")
        tk.Label(multa_inner, text="% sobre:", bg=COLOR_PANEL,
                fg=COLOR_SUBTLE, font=("Verdana", 8)).grid(
            row=0, column=2, sticky="w", padx=(0, 4))

        self.cob_multa_sobre = tk.StringVar(value="atualizado")
        self.rb_multa_atu = tk.Radiobutton(
            multa_inner, text="Valor atualizado",
            variable=self.cob_multa_sobre, value="atualizado",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white",
            state="disabled")
        self.rb_multa_atu.grid(row=0, column=3, sticky="w", padx=4)
        self.rb_multa_ori = tk.Radiobutton(
            multa_inner, text="Valor original",
            variable=self.cob_multa_sobre, value="original",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white",
            state="disabled")
        self.rb_multa_ori.grid(row=0, column=4, sticky="w", padx=4)

        # === Bloco: Juros ===
        juros_frame = tk.LabelFrame(tab, text=" Juros de Mora ",
                                   bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                   font=("Verdana", 8, "bold"),
                                   bd=1, relief="solid")
        juros_frame.pack(fill="x", pady=(0, 6))
        juros_inner = tk.Frame(juros_frame, bg=COLOR_PANEL)
        juros_inner.pack(fill="x", padx=8, pady=6)

        self.cob_juros_var = tk.BooleanVar(value=True)
        tk.Checkbutton(juros_inner, text="Aplicar juros",
                      variable=self.cob_juros_var, bg=COLOR_PANEL,
                      fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
                      activebackground=COLOR_PANEL, selectcolor="white",
                      command=self._cob_toggle_juros).grid(
            row=0, column=0, sticky="w", padx=4)

        self.e_cob_juros_pct = _entry_make(juros_inner, width=6, placeholder="1")
        self.e_cob_juros_pct.grid(row=0, column=1, sticky="w", padx=4)
        _entry_set(self.e_cob_juros_pct, "1")
        tk.Label(juros_inner, text="% a.m.  a partir de:", bg=COLOR_PANEL,
                fg=COLOR_SUBTLE, font=("Verdana", 8)).grid(
            row=0, column=2, sticky="w", padx=(0, 4))

        self.cob_juros_desde = tk.StringVar(value="notificacao")
        self.rb_juros_notif = tk.Radiobutton(
            juros_inner, text="Data de Notificação:",
            variable=self.cob_juros_desde, value="notificacao",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white")
        self.rb_juros_notif.grid(row=0, column=3, sticky="w", padx=4)
        self.e_cob_data_notif = _entry_make(juros_inner, width=14,
                                           placeholder="DD/MM/AAAA")
        self.e_cob_data_notif.grid(row=0, column=4, sticky="w", padx=4)
        _mask_full_date(self.e_cob_data_notif)
        self.rb_juros_orig = tk.Radiobutton(
            juros_inner, text="Data de origem do débito",
            variable=self.cob_juros_desde, value="origem",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white")
        self.rb_juros_orig.grid(row=0, column=5, sticky="w", padx=4)

        # Aviso explicativo sobre quando usar os juros
        tk.Label(juros_inner,
                text="ℹ Juros de mora habilitados por padrão. Correm a partir da "
                "Data de Notificação (Portaria IPREM 51/2022, art. 28). Informe "
                "a data ao lado. Se ainda não houve notificação formal, desmarque "
                "ou selecione \"Data de origem do débito\".",
                bg=COLOR_PANEL, fg="#9C7700", font=("Verdana", 7),
                wraplength=820, justify="left").grid(
            row=1, column=0, columnspan=6, sticky="w", padx=4, pady=(4, 0))

        # === Bloco: Honorários ===
        honor_frame = tk.LabelFrame(tab, text=" Honorários ",
                                   bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                   font=("Verdana", 8, "bold"),
                                   bd=1, relief="solid")
        honor_frame.pack(fill="x", pady=(0, 6))
        honor_inner = tk.Frame(honor_frame, bg=COLOR_PANEL)
        honor_inner.pack(fill="x", padx=8, pady=6)

        self.cob_honor_var = tk.BooleanVar(value=False)
        tk.Checkbutton(honor_inner, text="Aplicar honorários",
                      variable=self.cob_honor_var, bg=COLOR_PANEL,
                      fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
                      activebackground=COLOR_PANEL, selectcolor="white",
                      command=self._cob_toggle_honor).grid(
            row=0, column=0, sticky="w", padx=4)
        self.e_cob_honor_pct = _entry_make(honor_inner, width=6, placeholder="10")
        self.e_cob_honor_pct.grid(row=0, column=1, sticky="w", padx=4)
        _entry_set(self.e_cob_honor_pct, "10")
        self.e_cob_honor_pct.configure(state="disabled")
        tk.Label(honor_inner, text="% sobre subtotal (atualizado + multa + juros)",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 8)).grid(
            row=0, column=2, sticky="w", padx=(0, 4))

        # === Bloco: Abatimentos (parcelas + 13º) ===
        abat_frame = tk.LabelFrame(tab, text=" Abatimentos (opcionais) ",
                                  bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                  font=("Verdana", 8, "bold"),
                                  bd=1, relief="solid")
        abat_frame.pack(fill="x", pady=(0, 6))
        abat_inner = tk.Frame(abat_frame, bg=COLOR_PANEL)
        abat_inner.pack(fill="x", padx=8, pady=6)

        # Parcelas pagas
        self.cob_parcelas_var = tk.BooleanVar(value=False)
        tk.Checkbutton(abat_inner, text="Abater parcelas pagas",
                      variable=self.cob_parcelas_var, bg=COLOR_PANEL,
                      fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
                      activebackground=COLOR_PANEL, selectcolor="white",
                      command=self._cob_toggle_parcelas).grid(
            row=0, column=0, sticky="w", padx=4)

        self.cob_parcelas_corrigir = tk.StringVar(value="sim")
        self.rb_parc_sim = tk.Radiobutton(
            abat_inner, text="Corrigir cada parcela (IPCA da data de pgto)",
            variable=self.cob_parcelas_corrigir, value="sim",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white",
            state="disabled")
        self.rb_parc_sim.grid(row=0, column=1, sticky="w", padx=(20, 4))
        self.rb_parc_nao = tk.Radiobutton(
            abat_inner, text="Soma simples (sem correção)",
            variable=self.cob_parcelas_corrigir, value="nao",
            bg=COLOR_PANEL, fg=COLOR_BCB_BLUE, font=("Verdana", 8),
            activebackground=COLOR_PANEL, selectcolor="white",
            state="disabled")
        self.rb_parc_nao.grid(row=0, column=2, sticky="w", padx=4)

        # Tabelinha de parcelas
        self.parc_frame = tk.Frame(abat_inner, bg=COLOR_PANEL)
        self.parc_frame.grid(row=1, column=0, columnspan=3, sticky="ew",
                            padx=4, pady=(6, 0))

        hdr_parc = tk.Frame(self.parc_frame, bg=COLOR_BCB_BLUE)
        hdr_parc.pack(fill="x")
        tk.Label(hdr_parc, text="#", bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 8, "bold"), width=4, padx=4, pady=3).pack(side="left")
        tk.Label(hdr_parc, text="Data Pagamento (DD/MM/AAAA)",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 8, "bold"), width=28, padx=4, pady=3).pack(side="left", padx=1)
        tk.Label(hdr_parc, text="Valor (R$)",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 8, "bold"), width=18, padx=4, pady=3).pack(side="left", padx=1)
        tk.Label(hdr_parc, text="",
                bg=COLOR_BCB_BLUE, width=6).pack(side="left")

        self.parc_rows_frame = tk.Frame(self.parc_frame, bg=COLOR_PANEL)
        self.parc_rows_frame.pack(fill="x")
        self.cob_parcelas = []  # lista de dicts {data, valor, frame, lbl}

        btn_parc = tk.Frame(self.parc_frame, bg=COLOR_PANEL)
        btn_parc.pack(fill="x", pady=(2, 0))
        self.btn_add_parc = ttk.Button(btn_parc, text="+ Adicionar parcela",
                                       style="BCBSmall.TButton",
                                       command=self._cob_add_parcela)
        self.btn_add_parc.pack(side="left")
        self.btn_add_parc.configure(state="disabled")

        # Linha 2: 13º crédito
        sep = ttk.Separator(abat_inner, orient="horizontal")
        sep.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(8, 6))

        self.cob_c13_var = tk.BooleanVar(value=False)
        tk.Checkbutton(abat_inner, text="13º Salário (IPC-FIPE)",
                      variable=self.cob_c13_var, bg=COLOR_PANEL,
                      fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold"),
                      activebackground=COLOR_PANEL, selectcolor="white",
                      command=self._cob_toggle_c13).grid(
            row=3, column=0, sticky="w", padx=4)
        c13_inner = tk.Frame(abat_inner, bg=COLOR_PANEL)
        c13_inner.grid(row=3, column=1, columnspan=2, sticky="w", padx=4)

        # Operação: Somar ou Subtrair
        self.cob_c13_op = tk.StringVar(value="subtrair")
        self.rb_c13_subtr = tk.Radiobutton(
            c13_inner, text="Subtrair (−)", variable=self.cob_c13_op,
            value="subtrair", bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            font=("Verdana", 8), activebackground=COLOR_PANEL,
            selectcolor="white", state="disabled")
        self.rb_c13_subtr.pack(side="left", padx=(0, 4))
        self.rb_c13_somar = tk.Radiobutton(
            c13_inner, text="Somar (+)", variable=self.cob_c13_op,
            value="somar", bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
            font=("Verdana", 8), activebackground=COLOR_PANEL,
            selectcolor="white", state="disabled")
        self.rb_c13_somar.pack(side="left", padx=(0, 12))

        tk.Label(c13_inner, text="Data do Fato Gerador:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 8, "bold")).pack(side="left")
        self.e_cob_c13_data = _entry_make(c13_inner, width=14,
                                         placeholder="DD/MM/AAAA")
        self.e_cob_c13_data.pack(side="left", padx=4)
        _mask_full_date(self.e_cob_c13_data)
        self.e_cob_c13_data.configure(state="disabled")
        tk.Label(c13_inner, text="Valor (R$):", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 8, "bold")).pack(
            side="left", padx=(10, 4))
        self.e_cob_c13_valor = _entry_make(c13_inner, width=14, placeholder="0,00")
        self.e_cob_c13_valor.pack(side="left", padx=4)
        _bind_valor_format(self.e_cob_c13_valor)
        self.e_cob_c13_valor.configure(state="disabled")

        # === Botão Calcular ===
        btn_bar = tk.Frame(tab, bg=COLOR_PANEL)
        btn_bar.pack(fill="x", pady=(4, 4))
        self.btn_cob_calc = ttk.Button(btn_bar, text="▶ Calcular cobrança",
                                       style="BCB.TButton",
                                       command=self._cob_calc)
        self.btn_cob_calc.pack(side="right")
        ttk.Button(btn_bar, text="Limpar tudo", style="BCBSecondary.TButton",
                  command=self._cob_clear).pack(side="right", padx=(0, 6))
        # Botão à esquerda: abrir simulador IRPF
        ttk.Button(btn_bar, text="🧮  Simulador IRPF (Receita Federal)",
                  style="BCBSecondary.TButton",
                  command=self._abrir_simulador_irpf).pack(side="left")

        # === Resultado ===
        self.cob_result_frame = tk.Frame(tab, bg=COLOR_PANEL,
                                        highlightthickness=1,
                                        highlightbackground="#cccccc")
        self.cob_result_frame.pack(fill="both", expand=True, pady=(4, 0))

        self.cob_resultado = None

    def _abrir_simulador_irpf(self):
        """Abre janela modal com o simulador de IRPF (alíquotas efetivas
        oficiais da Receita Federal). Resultado pode ser copiado para
        a área de transferência ou exportado em PDF."""
        win = tk.Toplevel(self)
        win.title("Simulador IRPF — Alíquota Efetiva (Receita Federal)")
        win.configure(bg=COLOR_PANEL)
        win.transient(self)
        win.grab_set()
        # Dimensiona a janela sem ultrapassar a tela (importante em notebooks,
        # onde 780 px de altura não cabem). Como o corpo já tem rolagem, o
        # conteúdo continua todo acessível.
        try:
            self.update_idletasks()
            sw = win.winfo_screenwidth()
            sh = win.winfo_screenheight()
            larg = min(760, max(640, sw - 80))
            alt = min(780, max(480, sh - 120))
            x = self.winfo_rootx() + 40
            y = self.winfo_rooty() + 20
            # Não deixar a janela nascer fora da tela
            if y + alt > sh:
                y = max(0, sh - alt - 40)
            if x + larg > sw:
                x = max(0, sw - larg - 20)
            win.geometry(f"{larg}x{alt}+{x}+{y}")
        except Exception:
            win.geometry("760x600")
        win.minsize(620, 460)

        # Cabeçalho azul
        hdr = tk.Frame(win, bg=COLOR_BCB_BLUE)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Simulador de Alíquota Efetiva — IRPF",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 11, "bold"), padx=14, pady=10).pack(anchor="w")
        tk.Label(hdr, text="Reproduz o cálculo do simulador oficial da Receita "
                "Federal (https://www27.receita.fazenda.gov.br/simulador-irpf/). "
                "Usa as tabelas progressivas mensais publicadas em lei.",
                bg=COLOR_BCB_BLUE, fg="#cfe0f5",
                font=("Verdana", 8), wraplength=680, justify="left",
                padx=14, pady=8).pack(anchor="w", pady=(0, 8))

        # Body rolável (Canvas + Scrollbar)
        outer = tk.Frame(win, bg=COLOR_PANEL)
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, bg=COLOR_PANEL, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical",
                                  command=canvas.yview)
        body = tk.Frame(canvas, bg=COLOR_PANEL, padx=18, pady=14)

        body.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        body_window = canvas.create_window((0, 0), window=body, anchor="nw")

        def _resize_body(event):
            canvas.itemconfig(body_window, width=event.width)
        canvas.bind("<Configure>", _resize_body)

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll com a roda do mouse
        def _on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # Limpar binding quando popup fechar
        def _on_close():
            canvas.unbind_all("<MouseWheel>")
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_close)

        # === Bloco entradas ===
        ent_frame = tk.LabelFrame(body, text=" Dados informados ",
                                 bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                                 font=("Verdana", 8, "bold"))
        ent_frame.pack(fill="x", pady=(0, 10))

        grid = tk.Frame(ent_frame, bg=COLOR_PANEL)
        grid.pack(fill="x", padx=12, pady=10)

        # Ano e mês
        tk.Label(grid, text="Ano-calendário:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=0, sticky="e", padx=4, pady=4)
        anos_disp = ["2023", "2024", "2025", "2026"]
        self.cb_irpf_ano = ttk.Combobox(grid, values=anos_disp,
                                       state="readonly", width=10)
        self.cb_irpf_ano.set("2026")
        self.cb_irpf_ano.grid(row=0, column=1, sticky="w", padx=4, pady=4)

        tk.Label(grid, text="Mês:", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=0, column=2, sticky="e", padx=(20, 4), pady=4)
        meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio",
                      "Junho", "Julho", "Agosto", "Setembro", "Outubro",
                      "Novembro", "Dezembro"]
        self.cb_irpf_mes = ttk.Combobox(grid, values=meses_nomes,
                                       state="readonly", width=14)
        self.cb_irpf_mes.set(meses_nomes[date.today().month - 1])
        self.cb_irpf_mes.grid(row=0, column=3, sticky="w", padx=4, pady=4)

        # Rendimentos
        tk.Label(grid, text="1. Rendimentos tributáveis:",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9, "bold")).grid(
            row=1, column=0, columnspan=2, sticky="e", padx=4, pady=4)
        self.e_irpf_rend = _entry_make(grid, width=16, placeholder="0,00")
        self.e_irpf_rend.grid(row=1, column=2, columnspan=2,
                             sticky="w", padx=4, pady=4)
        _bind_valor_format(self.e_irpf_rend)
        tk.Label(grid,
                text="(salário, pensão etc — sem incluir parcela isenta de aposentado 65+)",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
            row=2, column=0, columnspan=4, sticky="w", padx=4)

        # Linha divisória
        ttk.Separator(grid, orient="horizontal").grid(
            row=3, column=0, columnspan=4, sticky="ew", pady=(8, 4))

        tk.Label(grid, text="2. Deduções", bg=COLOR_PANEL,
                fg=COLOR_BCB_BLUE, font=("Verdana", 9, "bold")).grid(
            row=4, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 4))

        # Previdência oficial
        tk.Label(grid, text="2.1 Previdência oficial:",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9)).grid(
            row=5, column=0, columnspan=2, sticky="e", padx=4, pady=2)
        self.e_irpf_prev = _entry_make(grid, width=16, placeholder="0,00")
        self.e_irpf_prev.grid(row=5, column=2, columnspan=2,
                             sticky="w", padx=4, pady=2)
        _bind_valor_format(self.e_irpf_prev)

        # Dependentes
        tk.Label(grid, text="2.2 Dependentes (quantidade):",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9)).grid(
            row=6, column=0, columnspan=2, sticky="e", padx=4, pady=2)
        self.e_irpf_deps = _entry_make(grid, width=8, placeholder="0")
        self.e_irpf_deps.grid(row=6, column=2, sticky="w", padx=4, pady=2)
        _entry_set(self.e_irpf_deps, "0")
        tk.Label(grid, text="× R$ 189,59 mensais por dependente",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
            row=6, column=3, sticky="w", padx=4)

        # Pensão alimentícia
        tk.Label(grid, text="2.3 Pensão alimentícia:",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9)).grid(
            row=7, column=0, columnspan=2, sticky="e", padx=4, pady=2)
        self.e_irpf_pensao = _entry_make(grid, width=16, placeholder="0,00")
        self.e_irpf_pensao.grid(row=7, column=2, columnspan=2,
                               sticky="w", padx=4, pady=2)
        _bind_valor_format(self.e_irpf_pensao)

        # Outras deduções
        tk.Label(grid, text="2.4 Outras deduções:",
                bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9)).grid(
            row=8, column=0, columnspan=2, sticky="e", padx=4, pady=2)
        self.e_irpf_outras = _entry_make(grid, width=16, placeholder="0,00")
        self.e_irpf_outras.grid(row=8, column=2, columnspan=2,
                               sticky="w", padx=4, pady=2)
        _bind_valor_format(self.e_irpf_outras)
        tk.Label(grid,
                text="(Previdência privada, Funpresp, FAPI, Carnê-Leão, Livro Caixa)",
                bg=COLOR_PANEL, fg=COLOR_SUBTLE, font=("Verdana", 7)).grid(
            row=9, column=0, columnspan=4, sticky="w", padx=4)

        # Botões
        btn_bar = tk.Frame(body, bg=COLOR_PANEL)
        btn_bar.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_bar, text="▶ Calcular IRPF",
                  style="BCB.TButton",
                  command=lambda: self._irpf_calc(win)).pack(side="right")
        ttk.Button(btn_bar, text="Limpar",
                  style="BCBSecondary.TButton",
                  command=self._irpf_limpar).pack(side="right", padx=(0, 6))
        ttk.Button(btn_bar, text="Fechar",
                  style="BCBSecondary.TButton",
                  command=win.destroy).pack(side="left")

        # Resultado
        self.irpf_result_frame = tk.Frame(body, bg=COLOR_PANEL,
                                         highlightthickness=1,
                                         highlightbackground="#cccccc")
        self.irpf_result_frame.pack(fill="both", expand=True)
        self._irpf_win = win

        # Aplica o tamanho de fonte escolhido (Pequeno/Médio/Grande) também
        # nesta janela, registrando seus widgets e reescalando.
        if getattr(self, "_tamanho_atual", "auto") != "auto":
            self.after(50, lambda: self._aplicar_tamanho(
                self._tamanho_atual, salvar=False))

    def _irpf_limpar(self):
        for e in [self.e_irpf_rend, self.e_irpf_prev, self.e_irpf_pensao,
                 self.e_irpf_outras]:
            _entry_set(e, "")
        _entry_set(self.e_irpf_deps, "0")
        for w in self.irpf_result_frame.winfo_children():
            w.destroy()

    def _irpf_calc(self, win):
        # Coleta
        try:
            ano = int(self.cb_irpf_ano.get())
        except Exception:
            return messagebox.showerror("Validação", "Ano inválido.", parent=win)
        meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio",
                      "Junho", "Julho", "Agosto", "Setembro", "Outubro",
                      "Novembro", "Dezembro"]
        try:
            mes = meses_nomes.index(self.cb_irpf_mes.get()) + 1
        except Exception:
            return messagebox.showerror("Validação", "Mês inválido.", parent=win)

        def _num(entry, label, obrig=True):
            v = _entry_value(entry).strip()
            if not v:
                if obrig:
                    raise ValueError(f"Informe o campo '{label}'.")
                return Decimal("0")
            val = parse_valor_br(v)
            if val is None or val < 0:
                raise ValueError(f"Valor inválido em '{label}'.")
            return val

        try:
            rend = _num(self.e_irpf_rend, "Rendimentos tributáveis")
            prev = _num(self.e_irpf_prev, "Previdência oficial", obrig=False)
            pensao = _num(self.e_irpf_pensao, "Pensão alimentícia", obrig=False)
            outras = _num(self.e_irpf_outras, "Outras deduções", obrig=False)
            deps_s = _entry_value(self.e_irpf_deps).strip() or "0"
            deps = int(deps_s)
            if deps < 0:
                raise ValueError("Dependentes não pode ser negativo.")
        except ValueError as e:
            return messagebox.showerror("Validação", str(e), parent=win)
        except Exception:
            return messagebox.showerror("Validação",
                "Quantidade de dependentes inválida.", parent=win)

        try:
            res = calcular_irpf({
                "ano": ano, "mes": mes,
                "rendimentos": rend,
                "previdencia_oficial": prev,
                "dependentes": deps,
                "pensao_alimenticia": pensao,
                "outras_deducoes": outras,
            })
        except Exception as e:
            return messagebox.showerror("Erro no cálculo", str(e), parent=win)

        self._render_irpf_result(res)

    def _render_irpf_result(self, res):
        for w in self.irpf_result_frame.winfo_children():
            w.destroy()

        # Cabeçalho do resultado
        tk.Label(self.irpf_result_frame,
                text=f"Cálculo IRPF — {res['mes']:02d}/{res['ano']}",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 10, "bold"),
                padx=10, pady=6, anchor="w").pack(fill="x")

        body = tk.Frame(self.irpf_result_frame, bg=COLOR_RESULT_BG)
        body.pack(fill="both", expand=True, padx=2, pady=2)

        grid = tk.Frame(body, bg=COLOR_RESULT_BG)
        grid.pack(fill="x", padx=14, pady=10)
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)

        def linha(r, label, value, bold=False, color=COLOR_BCB_BLUE):
            tk.Label(grid, text=label, bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                    font=("Verdana", 9), anchor="w").grid(
                row=r, column=0, sticky="ew", padx=6, pady=1)
            tk.Label(grid, text=value, bg=COLOR_RESULT_BG, fg=color,
                    font=("Consolas", 10, "bold" if bold else "normal"),
                    anchor="e").grid(
                row=r, column=1, sticky="ew", padx=6, pady=1)

        r = 0
        linha(r, "Rendimentos tributáveis:",
              f"R$ {fmt_brl(res['rendimentos'])}"); r += 1
        linha(r, "Previdência oficial:",
              f"R$ {fmt_brl(res['previdencia_oficial'])}"); r += 1
        if res['dependentes'] > 0:
            linha(r, f"Dependentes ({res['dependentes']} × R$ 189,59):",
                  f"R$ {fmt_brl(res['deducao_dependentes'])}"); r += 1
        if res['pensao_alimenticia'] > 0:
            linha(r, "Pensão alimentícia:",
                  f"R$ {fmt_brl(res['pensao_alimenticia'])}"); r += 1
        if res['outras_deducoes'] > 0:
            linha(r, "Outras deduções:",
                  f"R$ {fmt_brl(res['outras_deducoes'])}"); r += 1

        linha(r, "Total deduções (completo):",
              f"R$ {fmt_brl(res['deducoes_completas'])}"); r += 1
        if res['desconto_simplificado_disponivel'] > 0:
            linha(r, "Desconto simplificado disponível:",
                  f"R$ {fmt_brl(res['desconto_simplificado_disponivel'])}"); r += 1
        regime_str = ("Simplificado (mais benéfico)" if res['regime'] == "simplificado"
                     else "Completo (mais benéfico)")
        linha(r, "Regime adotado:", regime_str, bold=True); r += 1
        linha(r, "Dedução utilizada:",
              f"R$ {fmt_brl(res['deducao_utilizada'])}"); r += 1

        ttk.Separator(body, orient="horizontal").pack(fill="x", padx=14)

        grid2 = tk.Frame(body, bg=COLOR_RESULT_BG)
        grid2.pack(fill="x", padx=14, pady=10)
        grid2.columnconfigure(0, weight=1)
        grid2.columnconfigure(1, weight=1)

        def linha2(r, label, value, bold=False, color=COLOR_BCB_BLUE):
            tk.Label(grid2, text=label, bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                    font=("Verdana", 9), anchor="w").grid(
                row=r, column=0, sticky="ew", padx=6, pady=1)
            tk.Label(grid2, text=value, bg=COLOR_RESULT_BG, fg=color,
                    font=("Consolas", 10, "bold" if bold else "normal"),
                    anchor="e").grid(
                row=r, column=1, sticky="ew", padx=6, pady=1)

        r = 0
        linha2(r, "3. Base de cálculo (1 − 2):",
              f"R$ {fmt_brl(res['base_calculo'])}", bold=True); r += 1

        # === Demonstrativo da Apuração do Imposto (5 faixas) ===
        sep = ttk.Separator(body, orient="horizontal")
        sep.pack(fill="x", padx=14, pady=(4, 6))

        tk.Label(body, text="4. Demonstrativo da Apuração do Imposto",
                bg=COLOR_RESULT_BG, fg=COLOR_BCB_BLUE,
                font=("Verdana", 9, "bold"),
                padx=14).pack(anchor="w")

        tbl = tk.Frame(body, bg=COLOR_RESULT_BG)
        tbl.pack(fill="x", padx=14, pady=(2, 8))

        # Cabeçalho da tabela
        hdr_bg = "#D5E8F0"
        col_widths = [16, 16, 12, 16]
        cols = ["Faixa da Base de Cálculo", "Base na faixa", "Alíquota",
               "Valor do Imposto"]
        for j, label in enumerate(cols):
            tk.Label(tbl, text=label, bg=hdr_bg, fg=COLOR_BCB_BLUE,
                    font=("Verdana", 8, "bold"),
                    padx=6, pady=4, width=col_widths[j],
                    relief="solid", bd=1).grid(row=0, column=j, sticky="ew")

        # Linhas (5 faixas)
        for i, f in enumerate(res['demonstrativo_faixas']):
            row = i + 1
            faixa_nome = ["1ª Faixa", "2ª Faixa", "3ª Faixa",
                         "4ª Faixa", "5ª Faixa"][i]
            tk.Label(tbl, text=faixa_nome, bg="white", fg=COLOR_TEXT,
                    font=("Verdana", 8), padx=6, pady=3,
                    width=col_widths[0], relief="solid", bd=1,
                    anchor="w").grid(row=row, column=0, sticky="ew")
            tk.Label(tbl, text=f"R$ {fmt_brl(f['base_faixa'])}",
                    bg="white", fg=COLOR_BCB_BLUE,
                    font=("Consolas", 9), padx=6, pady=3,
                    width=col_widths[1], relief="solid", bd=1,
                    anchor="e").grid(row=row, column=1, sticky="ew")
            aliq_str = f"{float(f['aliquota']) * 100:.1f}%".replace(".", ",")
            tk.Label(tbl, text=aliq_str, bg="white", fg=COLOR_BCB_BLUE,
                    font=("Consolas", 9), padx=6, pady=3,
                    width=col_widths[2], relief="solid", bd=1,
                    anchor="center").grid(row=row, column=2, sticky="ew")
            tk.Label(tbl, text=f"R$ {fmt_brl(f['valor_imposto'])}",
                    bg="white", fg=COLOR_BCB_BLUE,
                    font=("Consolas", 9), padx=6, pady=3,
                    width=col_widths[3], relief="solid", bd=1,
                    anchor="e").grid(row=row, column=3, sticky="ew")

        # Linha Total
        row_total = len(res['demonstrativo_faixas']) + 1
        tk.Label(tbl, text="Total", bg=hdr_bg, fg=COLOR_BCB_BLUE,
                font=("Verdana", 8, "bold"), padx=6, pady=4,
                width=col_widths[0], relief="solid", bd=1,
                anchor="w").grid(row=row_total, column=0, sticky="ew")
        tk.Label(tbl, text="", bg=hdr_bg, padx=6, pady=4,
                width=col_widths[1], relief="solid",
                bd=1).grid(row=row_total, column=1, sticky="ew")
        tk.Label(tbl, text="—", bg=hdr_bg, fg=COLOR_BCB_BLUE,
                font=("Verdana", 8, "bold"), padx=6, pady=4,
                width=col_widths[2], relief="solid", bd=1,
                anchor="center").grid(row=row_total, column=2, sticky="ew")
        tk.Label(tbl, text=f"R$ {fmt_brl(res['imposto_bruto'])}",
                bg=hdr_bg, fg=COLOR_BCB_BLUE,
                font=("Consolas", 9, "bold"), padx=6, pady=4,
                width=col_widths[3], relief="solid", bd=1,
                anchor="e").grid(row=row_total, column=3, sticky="ew")

        # Parcela a deduzir (forma "simplificada" - validação)
        aliq_faixa_str = f"{float(res['aliquota_faixa']) * 100:.1f}%".replace(".", ",")
        tk.Label(body,
                text=f"  Forma simplificada: base × {aliq_faixa_str} "
                f"− parcela a deduzir R$ {fmt_brl(res['parcela_a_deduzir'])} "
                f"= R$ {fmt_brl(res['imposto_bruto'])}",
                bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                font=("Verdana", 7, "italic"),
                padx=14).pack(anchor="w", pady=(0, 4))

        # Redutor Lei 15.270/2025 (se aplicável)
        if res['redutor_lei_15270'] > 0:
            red_box = tk.Frame(body, bg="#FFF6D5",
                              highlightthickness=1,
                              highlightbackground="#C4A93B")
            red_box.pack(fill="x", padx=14, pady=(0, 6))
            tk.Label(red_box,
                    text=" Redutor (Lei nº 15.270/2025 - Reforma da Renda):",
                    bg="#FFF6D5", fg="#9C7700",
                    font=("Verdana", 9, "bold"),
                    padx=10, pady=6).pack(side="left")
            tk.Label(red_box,
                    text=f"−  R$ {fmt_brl(res['redutor_lei_15270'])}",
                    bg="#FFF6D5", fg="#9C7700",
                    font=("Consolas", 11, "bold"),
                    padx=10, pady=6).pack(side="right")

        # Imposto destacado
        imp_box = tk.Frame(body, bg=COLOR_BCB_BLUE)
        imp_box.pack(fill="x", padx=14, pady=(4, 4))
        label_imp = ("4. Imposto devido (após redutor)"
                    if res['redutor_lei_15270'] > 0
                    else "4. Imposto devido")
        tk.Label(imp_box, text=label_imp,
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 10, "bold"), padx=12, pady=6).pack(side="left")
        tk.Label(imp_box, text=f"R$ {fmt_brl(res['imposto'])}",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Consolas", 14, "bold"), padx=12, pady=6).pack(side="right")

        # Alíquota efetiva destacada
        aef_box = tk.Frame(body, bg="#E8F4E8")
        aef_box.pack(fill="x", padx=14, pady=(0, 10))
        aliq_ef_str = f"{res['aliquota_efetiva_pct']}%".replace(".", ",")
        tk.Label(aef_box, text="5. Alíquota efetiva (% do imposto sobre rendimentos)",
                bg="#E8F4E8", fg="#2A7A2A",
                font=("Verdana", 9, "bold"), padx=12, pady=6).pack(side="left")
        tk.Label(aef_box, text=aliq_ef_str,
                bg="#E8F4E8", fg="#2A7A2A",
                font=("Consolas", 13, "bold"), padx=12, pady=6).pack(side="right")

    def _cob_toggle_multa(self):
        st = "normal" if self.cob_multa_var.get() else "disabled"
        self.e_cob_multa_pct.configure(state=st)
        self.rb_multa_atu.configure(state=st)
        self.rb_multa_ori.configure(state=st)

    def _cob_toggle_juros(self):
        st = "normal" if self.cob_juros_var.get() else "disabled"
        self.e_cob_juros_pct.configure(state=st)
        self.rb_juros_notif.configure(state=st)
        self.rb_juros_orig.configure(state=st)
        self.e_cob_data_notif.configure(state=st)

    def _cob_toggle_honor(self):
        st = "normal" if self.cob_honor_var.get() else "disabled"
        self.e_cob_honor_pct.configure(state=st)

    def _cob_toggle_parcelas(self):
        st = "normal" if self.cob_parcelas_var.get() else "disabled"
        self.rb_parc_sim.configure(state=st)
        self.rb_parc_nao.configure(state=st)
        self.btn_add_parc.configure(state=st)
        for p in self.cob_parcelas:
            p["data"].configure(state=st)
            p["valor"].configure(state=st)
            p["btn_rm"].configure(state=st)

    def _cob_toggle_c13(self):
        st = "normal" if self.cob_c13_var.get() else "disabled"
        self.e_cob_c13_data.configure(state=st)
        self.e_cob_c13_valor.configure(state=st)
        self.rb_c13_somar.configure(state=st)
        self.rb_c13_subtr.configure(state=st)

    def _cob_add_parcela(self):
        rowf = tk.Frame(self.parc_rows_frame, bg=COLOR_PANEL)
        rowf.pack(fill="x", padx=2, pady=1)

        num = len(self.cob_parcelas) + 1
        lbl = tk.Label(rowf, text=str(num), bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                      font=("Verdana", 8), width=4)
        lbl.pack(side="left", padx=2)

        e_data = _entry_make(rowf, width=24, placeholder="DD/MM/AAAA")
        e_data.pack(side="left", padx=1)
        _mask_full_date(e_data)

        e_valor = _entry_make(rowf, width=14, placeholder="0,00")
        e_valor.pack(side="left", padx=1)
        _bind_valor_format(e_valor)

        row_data = {"frame": rowf, "lbl": lbl, "data": e_data, "valor": e_valor}

        btn_rm = tk.Button(
            rowf, text="✕", bg="#fdecea", fg=COLOR_ERROR,
            font=("Verdana", 8, "bold"), bd=0, padx=6, pady=1,
            command=lambda r=row_data: self._cob_remove_parcela(r),
            cursor="hand2")
        btn_rm.pack(side="left", padx=4)
        row_data["btn_rm"] = btn_rm

        self.cob_parcelas.append(row_data)

    def _cob_remove_parcela(self, row_data):
        row_data["frame"].destroy()
        self.cob_parcelas.remove(row_data)
        for i, r in enumerate(self.cob_parcelas, start=1):
            r["lbl"].config(text=str(i))

    # ---- Várias competências (opcional) ----
    def _cob_toggle_multimes(self):
        ativo = self.cob_multimes_var.get()
        estado = "normal" if ativo else "disabled"
        try:
            self.btn_cob_add_comp.config(state=estado)
        except Exception:
            pass
        # Habilita/desabilita os campos das linhas existentes
        for r in getattr(self, "cob_comp_rows", []):
            for k in ("comp", "desc", "valor"):
                try:
                    r[k].config(state=estado)
                except Exception:
                    pass
            try:
                r["btn_rm"].config(state=estado)
            except Exception:
                pass
        # Ao ativar pela 1ª vez, já cria uma linha em branco
        if ativo and not self.cob_comp_rows:
            self._cob_add_comp_row()
        # Habilita/desabilita os campos do Débito Original (modo único)
        estado_unico = "disabled" if ativo else "normal"
        for w in (self.e_cob_data_origem, self.e_cob_valor_origem):
            try:
                w.config(state=estado_unico)
            except Exception:
                pass

    def _cob_add_comp_row(self):
        rowf = tk.Frame(self.cob_comp_rows_frame, bg=COLOR_PANEL)
        rowf.pack(fill="x", padx=2, pady=1)

        num = len(self.cob_comp_rows) + 1
        lbl = tk.Label(rowf, text=str(num), bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                      font=("Verdana", 8), width=4)
        lbl.pack(side="left", padx=2)

        e_comp = _entry_make(rowf, width=22, placeholder="MM/AAAA")
        e_comp.pack(side="left", padx=1)
        _mask_month_year(e_comp)

        e_desc = _entry_make(rowf, width=26, placeholder="(opcional)")
        e_desc.pack(side="left", padx=1)

        e_valor = _entry_make(rowf, width=14, placeholder="0,00")
        e_valor.pack(side="left", padx=1)
        _bind_valor_format(e_valor)

        row_data = {"frame": rowf, "lbl": lbl, "comp": e_comp,
                   "desc": e_desc, "valor": e_valor}

        btn_rm = tk.Button(
            rowf, text="✕", bg="#fdecea", fg=COLOR_ERROR,
            font=("Verdana", 8, "bold"), bd=0, padx=6, pady=1,
            command=lambda r=row_data: self._cob_remove_comp_row(r),
            cursor="hand2")
        btn_rm.pack(side="left", padx=4)
        row_data["btn_rm"] = btn_rm

        self.cob_comp_rows.append(row_data)
        # Reaplica o tamanho de fonte se não estiver em "auto"
        if getattr(self, "_tamanho_atual", "auto") != "auto":
            self.after(10, self.reaplicar_tamanho)

    def _cob_remove_comp_row(self, row_data):
        row_data["frame"].destroy()
        self.cob_comp_rows.remove(row_data)
        for i, r in enumerate(self.cob_comp_rows, start=1):
            r["lbl"].config(text=str(i))

    def _cob_clear(self):
        if not messagebox.askyesno("Limpar tudo",
                "Deseja limpar todos os campos da Cobrança Amigável?"):
            return
        # 1) Limpa o resultado e seus dados
        for w in self.cob_result_frame.winfo_children():
            w.destroy()
        self.cob_resultado = None

        # 2) Débito Original
        for e in (self.e_cob_data_origem, self.e_cob_valor_origem):
            try:
                e.configure(state="normal")
                _entry_set(e, "")
            except Exception:
                pass
        # Data de atualização volta para hoje (default original)
        try:
            _entry_set(self.e_cob_data_atual,
                      date.today().strftime("%d/%m/%Y"))
        except Exception:
            pass
        try:
            self.cb_cob_indice.set("IPCA")
        except Exception:
            pass

        # 3) Várias competências (desmarca, esvazia a tabela)
        try:
            self.cob_multimes_var.set(False)
        except Exception:
            pass
        for r in list(getattr(self, "cob_comp_rows", [])):
            try:
                r["frame"].destroy()
            except Exception:
                pass
        self.cob_comp_rows = []
        # Reaplica o estado (libera o Débito Original e oculta as linhas)
        try:
            self._cob_toggle_multimes()
        except Exception:
            pass

        # 4) Multa
        try:
            self.cob_multa_var.set(False)
            _entry_set(self.e_cob_multa_pct, "10")
            self.cob_multa_sobre.set("atualizado")
            self._cob_toggle_multa()
        except Exception:
            pass

        # 5) Juros (volta ao default: ligado, 1%, desde notificação)
        try:
            self.cob_juros_var.set(True)
            _entry_set(self.e_cob_juros_pct, "1")
            self.cob_juros_desde.set("notificacao")
            _entry_set(self.e_cob_data_notif, "")
            self._cob_toggle_juros()
        except Exception:
            pass

        # 6) Honorários
        try:
            self.cob_honor_var.set(False)
            _entry_set(self.e_cob_honor_pct, "10")
            self._cob_toggle_honor()
        except Exception:
            pass

        # 7) Abatimentos: parcelas pagas e 13º
        try:
            self.cob_parcelas_var.set(False)
            self.cob_parcelas_corrigir.set("sim")
        except Exception:
            pass
        for p in list(getattr(self, "cob_parcelas", [])):
            try:
                p["frame"].destroy()
            except Exception:
                pass
        self.cob_parcelas = []
        try:
            self.cob_c13_var.set(False)
            self.cob_c13_op.set("subtrair")
            _entry_set(self.e_cob_c13_data, "")
            _entry_set(self.e_cob_c13_valor, "")
            self._cob_toggle_parcelas()
            self._cob_toggle_c13()
        except Exception:
            pass

    def _cob_calc(self):
        modo_multimes = self.cob_multimes_var.get()

        if modo_multimes:
            # Lê as competências da tabela
            competencias = []
            for i, r in enumerate(self.cob_comp_rows, start=1):
                comp_txt = _entry_value(r["comp"]).strip()
                valor_txt = _entry_value(r["valor"]).strip()
                # Pula linhas totalmente vazias
                if not comp_txt and not valor_txt:
                    continue
                my = parse_month_year(comp_txt)
                if not my:
                    return messagebox.showerror("Validação",
                        f"Competência da linha {i} inválida. Use MM/AAAA.")
                try:
                    cval = parse_valor_br(valor_txt)
                    if not cval or cval <= 0:
                        raise ValueError()
                except Exception:
                    return messagebox.showerror("Validação",
                        f"Valor da linha {i} inválido. Use formato 1.234,56.")
                mes, ano = my
                competencias.append({
                    "data": date(ano, mes, 1),
                    "valor": cval,
                    "descricao": _entry_value(r["desc"]).strip(),
                })
            if not competencias:
                return messagebox.showerror("Validação",
                    "Adicione pelo menos uma competência com mês e valor, "
                    "ou desmarque \"Lançar débito em várias competências\".")
            data_atual = parse_date_br(_entry_value(self.e_cob_data_atual))
            if not data_atual:
                return messagebox.showerror("Validação",
                    "Data de Atualização inválida.")
            # data_origem/valor_origem derivam das competências (na função)
            config = {
                "competencias": competencias,
                "data_origem": min(c["data"] for c in competencias),
                "valor_origem": Decimal("0"),
                "data_atualizacao": data_atual,
                "indice_key": self.cb_cob_indice.get(),
            }
        else:
            # Validação do modo de débito único (comportamento original)
            data_origem = parse_date_br(_entry_value(self.e_cob_data_origem))
            if not data_origem:
                return messagebox.showerror("Validação",
                    "Data de Origem inválida. Use DD/MM/AAAA.")
            try:
                valor_origem = parse_valor_br(_entry_value(self.e_cob_valor_origem))
                if not valor_origem or valor_origem <= 0:
                    raise ValueError()
            except Exception:
                return messagebox.showerror("Validação",
                    "Valor Original inválido. Use formato 1.234,56.")

            data_atual = parse_date_br(_entry_value(self.e_cob_data_atual))
            if not data_atual:
                return messagebox.showerror("Validação",
                    "Data de Atualização inválida.")
            if data_atual < data_origem:
                return messagebox.showerror("Validação",
                    "Data de Atualização deve ser posterior à Data de Origem.")

            config = {
                "data_origem": data_origem,
                "valor_origem": valor_origem,
                "data_atualizacao": data_atual,
                "indice_key": self.cb_cob_indice.get(),
            }

        # Multa
        if self.cob_multa_var.get():
            try:
                pct_str = _entry_value(self.e_cob_multa_pct).replace(",", ".") or "10"
                config["aplicar_multa"] = True
                config["multa_pct"] = Decimal(pct_str) / Decimal("100")
                config["multa_sobre"] = self.cob_multa_sobre.get()
            except Exception:
                return messagebox.showerror("Validação", "Percentual de multa inválido.")

        # Juros
        if self.cob_juros_var.get():
            try:
                pct_str = _entry_value(self.e_cob_juros_pct).replace(",", ".") or "1"
                config["aplicar_juros"] = True
                config["juros_mensais_pct"] = Decimal(pct_str) / Decimal("100")
                config["juros_desde"] = self.cob_juros_desde.get()
                if self.cob_juros_desde.get() == "notificacao":
                    dn = parse_date_br(_entry_value(self.e_cob_data_notif))
                    if not dn:
                        return messagebox.showerror("Validação",
                            "Data de Notificação inválida.")
                    config["data_notificacao"] = dn
            except Exception:
                return messagebox.showerror("Validação", "Percentual de juros inválido.")

        # Honorários
        if self.cob_honor_var.get():
            try:
                pct_str = _entry_value(self.e_cob_honor_pct).replace(",", ".") or "10"
                config["aplicar_honorarios"] = True
                config["honorarios_pct"] = Decimal(pct_str) / Decimal("100")
            except Exception:
                return messagebox.showerror("Validação", "Percentual de honorários inválido.")

        # Parcelas
        if self.cob_parcelas_var.get():
            parcs = []
            for i, p in enumerate(self.cob_parcelas, start=1):
                d_str = _entry_value(p["data"]).strip()
                v_str = _entry_value(p["valor"]).strip()
                if not d_str and not v_str:
                    continue
                dt = parse_date_br(d_str)
                if not dt:
                    return messagebox.showerror("Validação",
                        f"Parcela {i}: data inválida.")
                try:
                    val = parse_valor_br(v_str)
                    if val is None or val <= 0:
                        raise ValueError()
                except Exception:
                    return messagebox.showerror("Validação",
                        f"Parcela {i}: valor inválido.")
                parcs.append({"data": dt, "valor": val})
            config["parcelas_pagas"] = parcs
            config["parcelas_corrigir"] = (self.cob_parcelas_corrigir.get() == "sim")

        # 13º
        if self.cob_c13_var.get():
            d13 = parse_date_br(_entry_value(self.e_cob_c13_data))
            if not d13:
                return messagebox.showerror("Validação",
                    "Data do Fato Gerador do 13º inválida.")
            try:
                v13 = parse_valor_br(_entry_value(self.e_cob_c13_valor))
                if not v13 or v13 <= 0:
                    raise ValueError()
            except Exception:
                return messagebox.showerror("Validação", "Valor do 13º inválido.")
            config["credito_13"] = {
                "data_fato_gerador": d13,
                "valor": v13,
                "operacao": self.cob_c13_op.get(),   # "somar" ou "subtrair"
            }

        # Executar
        self.btn_cob_calc.config(state="disabled")
        self._cob_clear()
        loading = tk.Label(self.cob_result_frame,
                          text="⟳  Calculando...", bg=COLOR_PANEL,
                          fg=COLOR_BCB_BLUE, font=("Verdana", 9, "italic"),
                          pady=20)
        loading.pack()
        limpar_cache_api()

        def progress(msg):
            try:
                self.after(0, lambda: loading.config(text="⟳  " + msg))
            except Exception:
                pass

        def worker():
            try:
                res = calcular_cobranca_amigavel(config, progress_cb=progress)
                self.after(0, lambda: self._render_cob_result(res))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: messagebox.showerror("Erro no cálculo", err))
            finally:
                self.after(0, lambda: self.btn_cob_calc.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _render_cob_result(self, res):
        self._cob_clear()
        self.cob_resultado = res

        box = tk.Frame(self.cob_result_frame, bg=COLOR_RESULT_BG)
        box.pack(fill="both", expand=True, padx=2, pady=2)

        # Título
        tk.Label(box, text="Resultado da Cobrança Amigável",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 10, "bold"),
                padx=10, pady=6, anchor="w").pack(fill="x")

        # Barra de exportação
        toolbar = tk.Frame(box, bg=COLOR_RESULT_BG)
        toolbar.pack(fill="x", padx=10, pady=(8, 0))
        tk.Label(toolbar, text="Exportar:", bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                font=("Verdana", 8)).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="CSV…", style="BCBSmall.TButton",
                  command=self._cob_export_csv).pack(side="left", padx=2)
        if HAS_XLSX:
            ttk.Button(toolbar, text="XLSX…", style="BCBSmall.TButton",
                      command=self._cob_export_xlsx).pack(side="left", padx=2)
        if HAS_PDF:
            ttk.Button(toolbar, text="PDF…", style="BCBSmall.TButton",
                      command=self._cob_export_pdf).pack(side="left", padx=2)

        grid = tk.Frame(box, bg=COLOR_RESULT_BG)
        grid.pack(fill="x", padx=14, pady=12)
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)

        def linha(row, label, value, bold=False, color=COLOR_BCB_BLUE):
            tk.Label(grid, text=label, bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                    font=("Verdana", 9), anchor="w").grid(
                row=row, column=0, sticky="ew", padx=6, pady=2)
            tk.Label(grid, text=value, bg=COLOR_RESULT_BG, fg=color,
                    font=("Consolas", 10, "bold" if bold else "normal"),
                    anchor="e").grid(
                row=row, column=1, sticky="ew", padx=6, pady=2)

        r = 0
        linha(r, "Período de correção:",
              f"{res['data_origem'].strftime('%d/%m/%Y')} → "
              f"{res['data_atualizacao'].strftime('%d/%m/%Y')}"); r += 1
        linha(r, "Índice:", res['indice_nome']); r += 1
        linha(r, "Fator de correção:", fmt_fator(res['fator'])); r += 1
        linha(r, "Valor original:", f"R$ {fmt_brl(res['valor_origem'])}"); r += 1
        linha(r, "Valor atualizado:", f"R$ {fmt_brl(res['valor_atualizado'])}",
              bold=True); r += 1

        if res["aplicar_multa"]:
            sobre = "atualizado" if res["multa_sobre"] == "atualizado" else "original"
            pct = float(res["multa_pct"]) * 100
            linha(r, f"Multa ({pct:.0f}% sobre valor {sobre}):",
                  f"R$ {fmt_brl(res['multa'])}"); r += 1
        if res["aplicar_juros"]:
            pct = float(res["juros_mensais_pct"]) * 100
            linha(r, f"Juros ({pct:.1f}% a.m. × {float(res['meses_atraso_juros']):.2f} meses):".replace(".", ","),
                  f"R$ {fmt_brl(res['juros'])}"); r += 1

        linha(r, "Subtotal:", f"R$ {fmt_brl(res['subtotal'])}", bold=True); r += 1

        if res["aplicar_honorarios"]:
            pct = float(res["honorarios_pct"]) * 100
            linha(r, f"Honorários ({pct:.0f}% sobre subtotal):",
                  f"R$ {fmt_brl(res['honorarios'])}"); r += 1

        if res["parcelas_pagas"]:
            linha(r, f"(−) Parcelas pagas ({len(res['parcelas_pagas'])}):",
                  f"R$ {fmt_brl(res['total_parcelas_corrigido'])}",
                  color=COLOR_ERROR); r += 1
            for i, p in enumerate(res["parcelas_pagas"], start=1):
                linha(r, f"     #{i} pgto {p['data'].strftime('%d/%m/%Y')} "
                      f"(orig R$ {fmt_brl(p['valor'])}, fator {fmt_fator(p['fator'])}):",
                      f"R$ {fmt_brl(p['valor_corrigido'])}",
                      color=COLOR_SUBTLE); r += 1

        if res["credito_13"]:
            c13 = res["credito_13"]
            op_c13 = c13.get("operacao", "subtrair")
            sinal_c13 = "(+)" if op_c13 == "somar" else "(−)"
            cor_c13 = COLOR_VERDE if op_c13 == "somar" else COLOR_ERROR
            linha(r, f"{sinal_c13} 13º salário (IPC-FIPE, fator {fmt_fator(c13['fator'])}):",
                  f"R$ {fmt_brl(c13['valor_corrigido'])}",
                  color=cor_c13); r += 1

        ttk.Separator(box, orient="horizontal").pack(fill="x", padx=14)

        # Total final em destaque
        total_box = tk.Frame(box, bg=COLOR_BCB_BLUE)
        total_box.pack(fill="x", padx=14, pady=(8, 12))
        tk.Label(total_box, text="TOTAL FINAL", bg=COLOR_BCB_BLUE, fg="white",
                font=("Verdana", 11, "bold"), padx=12, pady=8).pack(side="left")
        tk.Label(total_box, text=f"R$ {fmt_brl(res['total'])}",
                bg=COLOR_BCB_BLUE, fg="white",
                font=("Consolas", 16, "bold"), padx=12, pady=8).pack(side="right")

        # Detalhamento mês a mês (colapsável) — só aparece em modo competências
        competencias = res.get("competencias") or []
        if competencias:
            det_header = tk.Frame(box, bg=COLOR_RESULT_BG)
            det_header.pack(fill="x", padx=14, pady=(4, 0))
            det_visible = tk.BooleanVar(value=False)
            det_frame = tk.Frame(box, bg=COLOR_RESULT_BG)

            def _toggle_det(dv=det_visible, df=det_frame, dh=det_header):
                if dv.get():
                    df.pack_forget()
                    dv.set(False)
                    for w in dh.winfo_children():
                        w.destroy()
                    _build_det_btn(dh, dv, df)
                else:
                    dv.set(True)
                    df.pack(fill="x", padx=14, pady=(0, 8))
                    for w in dh.winfo_children():
                        w.destroy()
                    _build_det_btn(dh, dv, df)

            def _build_det_btn(dh, dv, df):
                arrow = "▼" if dv.get() else "▶"
                tk.Button(dh,
                    text=f"{arrow}  Detalhamento mês a mês ({len(competencias)} competências)",
                    bg=COLOR_RESULT_BG, fg=COLOR_BCB_BLUE,
                    font=("Verdana", 8, "underline"),
                    relief="flat", cursor="hand2", bd=0,
                    command=lambda: _toggle_det()
                ).pack(side="left")

            _build_det_btn(det_header, det_visible, det_frame)

            # Preenche a tabela quando expandida
            def _fill_det_table(df=det_frame, comps=competencias, res=res):
                for w in df.winfo_children():
                    w.destroy()
                aplica_juros = res.get("aplicar_juros", False)
                cols = ["Competência", "Descrição", "Valor Original", "Fator",
                        "Valor Corrigido"]
                if aplica_juros:
                    cols += ["Meses Juros", "Juros"]

                tbl = tk.Frame(df, bg=COLOR_RESULT_BG)
                tbl.pack(fill="x")

                hdr_bg = COLOR_TABLE_HEADER
                for ci, ch in enumerate(cols):
                    tk.Label(tbl, text=ch, bg=hdr_bg, fg="white",
                             font=("Verdana", 8, "bold"),
                             padx=6, pady=4, anchor="center",
                             relief="flat").grid(
                        row=0, column=ci, sticky="ew", padx=1, pady=(0, 1))
                    tbl.columnconfigure(ci, weight=1)

                for ri, det in enumerate(comps, start=1):
                    bg = COLOR_TABLE_ALT if ri % 2 == 0 else COLOR_PANEL
                    competencia = det["data"].strftime("%m/%Y")
                    descricao   = det.get("descricao", "") or ""
                    v_orig      = f"R$ {fmt_brl(det['valor'])}"
                    v_fator     = fmt_fator(det["fator"])
                    v_corr      = f"R$ {fmt_brl(det['valor_corrigido'])}"

                    vals = [competencia, descricao, v_orig, v_fator, v_corr]
                    if aplica_juros:
                        mj = det.get("meses_juros", 0)
                        jj = det.get("juros", 0)
                        vals += [f"{float(mj):.0f}", f"R$ {fmt_brl(jj)}"]

                    for ci, val in enumerate(vals):
                        anchor = "e" if ci >= 2 else "w"
                        tk.Label(tbl, text=val, bg=bg,
                                 font=("Consolas", 8),
                                 fg=COLOR_TEXT,
                                 padx=6, pady=3,
                                 anchor=anchor).grid(
                            row=ri, column=ci, sticky="ew", padx=1, pady=0)

            # Monkey-patch toggle para também preencher a tabela
            orig_toggle = _toggle_det
            def _toggle_with_fill(dv=det_visible, df=det_frame, dh=det_header):
                was_visible = dv.get()
                orig_toggle()
                if not was_visible:
                    _fill_det_table(df)
            for w in det_header.winfo_children():
                w.configure(command=_toggle_with_fill)

        # Aviso de saldo credor (total negativo)
        if res["total"] < 0:
            warn = tk.Frame(box, bg="#FFF6D5",
                           highlightthickness=1, highlightbackground="#C4A93B")
            warn.pack(fill="x", padx=14, pady=(0, 12))
            tk.Label(
                warn,
                text=(" ⚠  SALDO CREDOR: a soma de parcelas pagas e "
                      "abatimentos é MAIOR que o débito atualizado. "
                      "Verifique se as parcelas estão corretas — o devedor "
                      "pode ter pago a mais. "),
                bg="#FFF6D5", fg="#9C7700",
                font=("Verdana", 8, "bold"),
                wraplength=900, justify="left", pady=8, padx=8
            ).pack(fill="x")

    def _cob_export_csv(self):
        if not self.cob_resultado:
            return messagebox.showwarning("Exportar",
                "Calcule a cobrança antes de exportar.")
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="cobranca_amigavel.csv")
        if not path:
            return
        try:
            exportar_csv_cobranca(path, self.cob_resultado)
            messagebox.showinfo("Exportar", f"CSV salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def _cob_export_xlsx(self):
        if not self.cob_resultado:
            return messagebox.showwarning("Exportar",
                "Calcule a cobrança antes de exportar.")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="cobranca_amigavel.xlsx")
        if not path:
            return
        try:
            exportar_xlsx_cobranca(path, self.cob_resultado)
            messagebox.showinfo("Exportar", f"XLSX salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def _cob_export_pdf(self):
        if not self.cob_resultado:
            return messagebox.showwarning("Exportar",
                "Calcule a cobrança antes de exportar.")
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="cobranca_amigavel.pdf")
        if not path:
            return
        try:
            exportar_pdf_cobranca(path, self.cob_resultado)
            messagebox.showinfo("Exportar", f"PDF salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    # ===================== Aba: Atraso de Parcela ========================= #

    def _build_atraso_tab(self):
        tab_outer = ttk.Frame(self.notebook, style="BCB.TFrame")
        self.notebook.add(tab_outer, text="Atraso de Parcela")

        # Scroll
        canvas = tk.Canvas(tab_outer, bg=COLOR_PANEL, highlightthickness=0)
        vsb = ttk.Scrollbar(tab_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=COLOR_PANEL)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(event):
            canvas.itemconfig(win_id, width=event.width)
        canvas.bind("<Configure>", _on_resize)

        def _on_frame_configure(_):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_frame_configure)

        def _on_mw(e):
            canvas.yview_scroll(-1 * int(e.delta / 120), "units")
        canvas.bind("<MouseWheel>", _on_mw)
        inner.bind("<MouseWheel>", _on_mw)

        pad = tk.Frame(inner, bg=COLOR_PANEL, padx=20, pady=16)
        pad.pack(fill="both", expand=True)

        # Título
        tk.Label(pad,
                 text="Atraso de Parcela  —  Lei 13.275/2002",
                 bg=COLOR_PANEL, fg=COLOR_BCB_BLUE,
                 font=("Verdana", 12, "bold")).pack(anchor="w", pady=(0, 4))
        tk.Label(pad,
                 text=("Atualiza uma parcela contratual em atraso: "
                       "correção monetária + multa 10% + juros 1% a.m."),
                 bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 8)).pack(anchor="w", pady=(0, 12))
        tk.Frame(pad, bg="#dde3ea", height=1).pack(fill="x", pady=(0, 14))

        # Formulário
        form = tk.Frame(pad, bg=COLOR_PANEL)
        form.pack(fill="x")

        def lbl(parent, txt, row, col, **kw):
            tk.Label(parent, text=txt, bg=COLOR_PANEL, fg=COLOR_TEXT,
                     font=("Verdana", 9), anchor="w", **kw).grid(
                row=row, column=col, sticky="w", padx=(0, 8), pady=4)

        # Valor da parcela
        lbl(form, "Valor da parcela (R$):", 0, 0)
        self.e_atr_valor = _entry_make(form, width=16, placeholder="0,00")
        _bind_valor_format(self.e_atr_valor)
        self.e_atr_valor.grid(row=0, column=1, sticky="w", pady=4)

        # Data de vencimento
        lbl(form, "Data de vencimento:", 1, 0)
        self.e_atr_venc = _entry_make(form, width=14, placeholder="DD/MM/AAAA")
        _mask_full_date(self.e_atr_venc)
        self.e_atr_venc.grid(row=1, column=1, sticky="w", pady=4)

        # Data de pagamento
        lbl(form, "Data de pagamento / atualização:", 2, 0)
        self.e_atr_pag = _entry_make(form, width=14, placeholder="DD/MM/AAAA")
        _mask_full_date(self.e_atr_pag)
        self.e_atr_pag.grid(row=2, column=1, sticky="w", pady=4)
        tk.Label(form, text="(data em que a parcela foi ou será paga)",
                 bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 7)).grid(row=2, column=2, sticky="w", padx=4)

        # Índice
        lbl(form, "Índice de correção:", 3, 0)
        self.cb_atr_indice = ttk.Combobox(
            form, state="readonly", width=22,
            values=[v["name"] for v in INDICES.values() if v["serie"] != 11])
        self.cb_atr_indice.set(INDICES["IPCA"]["name"])
        self.cb_atr_indice.grid(row=3, column=1, sticky="w", pady=4)

        # Multa %
        lbl(form, "Multa (%):", 4, 0)
        self.e_atr_multa = _entry_make(form, width=8, placeholder="10")
        self.e_atr_multa.grid(row=4, column=1, sticky="w", pady=4)
        tk.Label(form, text="% sobre valor atualizado",
                 bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 7)).grid(row=4, column=2, sticky="w", padx=4)

        # Juros %
        lbl(form, "Juros mensais (%):", 5, 0)
        self.e_atr_juros = _entry_make(form, width=8, placeholder="1")
        self.e_atr_juros.grid(row=5, column=1, sticky="w", pady=4)
        tk.Label(form, text="% a.m. sobre valor atualizado",
                 bg=COLOR_PANEL, fg=COLOR_SUBTLE,
                 font=("Verdana", 7)).grid(row=5, column=2, sticky="w", padx=4)

        # Botões
        btn_bar = tk.Frame(pad, bg=COLOR_PANEL)
        btn_bar.pack(fill="x", pady=(16, 0))
        ttk.Button(btn_bar, text="↺ Limpar",
                   style="BCBSmall.TButton",
                   command=self._atraso_limpar).pack(side="left", padx=(0, 8))
        self.btn_atr_calc = ttk.Button(
            btn_bar, text="▶ Calcular atraso",
            style="BCBPrimary.TButton",
            command=self._atraso_calc)
        self.btn_atr_calc.pack(side="right")

        tk.Frame(pad, bg="#dde3ea", height=1).pack(fill="x", pady=(14, 4))

        # Frame de resultado
        self.atr_result_frame = tk.Frame(pad, bg=COLOR_PANEL)
        self.atr_result_frame.pack(fill="x", pady=(4, 0))

    # ------------------------------------------------------------------ #
    # Exportar PDF — Atraso de Parcela                                    #
    # ------------------------------------------------------------------ #

    def _atraso_limpar(self):
        """Limpa todos os campos da aba Atraso de Parcela."""
        for entry, placeholder in [
            (self.e_atr_valor, "0,00"),
            (self.e_atr_venc,  "DD/MM/AAAA"),
            (self.e_atr_pag,   "DD/MM/AAAA"),
            (self.e_atr_multa, "10"),
            (self.e_atr_juros, "1"),
        ]:
            entry.delete(0, "end")
            entry.insert(0, placeholder)
            entry.configure(fg=COLOR_SUBTLE if hasattr(entry, "_is_placeholder")
                            else COLOR_TEXT)
        self.cb_atr_indice.set(INDICES["IPCA"]["name"])
        for w in self.atr_result_frame.winfo_children():
            w.destroy()

    def _atraso_calc(self):
        """Lê os campos da aba Atraso, valida e dispara o cálculo."""
        try:
            valor = parse_valor_br(_entry_value(self.e_atr_valor))
            if valor is None or valor <= 0:
                raise ValueError()
        except Exception:
            return messagebox.showerror("Validação", "Valor da parcela inválido.")

        data_venc = parse_date_br(_entry_value(self.e_atr_venc))
        if not data_venc:
            return messagebox.showerror("Validação", "Data de vencimento inválida.")
        data_pag = parse_date_br(_entry_value(self.e_atr_pag))
        if not data_pag:
            return messagebox.showerror("Validação", "Data de pagamento/atualização inválida.")

        try:
            multa_str = _entry_value(self.e_atr_multa).replace(",", ".")
            multa_pct = Decimal(multa_str) / Decimal("100")
        except Exception:
            multa_pct = Decimal("0.10")
        try:
            juros_str = _entry_value(self.e_atr_juros).replace(",", ".")
            juros_pct = Decimal(juros_str) / Decimal("100")
        except Exception:
            juros_pct = Decimal("0.01")

        indice_nome = self.cb_atr_indice.get()
        indice_key = next(
            (k for k, v in INDICES.items() if v["name"] == indice_nome), "IPCA")

        config = {
            "data_vencimento":   data_venc,
            "data_pagamento":    data_pag,
            "valor_parcela":     valor,
            "multa_pct":         multa_pct,
            "juros_mensais_pct": juros_pct,
            "indice_key":        indice_key,
        }

        for w in self.atr_result_frame.winfo_children():
            w.destroy()
        self.btn_atr_calc.config(state="disabled")
        loading = tk.Label(self.atr_result_frame,
                           text="⟳  Calculando...", bg=COLOR_PANEL,
                           fg=COLOR_BCB_BLUE, font=("Verdana", 9, "italic"),
                           pady=20)
        loading.pack()
        limpar_cache_api()

        def progress(msg):
            try:
                self.after(0, lambda: loading.config(text="⟳  " + msg))
            except Exception:
                pass

        def worker():
            try:
                res = calcular_atraso_parcela(config, progress_cb=progress)
                self.after(0, lambda: self._render_atraso_result(res))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: (
                    loading.destroy(),
                    messagebox.showerror("Erro no cálculo", err),
                    self.btn_atr_calc.config(state="normal")
                ))

        threading.Thread(target=worker, daemon=True).start()

    def _render_atraso_result(self, res):
        """Exibe o resultado do cálculo de atraso de parcela."""
        self.btn_atr_calc.config(state="normal")
        self.atr_resultado = res

        for w in self.atr_result_frame.winfo_children():
            w.destroy()

        frame = tk.Frame(self.atr_result_frame, bg=COLOR_RESULT_BG,
                         relief="flat", bd=1)
        frame.pack(fill="x", pady=(8, 0))

        def row(label, value, bold=False, color=COLOR_TEXT):
            r = tk.Frame(frame, bg=COLOR_RESULT_BG)
            r.pack(fill="x", padx=12, pady=2)
            tk.Label(r, text=label, bg=COLOR_RESULT_BG, fg=COLOR_SUBTLE,
                     font=("Verdana", 8), anchor="w", width=34).pack(side="left")
            tk.Label(r, text=value, bg=COLOR_RESULT_BG, fg=color,
                     font=("Verdana", 9, "bold" if bold else "normal"),
                     anchor="e").pack(side="right")

        tk.Frame(frame, bg=COLOR_BCB_BLUE, height=3).pack(fill="x")
        tk.Label(frame, text="Resultado — Atraso de Parcela",
                 bg=COLOR_RESULT_BG, fg=COLOR_BCB_BLUE,
                 font=("Verdana", 10, "bold"), pady=8).pack()

        row("Índice utilizado:", res["indice_nome"])
        row("Valor original:", fmt_brl(res["valor_parcela"]))
        row("Fator de correção:", fmt_fator(res["fator"]))
        row("Valor atualizado:", fmt_brl(res["valor_atualizado"]))
        row(f"Multa ({float(res['multa_pct']*100):.0f}%):",
            fmt_brl(res["multa"]))
        row(f"Juros ({fot(res['juros_mensais_pct']*100):.1f}% × "
            f"{float(res['meses_atraso_juros']):.2f} meses):".replace(".", ","),
            fmt_brl(res["juros"]))

        tk.Frame(frame, bg="#cccccc", height=1).pack(fill="x", padx=12, pady=4)

        row("TOTAL A PAGAR:", fmt_brl(res["total"]),
            bold=True, color=COLOR_RESULT_OK)

        btn_row = tk.Frame(frame, bg=COLOR_RESULT_BG)
        btn_row.pack(fill="x", padx=12, pady=(8, 12))
        ttk.Button(btn_row, text="⬇ Exportar PDF",
                   style="BCBSmall.TButton",
                   command=self._atraso_export_pdf).pack(side="right")

    def _atraso_export_pdf(self):
        path = filedialog.asksaveasfilename(
                        title="Salvar PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="atraso_parcela.pdf")
        if not path:
            return
        try:
            exportar_pdf_atraso(path, self.atr_resultado)
            messagebox.showinfo("Exportar", f"PDF salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))


# ===================== Ponto de entrada ===================== #

if __name__ == "__main__":
    app = CalculadoraApp()
    app.mainloop()
